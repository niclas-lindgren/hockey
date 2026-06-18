from datetime import date

import openpyxl

from tournament_scheduler.models import SeasonPlan, Team, Tournament
from tournament_scheduler.cli.rvv_cli import main as rvv_main
from tournament_scheduler.pipeline.manual_adjustment_workflow import ManualAdjustmentWorkflow
from tournament_scheduler.pipeline.stage3_helpers import _plan_to_dict
from tournament_scheduler.pipeline.state import PipelineState, StageName, StageStatus
from tournament_scheduler.season_planner import SeasonPlanner


def _make_plan() -> SeasonPlan:
    t1_teams = [
        Team(club="Jar", label="Jar 1", age_group="U10"),
        Team(club="Skien", label="Skien 1", age_group="U10"),
    ]
    t2_teams = [
        Team(club="Jar", label="Jar 1", age_group="U10"),
        Team(club="Kongsberg", label="Kongsberg 1", age_group="U10"),
    ]

    t1 = Tournament(
        id="pin12345",
        date=date(2027, 1, 16),
        arena="Jarhallen",
        age_group="U10",
        host_club="Jar",
        teams=t1_teams,
        games=SeasonPlanner.generate_round_robin_games(t1_teams, 1),
    )
    t2 = Tournament(
        id="move1234",
        date=date(2027, 1, 23),
        arena="Skien ishall",
        age_group="U10",
        host_club="Skien",
        teams=t2_teams,
        games=SeasonPlanner.generate_round_robin_games(t2_teams, 1),
    )
    plan = SeasonPlan(
        tournaments=[t1, t2],
        start_date=date(2027, 1, 1),
        end_date=date(2027, 3, 31),
        manual_adjustments={
            "locked_dates": ["2027-01-16"],
            "banned_dates": ["2027-01-23"],
            "forced_host_clubs": ["Jar"],
            "excluded_host_clubs": ["Skien"],
            "pinned_tournament_ids": ["pin12345"],
        },
    )
    return plan


def _write_state(tmp_path):
    work_dir = tmp_path / "pipeline"
    state = PipelineState(work_dir)

    input_file = tmp_path / "input.xlsx"
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    settings.append(["start_date", "2027-01-01"])
    settings.append(["end_date", "2027-03-31"])
    ages = wb.create_sheet("Aldersgrupper")
    ages.append(["age_group", "parallel_games", "round_length_minutes"])
    ages.append(["U10", 1, 10])
    teams_sheet = wb.create_sheet("Lag")
    teams_sheet.append(["club", "label", "age_group"])
    teams_sheet.append(["Jar", "Jar 1", "U10"])
    teams_sheet.append(["Skien", "Skien 1", "U10"])
    teams_sheet.append(["Kongsberg", "Kongsberg 1", "U10"])
    wb.save(input_file)
    state.write_stage(
        StageName.CONFIG,
        {
            "input_path": str(input_file),
            "teams": [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
                {"club": "Skien", "label": "Skien 1", "age_group": "U10"},
                {"club": "Kongsberg", "label": "Kongsberg 1", "age_group": "U10"},
            ],
            "round_length_minutes": {"U10": 10},
        },
        status=StageStatus.DONE,
    )

    plan = _make_plan()
    state.write_stage(
        StageName.PLANNING,
        {"plan": _plan_to_dict(plan), "rules_report": []},
        status=StageStatus.DONE,
    )
    return state, plan


def _make_plan_with_game_count_spread() -> SeasonPlan:
    """Return a plan where Jar 1 plays 4 games and Skien 1 plays 1 game (spread=3)."""
    jar = Team(club="Jar", label="Jar 1", age_group="U10")
    skien = Team(club="Skien", label="Skien 1", age_group="U10")
    kongsberg = Team(club="Kongsberg", label="Kongsberg 1", age_group="U10")

    # t_base: Jar vs Skien (1 game each)
    t_base = Tournament(
        id="spread_base",
        date=date(2027, 1, 16),
        arena="Jarhallen",
        age_group="U10",
        host_club="Jar",
        teams=[jar, skien],
        games=SeasonPlanner.generate_round_robin_games([jar, skien], 1),
    )
    # t2..t4: Jar vs Kongsberg (Jar gets 3 more games, Skien gets none)
    extra_tournaments = []
    for i, d in enumerate(
        [date(2027, 2, 6), date(2027, 2, 20), date(2027, 3, 6)],
        start=2,
    ):
        extra_tournaments.append(
            Tournament(
                id=f"spread_extra{i}",
                date=d,
                arena="Jarhallen",
                age_group="U10",
                host_club="Jar",
                teams=[jar, kongsberg],
                games=SeasonPlanner.generate_round_robin_games([jar, kongsberg], 1),
            )
        )

    return SeasonPlan(
        tournaments=[t_base, *extra_tournaments],
        start_date=date(2027, 1, 1),
        end_date=date(2027, 3, 31),
        manual_adjustments={},
    )


def _write_state_with_spread(tmp_path):
    """Write pipeline state containing a plan with a game-count spread violation."""
    work_dir = tmp_path / "pipeline_spread"
    state = PipelineState(work_dir)

    input_file = tmp_path / "input_spread.xlsx"
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    settings.append(["start_date", "2027-01-01"])
    settings.append(["end_date", "2027-03-31"])
    ages = wb.create_sheet("Aldersgrupper")
    ages.append(["age_group", "parallel_games", "round_length_minutes"])
    ages.append(["U10", 1, 10])
    teams_sheet = wb.create_sheet("Lag")
    teams_sheet.append(["club", "label", "age_group"])
    teams_sheet.append(["Jar", "Jar 1", "U10"])
    teams_sheet.append(["Skien", "Skien 1", "U10"])
    teams_sheet.append(["Kongsberg", "Kongsberg 1", "U10"])
    wb.save(input_file)
    state.write_stage(
        StageName.CONFIG,
        {
            "input_path": str(input_file),
            "teams": [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
                {"club": "Skien", "label": "Skien 1", "age_group": "U10"},
                {"club": "Kongsberg", "label": "Kongsberg 1", "age_group": "U10"},
            ],
            "round_length_minutes": {"U10": 10},
        },
        status=StageStatus.DONE,
    )

    plan = _make_plan_with_game_count_spread()
    state.write_stage(
        StageName.PLANNING,
        {"plan": _plan_to_dict(plan), "rules_report": []},
        status=StageStatus.DONE,
    )
    return state, plan


def test_post_patch_warnings_is_list(tmp_path):
    """post_patch_warnings must always be a list (never None) after apply()."""
    state, plan = _write_state(tmp_path)
    workflow = ManualAdjustmentWorkflow(state)

    result = workflow.apply(plan)

    assert result.success is True
    assert isinstance(result.post_patch_warnings, list)


def test_post_patch_warnings_reports_game_count_spread_violation(tmp_path):
    """A plan where Jar 1 plays 4 games and Skien 1 plays 1 game (spread=3 > default 2)
    must include a game-count spread warning string."""
    state, plan = _write_state_with_spread(tmp_path)
    workflow = ManualAdjustmentWorkflow(state)

    result = workflow.apply(plan)

    assert result.success is True
    assert result.post_patch_warnings, "Expected at least one post-patch warning for game count spread violation"
    warning_text = " ".join(result.post_patch_warnings)
    # The game-count spread warning contains the Norwegian word for spread.
    assert "spredning" in warning_text


def test_post_patch_warnings_deduplicated(tmp_path):
    """Warnings must be deduplicated — no identical string appears twice."""
    state, plan = _write_state_with_spread(tmp_path)
    workflow = ManualAdjustmentWorkflow(state)

    result = workflow.apply(plan)

    assert result.success is True
    assert len(result.post_patch_warnings) == len(set(result.post_patch_warnings)), (
        "Duplicate warning strings found in post_patch_warnings"
    )


def test_manual_adjustment_workflow_moves_banned_dates_and_updates_host(tmp_path):
    state, plan = _write_state(tmp_path)
    workflow = ManualAdjustmentWorkflow(state)

    result = workflow.apply(plan)

    assert result.success is True
    assert result.operation == "manual_adjustment"
    assert len(result.changes["moved"]) == 1
    assert len(result.changes["host_changes"]) == 1
    assert plan.tournaments[0].date == date(2027, 1, 16)
    assert plan.tournaments[0].host_club == "Jar"
    assert plan.tournaments[1].date != date(2027, 1, 23)
    assert plan.tournaments[1].host_club == "Jar"
    assert plan.tournaments[1].arena == "Jarhallen"
    assert plan.fairness_gate["status"] in {"pass", "warn", "fail"}
    assert result.conflicts == []
    assert plan.manual_adjustments["pinned_tournament_ids"] == ["pin12345"]


def test_adjust_cli_runs_end_to_end(tmp_path):
    state, _plan = _write_state(tmp_path)
    export_dir = tmp_path / "export"

    code = rvv_main(
        [
            "adjust",
            "--work-dir",
            str(state.work_dir),
            "--export-dir",
            str(export_dir),
            "--lock-date",
            "2027-01-16",
            "--ban-date",
            "2027-01-23",
            "--pin-tournament",
            "pin12345",
            "--force-host-club",
            "Jar",
            "--exclude-host-club",
            "Skien",
        ]
    )

    assert code == 0
    assert (export_dir / "season_plan.xlsx").exists()
    assert (export_dir / "season_plan.html").exists()
    assert (export_dir / "season_plan_report.html").exists()
