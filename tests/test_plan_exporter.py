"""Tests for SeasonPlanExporter (Excel export of season plans)."""

from datetime import date

import openpyxl
import pytest

from tournament_scheduler.models import Game, SeasonPlan, Team, Tournament
from tournament_scheduler.excel.plan_exporter import SeasonPlanExporter, _NORWEGIAN_WEEKDAYS


def _round_robin_games(teams):
    """Generate games with round_number set so tests verify the correct column values."""
    from tournament_scheduler.season_planner import SeasonPlanner
    return SeasonPlanner.generate_round_robin_games(teams, parallel_games=2)


@pytest.fixture
def sample_plan():
    teams_u10 = [Team(club=f"Club{i}", label=f"U10-{i}", age_group="U10") for i in range(5)]
    teams_ju11 = [Team(club=f"Club{i}", label=f"JU11-{i}", age_group="JU11") for i in range(4)]

    t1 = Tournament(
        date=date(2026, 10, 10),
        arena="Jarhallen",
        age_group="U10",
        teams=teams_u10,
        games=_round_robin_games(teams_u10),
        host_club="Jar",
    )
    t2 = Tournament(
        date=date(2026, 11, 7),
        arena="Bærum ishall",
        age_group="JU11",
        teams=teams_ju11,
        games=_round_robin_games(teams_ju11),
        host_club="Jutul",
    )
    return SeasonPlan(
        tournaments=[t1, t2],
        start_date=date(2026, 10, 1),
        end_date=date(2027, 4, 30),
        diversity_score=0.5,
        arena_counts={"Jarhallen": 1, "Bærum ishall": 1},
    )


class TestSeasonPlanExporter:
    """Test suite for SeasonPlanExporter.export — round-trips through openpyxl."""

    def test_export_writes_a_readable_workbook(self, sample_plan, tmp_path):
        output_path = tmp_path / "season_plan.xlsx"
        SeasonPlanExporter().export(sample_plan, str(output_path))

        assert output_path.exists()

        workbook = openpyxl.load_workbook(str(output_path))
        old_adjustment_sheet = "Fairness" + "justeringer"
        assert "Sesongoversikt" in workbook.sheetnames
        assert "Rettferdighetsjusteringer" in workbook.sheetnames
        assert old_adjustment_sheet not in workbook.sheetnames
        # One sheet per tournament, plus the overview sheet, plus the
        # fairness overview, plus one summary sheet per distinct club.
        distinct_clubs = {team.club for tournament in sample_plan.tournaments for team in tournament.teams}
        assert len(workbook.sheetnames) == 2 + len(sample_plan.tournaments) + len(distinct_clubs)

    def test_overview_rows_match_plan_tournaments(self, sample_plan, tmp_path):
        output_path = tmp_path / "season_plan.xlsx"
        SeasonPlanExporter().export(sample_plan, str(output_path))

        workbook = openpyxl.load_workbook(str(output_path))
        old_adjustment_sheet = "Fairness" + "justeringer"
        assert "Rettferdighetsjusteringer" in workbook.sheetnames
        assert old_adjustment_sheet not in workbook.sheetnames
        fairness = workbook["Rettferdighetsjusteringer"]
        fairness_rows = list(fairness.iter_rows(values_only=True))
        assert fairness_rows[0][0] == "Rettferdighetsjusteringer per lag"
        assert "fairness" not in str(fairness_rows[1][0]).lower()
        assert fairness_rows[3][0] == "Lag"
        overview = workbook["Sesongoversikt"]
        rows = list(overview.iter_rows(values_only=True))

        header, *data_rows = rows
        assert header == ("Dato", "Ukedag", "Aldersgruppe", "Arena", "Vertsklubb", "Lag", "Lengst anslått reise", "Starttid", "Sluttid")
        assert len(data_rows) == len(sample_plan.tournaments)

        for tournament, row in zip(sample_plan.tournaments, data_rows):
            expected_date = tournament.date.strftime("%d.%m.%Y")
            expected_weekday = _NORWEGIAN_WEEKDAYS[tournament.date.weekday()]
            expected_teams = ", ".join(team.label for team in tournament.teams)

            assert row[0] == expected_date
            assert row[1] == expected_weekday
            assert row[2] == tournament.age_group
            assert row[3] == tournament.arena
            assert row[4] == tournament.host_club
            assert row[5] == expected_teams
            # row[6] is "Lengst anslått reise" — varies by arena/teams
            # row[6] is "Lengst anslått reise" — may be None (empty cell) or a string
            assert row[6] is None or "~" in str(row[6])
            # row[7] is "Starttid" — empty string when start_time unset, becomes None in Excel
            assert row[7] == (tournament.start_time or None)
            # row[8] is "Sluttid" — empty unless round_length_for_age_group is passed to export()
            assert row[8] is None

    def test_per_tournament_sheets_match_game_lists(self, sample_plan, tmp_path):
        output_path = tmp_path / "season_plan.xlsx"
        SeasonPlanExporter().export(sample_plan, str(output_path))

        workbook = openpyxl.load_workbook(str(output_path))

        for tournament in sample_plan.tournaments:
            # Find the sheet whose title contains this tournament's date prefix.
            date_prefix = tournament.date.strftime("%d.%m")
            matching_sheets = [
                workbook[name] for name in workbook.sheetnames
                if name.startswith(date_prefix) and tournament.age_group in name
            ]
            assert len(matching_sheets) == 1, f"expected exactly one sheet for {tournament.date} {tournament.age_group}"
            sheet = matching_sheets[0]

            rows = list(sheet.iter_rows(values_only=True))
            # Find the header row for the games table.
            header_index = next(
                i for i, row in enumerate(rows)
                if row[:4] == ("Runde", "Hjemmelag", "Bortelag", "Parallellbane")
            )
            game_rows = rows[header_index + 1:header_index + 1 + len(tournament.games)]

            assert len(game_rows) == len(tournament.games)
            for game, row in zip(tournament.games, game_rows):
                assert row[0] == game.round_number
                assert row[1] == game.home.label
                assert row[2] == game.away.label
                assert row[3] == game.parallel_slot + 1  # exporter displays 1-based slots

    def test_sheet_titles_are_unique_and_within_excel_limits(self, sample_plan, tmp_path):
        output_path = tmp_path / "season_plan.xlsx"
        SeasonPlanExporter().export(sample_plan, str(output_path))

        workbook = openpyxl.load_workbook(str(output_path))
        titles = workbook.sheetnames

        assert len(titles) == len(set(titles)), "sheet titles must be unique"
        for title in titles:
            assert len(title) <= 31

    def test_club_sheet_titles_are_unique_and_sanitized_on_collision(self, tmp_path):
        # Two clubs whose names collide once sanitized/truncated to the same
        # base, forcing the numeric-suffix collision path to engage.
        long_name = "Sandefjord Penguins Ishockeyklubb Senior"
        teams_a = [Team(club=long_name, label="A1", age_group="U10")]
        teams_b = [Team(club=f"{long_name} (2)", label="B1", age_group="U10")]

        tournament = Tournament(
            date=date(2026, 9, 12),
            arena="Testhallen",
            age_group="U10",
            teams=teams_a + teams_b,
            games=_round_robin_games(teams_a + teams_b),
            host_club=long_name,
        )
        plan = SeasonPlan(tournaments=[tournament])

        output_path = tmp_path / "club_collision.xlsx"
        SeasonPlanExporter().export(plan, str(output_path))

        workbook = openpyxl.load_workbook(str(output_path))
        titles = workbook.sheetnames

        assert len(titles) == len(set(titles)), "sheet titles must be unique"
        for title in titles:
            assert len(title) <= 31

        club_titles = [title for title in titles if title.startswith("Klubb ")]
        assert len(club_titles) == 2, f"expected one summary sheet per club, got {club_titles}"

    def test_club_summary_sheet_lists_team_opponents_and_arena(self, sample_plan, tmp_path):
        output_path = tmp_path / "season_plan.xlsx"
        SeasonPlanExporter().export(sample_plan, str(output_path))

        workbook = openpyxl.load_workbook(str(output_path))
        club_sheets = [name for name in workbook.sheetnames if name.startswith("Klubb ")]

        # sample_plan teams use clubs "Club0".."Club4" (U10) and "Club0".."Club3" (JU11)
        assert sorted(club_sheets) == [f"Klubb Club{i}" for i in range(5)]

        sheet = workbook["Klubb Club0"]
        rows = list(sheet.iter_rows(values_only=True))

        header_index = next(
            i for i, row in enumerate(rows)
            if row[:6] == ("Lag", "Aldersgruppe", "Dato", "Ukedag", "Motstander(e)", "Vertsarena")
        )
        data_rows = [row for row in rows[header_index + 1:] if row[0] is not None]

        # Club0 has one team in each tournament (U10-0 and JU11-0)
        assert len(data_rows) == 2
        labels = {row[0] for row in data_rows}
        assert labels == {"U10-0", "JU11-0"}

        for row in data_rows:
            label, age_group, formatted_date, weekday, opponents, arena = row
            tournament = next(t for t in sample_plan.tournaments if t.age_group == age_group)
            assert formatted_date == tournament.date.strftime("%d.%m.%Y")
            assert weekday == _NORWEGIAN_WEEKDAYS[tournament.date.weekday()]
            assert arena == tournament.arena
            team = next(t for t in tournament.teams if t.label == label)
            expected_opponents = {
                (game.away if game.home is team else game.home).label
                for game in tournament.games
                if game.home is team or game.away is team
            }
            assert set(opponents.split(", ")) == expected_opponents


class TestVarnerArenaHomeTeamRegression:
    """Regression: Frisk Asker must be the home team in Excel output for Varner Arena tournaments."""

    def test_hjemmelag_column_shows_frisk_asker_for_varner_arena(self, tmp_path):
        """When a tournament is hosted at Varner Arena, all game rows in the
        per-tournament Excel sheet must show Frisk Asker in the Hjemmelag column."""
        from tournament_scheduler.season_planner import SeasonPlanner

        frisk = Team(club="Frisk Asker", label="Frisk Asker A", age_group="U10")
        visitors = [
            Team(club="Kongsberg", label="Kongsberg A", age_group="U10"),
            Team(club="Skien", label="Skien A", age_group="U10"),
            Team(club="Ringerike", label="Ringerike A", age_group="U10"),
        ]
        # Host-first ordering: Frisk Asker comes first
        teams = [frisk] + visitors
        games = SeasonPlanner.generate_round_robin_games(teams, parallel_games=2)

        tournament = Tournament(
            date=date(2026, 11, 14),
            arena="Varner Arena",
            age_group="U10",
            teams=teams,
            games=games,
            host_club="Frisk Asker",
        )
        plan = SeasonPlan(
            tournaments=[tournament],
            start_date=date(2026, 10, 1),
            end_date=date(2027, 4, 30),
            diversity_score=0.5,
            arena_counts={"Varner Arena": 1},
        )

        output_path = tmp_path / "varner_arena_plan.xlsx"
        SeasonPlanExporter().export(plan, str(output_path))

        workbook = openpyxl.load_workbook(str(output_path))

        # Find the per-tournament sheet for this tournament.
        # Sheet names use short date format "DD.MM" not the full date.
        date_prefix = tournament.date.strftime("%d.%m")
        matching_sheets = [
            name for name in workbook.sheetnames
            if name.startswith(date_prefix) and tournament.age_group in name
        ]
        assert len(matching_sheets) == 1, (
            f"Expected exactly one sheet for {tournament.date} {tournament.age_group}; "
            f"got {matching_sheets}"
        )
        sheet = workbook[matching_sheets[0]]
        rows = list(sheet.iter_rows(values_only=True))

        header_index = next(
            i for i, row in enumerate(rows)
            if row[:2] == ("Runde", "Hjemmelag")
        )
        game_rows = rows[header_index + 1:header_index + 1 + len(games)]

        assert len(game_rows) == len(games), "row count mismatch"
        # Frisk Asker must be home in every game they play (never away).
        frisk_rows = [
            (idx, row) for idx, row in enumerate(game_rows)
            if row[1] == frisk.label or row[2] == frisk.label
        ]
        assert len(frisk_rows) > 0, "Frisk Asker should appear in at least one game row"
        for idx, row in frisk_rows:
            home_label = row[1]
            assert home_label == frisk.label, (
                f"Game row {idx}: Frisk Asker must be Hjemmelag, got {home_label!r}"
            )


class TestDuplicateLabelDisambiguation:
    """Tests that teams with duplicate labels are correctly counted using team_key()."""

    def _make_plan_with_duplicate_labels(self):
        """Return a SeasonPlan where two teams share the same label but differ in club/age_group."""
        from tournament_scheduler.season_planner import SeasonPlanner

        # Two teams labelled "A-lag" in different clubs/age groups
        team_a1 = Team(club="Kongsberg", label="A-lag", age_group="U10")
        team_a2 = Team(club="Ringerike", label="A-lag", age_group="U10")
        team_b = Team(club="Skien", label="Skien U10", age_group="U10")
        team_c = Team(club="Jar", label="Jar U10", age_group="U10")

        teams = [team_a1, team_a2, team_b, team_c]
        games = SeasonPlanner.generate_round_robin_games(teams, parallel_games=2)

        tournament = Tournament(
            date=date(2026, 10, 10),
            arena="Kongsberg ishall",
            age_group="U10",
            teams=teams,
            games=games,
            host_club="Kongsberg",
        )
        return SeasonPlan(
            tournaments=[tournament],
            start_date=date(2026, 10, 1),
            end_date=date(2027, 4, 30),
            diversity_score=0.9,
            arena_counts={"Kongsberg ishall": 1},
        )

    def test_compute_team_game_counts_disambiguates_duplicate_labels(self):
        """compute_team_game_counts should produce distinct keys for teams sharing a label."""
        from tournament_scheduler.html.data_computation import compute_team_game_counts

        plan = self._make_plan_with_duplicate_labels()
        counts = compute_team_game_counts(plan)

        # There are 4 distinct teams: two named "A-lag" (different clubs) plus two unique ones.
        assert len(counts) == 4, (
            f"Expected 4 distinct team keys, got {len(counts)}: {list(counts.keys())}"
        )
        # All counts should be > 0 (each team plays some games).
        assert all(v > 0 for v in counts.values()), (
            f"All teams should have played at least one game; counts={counts}"
        )
        # The raw duplicate label should not appear alone as a key.
        assert "A-lag" not in counts, (
            "Raw duplicate label 'A-lag' should not appear as a key; keys should be disambiguated"
        )

    def test_review_summary_game_spread_uses_disambiguated_counts(self):
        """analyze_review_summary should show non-zero game counts for all teams."""
        from tournament_scheduler.html.data_computation import compute_team_game_counts
        from tournament_scheduler.html.renderers.review import analyze_review_summary

        plan = self._make_plan_with_duplicate_labels()
        # Attach team_game_counts to the plan as the HTML exporter does.
        plan.team_game_counts = compute_team_game_counts(plan)  # type: ignore[attr-defined]

        result = analyze_review_summary(plan)

        # The age_spread_summaries finding should not contain "0" (i.e. no missed lookups).
        spread_finding = next(
            (f for f in result["findings"] if "Avvik" in str(f.get("label", ""))),
            None,
        )
        assert spread_finding is not None, "Expected an 'Avvik' finding"
        # The finding text must not read "0 kamper" / reference zero counts if there are games.
        text = str(spread_finding.get("text", ""))
        # If all lookups returned 0, the spread would be 0 and the text would say "0 kamper".
        assert "har 0 kamper" not in text, (
            f"All lookups returned 0 — team_key disambiguation likely failed. Finding: {text}"
        )

    def test_opinionated_judgment_spread_uses_disambiguated_counts(self):
        """analyze_opinionated_judgment spread should reflect actual game counts, not all zeros."""
        from tournament_scheduler.html.data_computation import compute_team_game_counts
        from tournament_scheduler.html.renderers.judgment import analyze_opinionated_judgment

        plan = self._make_plan_with_duplicate_labels()
        team_game_counts = compute_team_game_counts(plan)

        result = analyze_opinionated_judgment(
            plan,
            team_game_counts=team_game_counts,
            club_stats={},
            team_travel={},
        )

        # The "Belastning" card should mention actual game counts, not "0 kamper".
        load_card = next(
            (text for label, text in result["cards"] if label == "Belastning"),
            None,
        )
        assert load_card is not None, "Expected a 'Belastning' card"
        assert "har 0 kamper" not in load_card, (
            f"All lookups returned 0 — team_key disambiguation likely failed. Card: {load_card}"
        )
