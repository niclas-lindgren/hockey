from __future__ import annotations

import csv
import json
from pathlib import Path

import openpyxl
import pytest

from tournament_scheduler.registrations import (
    RegistrationImportError,
    export_registrations,
    validate_registrations,
)


def _write_input_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    settings.append(["start_date", "2026-10-01"])
    settings.append(["end_date", "2027-04-30"])

    age_groups = wb.create_sheet("Aldersgrupper")
    age_groups.append(["age_group", "parallel_games"])
    age_groups.append(["U10", 3])
    age_groups.append(["U11", 2])

    teams = wb.create_sheet("Lag")
    teams.append(["club", "label", "age_group"])
    teams.append(["Kongsberg", "Kongsberg 1", "U10"])
    teams.append(["Jar", "Jar 1", "U10"])
    teams.append(["Skien", "Skien 1", "U11"])

    sources = wb.create_sheet("Kilder")
    sources.append(["name", "type", "url"])
    sources.append(["Secret source", "ical", "https://calendar.example.test/private.ics"])
    wb.save(path)


def _write_registration_csv(path: Path, rows: list[dict[str, str]]) -> None:
    headers = [
        "SharePoint ID",
        "Klubb",
        "Lag",
        "Aldergruppe",
        "Status",
        "Kontakt",
        "Kommentar",
        "Modified",
    ]
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def _default_rows() -> list[dict[str, str]]:
    return [
        {
            "SharePoint ID": "101",
            "Klubb": "Kongsberg",
            "Lag": "Kongsberg 1",
            "Aldergruppe": "U10",
            "Status": "Godkjent",
            "Kontakt": "parent@example.test",
            "Kommentar": "private note",
            "Modified": "2026-07-01T12:00:00Z",
        },
        {
            "SharePoint ID": "102",
            "Klubb": "Jar",
            "Lag": "Jar 2",
            "Aldergruppe": "U10",
            "Status": "approved",
            "Kontakt": "coach@example.test",
            "Kommentar": "rename from Jar 1",
            "Modified": "2026-07-02T12:00:00Z",
        },
        {
            "SharePoint ID": "103",
            "Klubb": "Skien",
            "Lag": "Skien 1",
            "Aldergruppe": "U11",
            "Status": "Trukket",
            "Kontakt": "withdrawn@example.test",
            "Kommentar": "withdrawn",
            "Modified": "2026-07-03T12:00:00Z",
        },
    ]


class TestRegistrationImport:
    def test_validate_csv_reports_diff_without_writing(self, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        registrations = tmp_path / "sharepoint.csv"
        output_path = tmp_path / "input.updated.xlsx"
        _write_input_workbook(input_path)
        _write_registration_csv(registrations, _default_rows())

        result = export_registrations(registrations, input_path=input_path, output_path=output_path, dry_run=True)

        assert result.dry_run is True
        assert result.written is False
        assert not output_path.exists()
        assert result.summary_counts == {
            "active": 2,
            "added": 1,
            "removed": 2,
            "changed": 1,
            "unchanged": 1,
            "rejected": 1,
        }
        assert result.diff.added == [{"club": "Jar", "label": "Jar 2", "age_group": "U10"}]
        assert {row["label"] for row in result.diff.removed} == {"Jar 1", "Skien 1"}
        assert result.diff.rejected[0].sharepoint_id == "103"
        assert result.source_fingerprint.startswith("sha256:")

    def test_non_dry_run_replaces_only_lag_sheet_and_writes_audit(self, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        registrations = tmp_path / "sharepoint.csv"
        output_path = tmp_path / "input.updated.xlsx"
        _write_input_workbook(input_path)
        _write_registration_csv(registrations, _default_rows())

        result = export_registrations(registrations, input_path=input_path, output_path=output_path)

        assert result.written is True
        assert output_path.exists()
        audit_path = output_path.with_suffix(".registrations.audit.json")
        assert audit_path.exists()

        wb = openpyxl.load_workbook(output_path, data_only=True)
        assert wb.sheetnames == ["Innstillinger", "Aldersgrupper", "Lag", "Kilder"]
        assert wb["Kilder"]["C2"].value == "https://calendar.example.test/private.ics"
        lag_rows = list(wb["Lag"].iter_rows(values_only=True))
        assert lag_rows == [
            ("club", "label", "age_group"),
            ("Jar", "Jar 2", "U10"),
            ("Kongsberg", "Kongsberg 1", "U10"),
        ]
        serialized_workbook_values = repr(lag_rows)
        assert "parent@example.test" not in serialized_workbook_values
        assert "private note" not in serialized_workbook_values

        audit = json.loads(audit_path.read_text(encoding="utf-8"))
        assert audit["source_fingerprint"] == result.source_fingerprint
        assert audit["included_sharepoint_ids"] == ["102", "101"]
        assert audit["active_team_count"] == 2

    def test_validate_accepts_xlsx_sharepoint_export(self, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        registrations = tmp_path / "sharepoint.xlsx"
        _write_input_workbook(input_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["SharePoint ID", "Klubb", "Lag", "Aldergruppe", "Status"])
        ws.append(["201", "Kongsberg", "Kongsberg 1", "U10", "current"])
        ws.append(["202", "Jar", "Jar 1", "U10", "accepted"])
        ws.append(["203", "Skien", "Skien 1", "U11", "active"])
        wb.save(registrations)

        result = validate_registrations(registrations, input_path=input_path)

        assert result.summary_counts["active"] == 3
        assert result.summary_counts["unchanged"] == 3

    def test_duplicate_sharepoint_ids_are_blocked(self, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        registrations = tmp_path / "sharepoint.csv"
        _write_input_workbook(input_path)
        rows = _default_rows()[:2]
        rows[1]["SharePoint ID"] = "101"
        _write_registration_csv(registrations, rows)

        with pytest.raises(RegistrationImportError, match="duplikat SharePoint-ID"):
            validate_registrations(registrations, input_path=input_path)

    def test_duplicate_team_identity_with_different_ids_is_blocked(self, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        registrations = tmp_path / "sharepoint.csv"
        _write_input_workbook(input_path)
        rows = _default_rows()[:2]
        rows[1]["Klubb"] = "Kongsberg"
        rows[1]["Lag"] = "Kongsberg 1"
        _write_registration_csv(registrations, rows)

        with pytest.raises(RegistrationImportError, match="finnes med flere SharePoint-ID"):
            validate_registrations(registrations, input_path=input_path)

    def test_unknown_status_and_unknown_controlled_values_are_actionable(self, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        registrations = tmp_path / "sharepoint.csv"
        _write_input_workbook(input_path)
        rows = _default_rows()[:1]
        rows[0]["Klubb"] = "Unknown"
        rows[0]["Aldergruppe"] = "U99"
        rows[0]["Status"] = "Maybe"
        _write_registration_csv(registrations, rows)

        with pytest.raises(RegistrationImportError) as exc_info:
            validate_registrations(registrations, input_path=input_path)

        message = str(exc_info.value)
        assert "ukjent status 'Maybe'" in message

        rows[0]["Status"] = "Godkjent"
        _write_registration_csv(registrations, rows)
        with pytest.raises(RegistrationImportError) as exc_info:
            validate_registrations(registrations, input_path=input_path)

        message = str(exc_info.value)
        assert "ukjent klubb 'Unknown'" in message
        assert "ukjent aldersgruppe 'U99'" in message

    def test_missing_required_columns_are_reported(self, tmp_path: Path) -> None:
        input_path = tmp_path / "input.xlsx"
        registrations = tmp_path / "sharepoint.csv"
        _write_input_workbook(input_path)
        with open(registrations, "w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=["Klubb", "Lag"])
            writer.writeheader()
            writer.writerow({"Klubb": "Kongsberg", "Lag": "Kongsberg 1"})

        with pytest.raises(RegistrationImportError, match="mangler påkrevde kolonner"):
            validate_registrations(registrations, input_path=input_path)
