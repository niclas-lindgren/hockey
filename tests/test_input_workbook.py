"""Tests for the public-sheet whitelist in tournament_scheduler.pipeline.input_workbook."""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from tournament_scheduler.pipeline.input_workbook import (
    PUBLIC_SHEET_WHITELIST,
    WorkbookInputError,
    assert_public_sheet,
    read_public_teams,
)


def _write_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    settings.append(["start_date", "2025-09-01"])

    age_groups = wb.create_sheet("Aldersgrupper")
    age_groups.append(["age_group", "parallel_games"])
    age_groups.append(["U10", 3])

    teams = wb.create_sheet("Lag")
    teams.append(["club", "label", "age_group"])
    teams.append(["Kongsberg", "Kongsberg U10A", "U10"])
    teams.append(["Skien", "Skien U10A", "U10"])
    teams.append(["Jar", "Jar U12A", "U12"])

    sources = wb.create_sheet("Kilder")
    sources.append(["name", "type", "url"])
    sources.append(["Kongsberg kalender", "ical", "https://secret.example.com/feed.ics"])

    wb.save(path)


class TestPublicSheetWhitelist:
    def test_whitelist_only_allows_lag(self):
        assert PUBLIC_SHEET_WHITELIST == ("Lag",)
        assert_public_sheet("Lag")  # does not raise

    @pytest.mark.parametrize("sheet_name", ["Innstillinger", "Aldersgrupper", "Kilder", "Datopreferanser"])
    def test_internal_sheets_are_rejected(self, sheet_name):
        with pytest.raises(WorkbookInputError):
            assert_public_sheet(sheet_name)

    def test_read_public_teams_returns_only_lag_rows(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _write_workbook(path)

        teams = read_public_teams(path)

        assert len(teams) == 3
        assert {t["club"] for t in teams} == {"Kongsberg", "Skien", "Jar"}
        assert {t["age_group"] for t in teams} == {"U10", "U12"}

    def test_read_public_teams_never_exposes_internal_sheet_content(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _write_workbook(path)

        teams = read_public_teams(path)

        serialized = str(teams)
        assert "secret.example.com" not in serialized
        assert "parallel_games" not in serialized

    def test_read_public_teams_missing_lag_sheet_returns_empty(self, tmp_path):
        path = tmp_path / "input.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Innstillinger"
        wb.save(path)

        assert read_public_teams(path) == []
