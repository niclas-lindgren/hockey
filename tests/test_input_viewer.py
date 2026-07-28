"""Tests for the public input.html viewer (issue #32)."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from tournament_scheduler.pipeline.input_viewer import generate_html


def _write_input_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    settings.append(["start_date", "2025-09-01"])

    age_groups = wb.create_sheet("Aldersgrupper")
    age_groups.append(["age_group", "parallel_games"])
    age_groups.append(["U10", 3])

    teams = wb.create_sheet("Lag")
    teams.append(["club", "label", "age_group"])
    teams.append(["Kongsberg", "Kongsberg U10A", "U10"])
    teams.append(["Skien", "Skien U10A", "U10"])
    teams.append(["Jar", "Jar U12A", "U12"])

    sources = wb.create_sheet("Kilder")
    sources.append(["name", "type", "url"])
    sources.append(["Hemmelig kilde", "ical", "https://secret.example.com/feed.ics"])

    wb.save(path)


class TestInputViewer:
    def test_generate_html_lists_teams_grouped_by_age_group(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(input_path)

        out_path = Path(
            generate_html(input_path=str(input_path), export_dir=str(tmp_path / "export"))
        )
        html = out_path.read_text(encoding="utf-8")

        assert out_path.name == "input.html"
        assert "Påmeldte lag" in html
        assert "Kongsberg U10A" in html
        assert "Skien U10A" in html
        assert "Jar U12A" in html
        assert ">U10<" in html
        assert ">U12<" in html

    def test_generate_html_shows_club_and_team_totals(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(input_path)

        html = Path(
            generate_html(input_path=str(input_path), export_dir=str(tmp_path / "export"))
        ).read_text(encoding="utf-8")

        assert "<strong>3</strong>&nbsp;klubber" in html
        assert "<strong>3</strong>&nbsp;lag" in html
        assert "<strong>2</strong>&nbsp;aldersgrupper" in html

    def test_generate_html_includes_search_and_filter_controls(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(input_path)

        html = Path(
            generate_html(input_path=str(input_path), export_dir=str(tmp_path / "export"))
        ).read_text(encoding="utf-8")

        assert 'id="filterAge"' in html
        assert 'id="filterClub"' in html
        assert 'id="filterSearch"' in html

    def test_generate_html_never_exposes_internal_sheets_or_workbook(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(input_path)

        html = Path(
            generate_html(input_path=str(input_path), export_dir=str(tmp_path / "export"))
        ).read_text(encoding="utf-8")

        assert "secret.example.com" not in html
        assert "Hemmelig kilde" not in html
        assert "parallel_games" not in html
        assert "input.xlsx" not in html

    def test_generate_html_handles_empty_lag_sheet(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Innstillinger"
        teams = wb.create_sheet("Lag")
        teams.append(["club", "label", "age_group"])
        wb.save(input_path)

        html = Path(
            generate_html(input_path=str(input_path), export_dir=str(tmp_path / "export"))
        ).read_text(encoding="utf-8")

        assert "Ingen lag er registrert" in html

    def test_generate_html_links_to_calendars_only_when_present(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(input_path)
        export_dir = tmp_path / "export"
        export_dir.mkdir()

        html_without = Path(
            generate_html(input_path=str(input_path), export_dir=str(export_dir))
        ).read_text(encoding="utf-8")
        assert "Skrapede kalendere" not in html_without

        (export_dir / "calendars.html").write_text("<html></html>", encoding="utf-8")
        html_with = Path(
            generate_html(
                input_path=str(input_path),
                export_dir=str(export_dir),
                calendars_path=str(export_dir / "calendars.html"),
            )
        ).read_text(encoding="utf-8")
        assert "Skrapede kalendere" in html_with
