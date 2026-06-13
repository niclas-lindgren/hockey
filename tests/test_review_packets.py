from datetime import date
import json
from pathlib import Path

import openpyxl

from tournament_scheduler.cli.rvv_cli import main as rvv_main
from tournament_scheduler.models import Game, SeasonPlan, Team, Tournament
from tournament_scheduler.pipeline.stage3_helpers import _plan_to_dict
from tournament_scheduler.pipeline.stage4_helpers import _dict_to_plan
from tournament_scheduler.pipeline.state import PipelineState, StageName, StageStatus
from tournament_scheduler.pipeline.stage4_export import run
from tournament_scheduler.season_planner import SeasonPlanner


def _make_plan() -> SeasonPlan:
    jar = Team(club="Jar", label="Jar U10A", age_group="U10")
    kongsberg = Team(club="Kongsberg", label="Kongsberg U10A", age_group="U10")
    skien = Team(club="Skien", label="Skien JU11A", age_group="JU11")
    holmen = Team(club="Holmen", label="Holmen JU11A", age_group="JU11")

    t1 = Tournament(
        id="jar12345",
        date=date(2025, 10, 5),
        arena="Jarhallen",
        age_group="U10",
        host_club="Jar",
        teams=[jar, kongsberg],
        games=[Game(home=jar, away=kongsberg, round_number=1)],
        start_time="09:00",
    )
    t2 = Tournament(
        id="ski12345",
        date=date(2025, 10, 12),
        arena="Skien ishall",
        age_group="JU11",
        host_club="Skien",
        teams=[skien, holmen],
        games=[Game(home=skien, away=holmen, round_number=1)],
        start_time="10:00",
    )
    return SeasonPlan(tournaments=[t1, t2], start_date=date(2025, 9, 1), end_date=date(2025, 12, 1))


def _write_state(tmp_path):
    state = PipelineState(tmp_path / "pipeline")
    input_file = tmp_path / "input.json"
    input_file.write_text(
        json.dumps(
            {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "age_groups": ["U10", "JU11"],
                "parallel_games": {"U10": 1, "JU11": 1},
                "round_length_minutes": {"U10": 15, "JU11": 12},
                "teams": [
                    {"club": "Jar", "label": "Jar U10A", "age_group": "U10"},
                    {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
                    {"club": "Skien", "label": "Skien JU11A", "age_group": "JU11"},
                    {"club": "Holmen", "label": "Holmen JU11A", "age_group": "JU11"},
                ],
            }
        ),
        encoding="utf-8",
    )
    state.write_stage(
        StageName.CONFIG,
        {
            "input_path": str(input_file),
            "teams": [
                {"club": "Jar", "label": "Jar U10A", "age_group": "U10"},
                {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
                {"club": "Skien", "label": "Skien JU11A", "age_group": "JU11"},
                {"club": "Holmen", "label": "Holmen JU11A", "age_group": "JU11"},
            ],
            "round_length_minutes": {"U10": 15, "JU11": 12},
        },
        status=StageStatus.DONE,
    )

    plan = _make_plan()
    state.write_stage(StageName.PLANNING, {"plan": _plan_to_dict(plan), "rules_report": []}, status=StageStatus.DONE)
    return state, plan


def test_stage4_writes_club_review_packets(tmp_path):
    state, _plan = _write_state(tmp_path)

    result = run(
        {"plan": _plan_to_dict(_make_plan())},
        state,
        export_dir=str(tmp_path / "export"),
    )

    review_dir = Path(result["output_files"]["review_packets"])
    kongsberg_dir = review_dir / "Kongsberg"
    assert review_dir.exists()
    assert kongsberg_dir.exists()
    assert (kongsberg_dir / "club_review.xlsx").exists()
    assert (kongsberg_dir / "club_review_spond.xlsx").exists()
    assert (kongsberg_dir / "club_review_spond_games.xlsx").exists()
    assert (kongsberg_dir / "manifest.json").exists()
    assert (kongsberg_dir / "response_template.json").exists()

    workbook = openpyxl.load_workbook(kongsberg_dir / "club_review_spond.xlsx")
    sheet = workbook["Sesongplan"]
    rows = list(sheet.iter_rows(values_only=True))
    assert len(rows) == 2
    assert rows[1][6] == "Jar"
    assert rows[1][7] == "Jar, Kongsberg"

    summary = openpyxl.load_workbook(kongsberg_dir / "club_review.xlsx")
    overview = summary["Oversikt"]
    assert overview[2][0].value == "Totalt antall relevante turneringer: 1"


def test_review_command_applies_change_request_and_reexports(tmp_path):
    state, plan = _write_state(tmp_path)

    # First generate review packets so the response template exists.
    run({"plan": _plan_to_dict(plan)}, state, export_dir=str(tmp_path / "export"))

    review_dir = Path(tmp_path / "export" / "review_packets" / "Kongsberg")
    response_path = review_dir / "response_template.json"
    response = json.loads(response_path.read_text(encoding="utf-8"))
    response["decision"] = "change_request"
    response["requested_changes"] = {
        "lock_dates": [],
        "ban_dates": [],
        "pin_tournaments": [],
        "force_host_clubs": ["Kongsberg"],
        "exclude_host_clubs": ["Jar"],
    }
    response_path.write_text(json.dumps(response, indent=2, ensure_ascii=False), encoding="utf-8")

    code = rvv_main(
        [
            "review",
            "--work-dir",
            str(state.work_dir),
            "--export-dir",
            str(tmp_path / "review-export"),
            "--response",
            str(review_dir),
        ]
    )

    assert code == 0
    assert (tmp_path / "review-export" / "season_plan.xlsx").exists()
    updated = _dict_to_plan(state.read_stage(StageName.PLANNING)["plan"])
    assert updated.tournaments[0].host_club == "Kongsberg"
    assert updated.tournaments[0].arena == "Kongsberghallen"
    assert updated.manual_adjustments["forced_host_clubs"] == ["Kongsberg"]
    assert updated.manual_adjustments["excluded_host_clubs"] == ["Jar"]
