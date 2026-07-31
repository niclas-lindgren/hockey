"""Tests for public activity/year-wheel JSON export (issue #33)."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from tournament_scheduler.pipeline.activity_export import build_activities_payload, has_activity_table, write_activities_json
from tournament_scheduler.pipeline.input_workbook import WorkbookInputError


def _workbook(path: Path, *, sheet_name: str = "Aktiviteter", rows: list[list[object]] | None = None) -> None:
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    settings.append(["start_date", "2026-01-01"])

    ws = wb.create_sheet(sheet_name)
    for row in rows or []:
        ws.append(row)

    secret = wb.create_sheet("Kilder")
    secret.append(["name", "type", "url"])
    secret.append(["Hemmelig", "ical", "https://secret.example.com/feed.ics"])
    wb.save(path)


class TestActivityExport:
    def test_returns_none_when_no_activity_sheet_exists(self, tmp_path):
        path = tmp_path / "input.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Innstillinger"
        wb.save(path)

        assert has_activity_table(path) is False
        assert build_activities_payload(path) is None

    def test_normalizes_records_and_sorts_by_actual_date(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Måned", "Dato", "Aktivitet", "Sted"],
                ["Mars", 3, "Klubbforum U13/U14", "Kongsberg"],
                ["Januar", 17, "Spillerutvikling JU14", "Sandefjord"],
            ],
        )

        payload = build_activities_payload(path, default_year=2026, generated_at="2026-07-29T12:00:00Z")

        assert payload["schema_version"] == 2
        assert payload["generated_at"] == "2026-07-29T12:00:00Z"
        assert payload["year"] == 2026
        assert payload["validation_warnings"] == []
        assert payload["activities"] == [
            {
                "date": "2026-01-17",
                "type": "spillerutviklingssamling",
                "category": "spillerutviklingssamling",
                "category_label": "Spillerutviklingssamling",
                "category_code": "SU",
                "raw_category": None,
                "age_groups": ["JU14"],
                "title": "Spillerutvikling JU14",
                "location": "Sandefjord",
                "description": None,
                "url": None,
            },
            {
                "date": "2026-03-03",
                "type": "annet",
                "category": "annet",
                "category_label": "Annen aktivitet",
                "category_code": "AN",
                "raw_category": None,
                "age_groups": ["U13", "U14"],
                "title": "Klubbforum U13/U14",
                "location": "Kongsberg",
                "description": None,
                "url": None,
            },
        ]

    def test_supports_explicit_columns_and_iso_dates(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["date", "title", "type", "age_groups", "location", "description", "url"],
                [
                    "2026-05-04",
                    "Jentesamling",
                    "samling",
                    "JU14, JU16",
                    "Jar",
                    "Ta med utstyr",
                    "https://example.com/info",
                ],
            ],
        )

        activity = build_activities_payload(path)["activities"][0]

        assert activity["date"] == "2026-05-04"
        assert activity["type"] == "unknown"
        assert activity["category"] == "unknown"
        assert activity["category_code"] == "?"
        assert activity["raw_category"] == "samling"
        assert activity["age_groups"] == ["JU14", "JU16"]
        assert activity["description"] == "Ta med utstyr"
        assert activity["url"] == "https://example.com/info"
        assert build_activities_payload(path)["validation_warnings"][0]["field"] == "category"

    def test_prefers_compact_activity_table_when_sheet_has_helper_headers(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            sheet_name="Årshjul",
            rows=[
                ["📅 Årshjul – aktiviteter"],
                [],
                ["MånedNr", "Måned", "Aktivitet", "Verdi", None, "Måned", "Antall", None, None, "Måned", "Dato", "Aktivitet", "Sted", "Verdi"],
                [1, "Januar", "46037 – Aktivitet 27 (Naja)", 1, None, "Januar", 2, None, None, "Desember", "2026-12-15", "RS U15", "Kongsberg", 1],
                [1, "Januar", "46053 – Aktivitet 26 (Magnus)", 1, None, "Januar", 2, None, None, "September", "2026-09-30", "S.Utvikling JU14", "Sandefjord", 1],
            ],
        )

        activities = build_activities_payload(path)["activities"]

        assert [activity["title"] for activity in activities] == ["S.Utvikling JU14", "RS U15"]
        assert activities[0]["location"] == "Sandefjord"
        assert "46037" not in json.dumps(activities, ensure_ascii=False)

    def test_finds_header_below_help_text_and_skips_example_rows(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Hjelp", "Denne teksten skal ikke eksporteres"],
                ["Eksempel", "17. januar", "Eksempelaktivitet U13"],
                ["Måned", "Dato", "Aktivitet", "Sted"],
                ["Eksempel", 1, "Eksempel: fyll inn aktivitet", "Test"],
                ["April", 9, "Dommerkurs U15", "Tønsberg"],
            ],
        )

        activities = build_activities_payload(path, default_year=2026)["activities"]

        assert [activity["title"] for activity in activities] == ["Dommerkurs U15"]
        assert activities[0]["date"] == "2026-04-09"

    def test_never_reads_internal_workbook_sheets(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Dato", "Aktivitet"],
                ["2026-02-01", "Regionsamling U13"],
            ],
        )

        payload = json.dumps(build_activities_payload(path), ensure_ascii=False)

        assert "secret.example.com" not in payload
        assert "Hemmelig" not in payload
        assert "Innstillinger" not in payload

    def test_malformed_date_reports_sheet_and_source_row(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Dato", "Aktivitet"],
                ["32.13.2026", "Spillerutvikling U13"],
            ],
        )

        with pytest.raises(WorkbookInputError, match="Aktiviteter!rad 2: ugyldig dato '32.13.2026'"):
            build_activities_payload(path)

    def test_missing_title_rows_are_skipped_with_warning(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Dato", "Aktivitet"],
                ["2026-02-01", ""],
                ["2026-02-02", "RS U15"],
            ],
        )

        payload = build_activities_payload(path)

        assert [activity["title"] for activity in payload["activities"]] == ["RS U15"]
        assert payload["validation_warnings"][0]["field"] == "title"
        assert payload["validation_warnings"][0]["row"] == 2
        assert "hoppes over" in payload["validation_warnings"][0]["message"]

    def test_day_and_month_without_default_year_is_actionable(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Måned", "Dato", "Aktivitet"],
                ["Januar", 17, "Spillerutvikling U13"],
            ],
        )

        with pytest.raises(WorkbookInputError, match="mangler år"):
            build_activities_payload(path)

    def test_normalizes_legacy_category_values_to_canonical_vocabulary(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Dato", "Aktivitet", "Type", "Aldersgruppe"],
                ["2026-01-01", "IA JU14", "IA", "JU14"],
                ["2026-01-02", "RS U15", "RS", "U15"],
                ["2026-01-03", "RM U16", "regionsmesterskap", "U16"],
                ["2026-01-04", "Turnering JU16", "regionsturneringju16", "JU16"],
            ],
        )

        payload = build_activities_payload(path)

        assert payload["validation_warnings"] == []
        assert [(a["category"], a["category_code"]) for a in payload["activities"]] == [
            ("spillerutviklingssamling", "SU"),
            ("regionslagssamling", "RS"),
            ("regionsmesterskap", "RM"),
            ("regionsturnering", "RT"),
        ]
        assert {item["id"] for item in payload["category_vocabulary"]} >= {
            "spillerutviklingssamling",
            "regionslagssamling",
            "regionsmesterskap",
            "regionsturnering",
            "unknown",
        }

    def test_unknown_category_and_age_group_emit_deterministic_warnings(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Dato", "Aktivitet", "Type", "Aldersgruppe"],
                ["2026-02-01", "Mystery event", "mystery", "Juniorer"],
            ],
        )

        payload = build_activities_payload(path)

        assert payload["activities"][0]["category"] == "unknown"
        assert payload["activities"][0]["category_code"] == "?"
        assert payload["activities"][0]["age_groups"] == []
        assert [(warning["field"], warning["raw_value"]) for warning in payload["validation_warnings"]] == [
            ("age_groups", "Juniorer"),
            ("category", "mystery"),
        ]

    def test_explicit_everyone_age_group_is_not_serialized_as_alle_lane(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Dato", "Aktivitet", "Type", "Aldersgruppe"],
                ["2026-02-01", "Felles samling", "SU", "Alle"],
            ],
        )

        activity = build_activities_payload(path)["activities"][0]

        assert activity["age_groups"] == ["ALL"]
        assert "Alle" not in activity["age_groups"]

    def test_write_activities_json_outputs_valid_json(self, tmp_path):
        path = tmp_path / "input.xlsx"
        _workbook(
            path,
            rows=[
                ["Dato", "Aktivitet"],
                ["2026-02-01", "Regionsamling U13"],
            ],
        )

        out = Path(write_activities_json(path, tmp_path / "export", generated_at="2026-07-29T12:00:00Z"))

        assert out.name == "activities.json"
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["schema_version"] == 2
        assert data["activities"][0]["date"] == "2026-02-01"
        assert data["activities"][0]["title"] == "Regionsamling U13"


# ---------------------------------------------------------------------------
# Power Automate content_json → schema-v2 tests (GitHub issue #55)
# ---------------------------------------------------------------------------

from tournament_scheduler.pipeline.activity_export import (
    ActivityJSONValidationError,
    build_activities_payload_from_values,
    validate_content_json,
)


class TestValidateContentJSON:
    def test_valid_payload_passes(self):
        payload = {
            "schemaVersion": 1,
            "worksheet": "Årshjul",
            "values": [["Måned", "Aktivitet"], ["Januar", "Klubbforum"]],
        }
        validate_content_json(payload)  # does not raise

    def test_rejects_non_dict(self):
        with pytest.raises(ActivityJSONValidationError, match="JSON-objekt"):
            validate_content_json([1, 2, 3])

    def test_rejects_wrong_schema_version(self):
        payload = {"schemaVersion": 2, "worksheet": "Årshjul", "values": []}
        with pytest.raises(ActivityJSONValidationError, match="schemaVersion må være 1"):
            validate_content_json(payload)

    def test_rejects_missing_schema_version(self):
        payload = {"worksheet": "Årshjul", "values": []}
        with pytest.raises(ActivityJSONValidationError, match="schemaVersion må være 1"):
            validate_content_json(payload)

    def test_rejects_wrong_worksheet(self):
        payload = {"schemaVersion": 1, "worksheet": "Aktiviteter", "values": []}
        with pytest.raises(ActivityJSONValidationError, match="worksheet må være"):
            validate_content_json(payload)

    def test_rejects_missing_values(self):
        payload = {"schemaVersion": 1, "worksheet": "Årshjul"}
        with pytest.raises(ActivityJSONValidationError, match="values må være"):
            validate_content_json(payload)

    def test_rejects_values_not_a_list(self):
        payload = {"schemaVersion": 1, "worksheet": "Årshjul", "values": "not-a-list"}
        with pytest.raises(ActivityJSONValidationError, match="values må være"):
            validate_content_json(payload)

    def test_rejects_empty_values(self):
        payload = {"schemaVersion": 1, "worksheet": "Årshjul", "values": []}
        with pytest.raises(ActivityJSONValidationError, match="values er tom"):
            validate_content_json(payload)

    def test_rejects_non_list_rows(self):
        payload = {"schemaVersion": 1, "worksheet": "Årshjul", "values": ["not-a-row"]}
        with pytest.raises(ActivityJSONValidationError, match="rad 1 er ikke en array"):
            validate_content_json(payload)

    def test_rejects_nested_object_in_cell(self):
        payload = {
            "schemaVersion": 1,
            "worksheet": "Årshjul",
            "values": [["Måned", {"nested": "object"}]],
        }
        with pytest.raises(ActivityJSONValidationError, match="er et objekt eller array"):
            validate_content_json(payload)

    def test_rejects_nested_array_in_cell(self):
        payload = {
            "schemaVersion": 1,
            "worksheet": "Årshjul",
            "values": [[["nested", "array"]]],
        }
        with pytest.raises(ActivityJSONValidationError, match="er et objekt eller array"):
            validate_content_json(payload)

    def test_allows_null_and_bool_cells(self):
        payload = {
            "schemaVersion": 1,
            "worksheet": "Årshjul",
            "values": [["Måned", "Aktivitet"], [None, True]],
        }
        validate_content_json(payload)  # does not raise

    def test_allows_numeric_cells(self):
        payload = {
            "schemaVersion": 1,
            "worksheet": "Årshjul",
            "values": [["Måned", "Dato"], ["Januar", 17]],
        }
        validate_content_json(payload)  # does not raise


class TestBuildActivitiesPayloadFromValues:
    def test_transforms_valid_values_to_schema_v2(self):
        values = [
            ["Måned", "Dato", "Aktivitet", "Sted"],
            ["Mars", 3, "Klubbforum U13/U14", "Kongsberg"],
            ["Januar", 17, "Spillerutvikling JU14", "Sandefjord"],
        ]
        payload = build_activities_payload_from_values(
            values, worksheet_name="Årshjul", default_year=2026, generated_at="2026-07-31T12:00:00Z"
        )

        assert payload["schema_version"] == 2
        assert payload["generated_at"] == "2026-07-31T12:00:00Z"
        assert payload["year"] == 2026
        assert len(payload["activities"]) == 2
        # Sorted by date: January before March.
        assert payload["activities"][0]["date"] == "2026-01-17"
        assert payload["activities"][0]["title"] == "Spillerutvikling JU14"
        assert payload["activities"][1]["date"] == "2026-03-03"
        assert payload["activities"][1]["title"] == "Klubbforum U13/U14"

    def test_infers_year_from_data_when_not_provided(self):
        values = [
            ["Dato", "Aktivitet"],
            ["2027-05-01", "Testaktivitet"],
        ]
        payload = build_activities_payload_from_values(
            values, worksheet_name="Årshjul", generated_at="2026-07-31T12:00:00Z"
        )
        assert payload["year"] == 2027

    def test_year_is_none_when_no_records(self):
        values = [
            ["Måned", "Aktivitet"],
            # No data rows with a title.
            ["", ""],
        ]
        payload = build_activities_payload_from_values(
            values, worksheet_name="Årshjul", default_year=None, generated_at="2026-07-31T12:00:00Z"
        )
        assert payload["year"] is None
        assert payload["activities"] == []

    def test_raises_when_no_header_found(self):
        values = [
            ["Kolonne1", "Kolonne2"],  # no recognised date/month/title headers
        ]
        with pytest.raises(WorkbookInputError, match="fant ikke en aktivitetstabell"):
            build_activities_payload_from_values(values, worksheet_name="Årshjul")

    def test_produces_same_output_as_xlsx_path_for_equivalent_data(self, tmp_path):
        """Smoke test: the JSON path yields the same schema-v2 output as the XLSX path."""
        rows = [
            ["Måned", "Dato", "Aktivitet", "Aldersgruppe", "Sted"],
            ["Februar", 14, "Spillerutvikling U9", "U9", "Jar"],
            ["Mars", 21, "Regionsturnering U12", "U12", "Holmen"],
        ]
        # XLSX path.
        xlsx_path = tmp_path / "input.xlsx"
        _workbook(xlsx_path, sheet_name="Årshjul", rows=rows)
        xlsx_payload = build_activities_payload(
            xlsx_path, default_year=2026, generated_at="2026-07-31T12:00:00Z"
        )

        # JSON path.
        json_payload = build_activities_payload_from_values(
            rows, worksheet_name="Årshjul", default_year=2026, generated_at="2026-07-31T12:00:00Z"
        )

        # Compare the activity arrays (generated_at differs).
        assert json_payload["schema_version"] == xlsx_payload["schema_version"]
        assert json_payload["year"] == xlsx_payload["year"]
        assert json_payload["activities"] == xlsx_payload["activities"]
        assert json_payload["validation_warnings"] == xlsx_payload["validation_warnings"]
        assert json_payload["category_vocabulary"] == xlsx_payload["category_vocabulary"]

    def test_empty_values_produces_empty_activities(self):
        values = [
            ["Dato", "Aktivitet"],
        ]
        payload = build_activities_payload_from_values(
            values, worksheet_name="Årshjul", default_year=2026, generated_at="2026-07-31T12:00:00Z"
        )
        assert payload["activities"] == []
        assert payload["year"] == 2026

    def test_skips_help_rows(self):
        values = [
            ["Måned", "Dato", "Aktivitet"],
            ["Help", 1, "Example activity to skip"],
            ["Mars", 15, "Klubbforum U13"],
        ]
        payload = build_activities_payload_from_values(
            values, worksheet_name="Årshjul", default_year=2026, generated_at="2026-07-31T12:00:00Z"
        )
        assert len(payload["activities"]) == 1
        assert payload["activities"][0]["title"] == "Klubbforum U13"


class TestBuildActivitiesPayloadJSONDispatch:
    def test_json_file_dispatches_to_power_automate_path(self, tmp_path):
        """build_activities_payload with a .json file uses the JSON path."""
        json_path = tmp_path / "activities.json"
        payload = {
            "schemaVersion": 1,
            "worksheet": "Årshjul",
            "values": [
                ["Dato", "Aktivitet", "Aldersgruppe"],
                ["2026-09-15", "Spillerutvikling U9", "U9"],
            ],
        }
        json_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

        result = build_activities_payload(json_path, default_year=2026, generated_at="2026-07-31T12:00:00Z")

        assert result is not None
        assert result["schema_version"] == 2
        assert len(result["activities"]) == 1
        assert result["activities"][0]["title"] == "Spillerutvikling U9"
        assert result["activities"][0]["age_groups"] == ["U9"]

    def test_json_file_returns_none_when_missing(self, tmp_path):
        json_path = tmp_path / "nonexistent.json"
        assert build_activities_payload(json_path) is None

    def test_json_file_returns_none_when_empty(self, tmp_path):
        json_path = tmp_path / "empty.json"
        json_path.write_text("", encoding="utf-8")
        assert build_activities_payload(json_path) is None

    def test_json_file_raises_on_invalid_json(self, tmp_path):
        json_path = tmp_path / "bad.json"
        json_path.write_text("not json", encoding="utf-8")
        with pytest.raises(WorkbookInputError, match="Kunne ikke lese JSON"):
            build_activities_payload(json_path)

    def test_json_file_raises_on_non_object(self, tmp_path):
        json_path = tmp_path / "array.json"
        json_path.write_text("[1, 2, 3]", encoding="utf-8")
        with pytest.raises(WorkbookInputError, match="må være et objekt"):
            build_activities_payload(json_path)

    def test_json_file_raises_on_contract_violation(self, tmp_path):
        json_path = tmp_path / "bad_contract.json"
        payload = {"schemaVersion": 2, "worksheet": "Feil", "values": []}
        json_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        with pytest.raises(ActivityJSONValidationError, match="schemaVersion må være 1"):
            build_activities_payload(json_path)
