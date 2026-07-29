"""Tests for the standalone activity calendar page (issues #33, #38, and #40)."""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

from tournament_scheduler.pipeline.activity_viewer import generate_activity_artifacts, generate_html


def _write_activity_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = "Innstillinger"
    ws = wb.create_sheet("Aktiviteter")
    ws.append(["Måned", "Dato", "Aktivitet", "Sted"])
    ws.append(["Januar", 17, "Spillerutvikling JU14", "Sandefjord"])
    ws.append(["Mars", 3, "Klubbforum U13/U14", "Kongsberg"])
    wb.save(path)


class TestActivityViewer:
    def test_generate_activity_artifacts_writes_json_and_standalone_page(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_activity_workbook(input_path)

        artifacts = generate_activity_artifacts(
            input_path=str(input_path),
            export_dir=str(tmp_path / "export"),
            default_year=2026,
            generated_at="2026-07-29T12:00:00Z",
        )

        assert artifacts is not None
        json_path = Path(artifacts["activities_json"])
        html_path = Path(artifacts["activities_html"])
        assert json_path.name == "activities.json"
        assert html_path.as_posix().endswith("activities/index.html")
        data = json.loads(json_path.read_text(encoding="utf-8"))
        assert data["schema_version"] == 2
        assert data["activities"][0]["date"] == "2026-01-17"
        assert data["activities"][0]["category"] == "spillerutviklingssamling"
        html = html_path.read_text(encoding="utf-8")
        assert "Aktivitetskalender for Region Viken Vest" in html
        assert "../activities.json" in html
        assert "SheetJS" not in html
        assert "xlsx" not in html.lower()

    def test_generate_activity_artifacts_returns_none_without_activity_sheet(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Innstillinger"
        wb.save(input_path)

        assert generate_activity_artifacts(input_path=str(input_path), export_dir=str(tmp_path / "export")) is None

    def test_page_contains_two_primary_views_filters_and_overlay_details(self, tmp_path):
        html_path = Path(generate_html(export_dir=str(tmp_path / "export")))
        html = html_path.read_text(encoding="utf-8")

        assert 'body class="view-overview"' in html
        assert "Sesongsløp" in html
        assert "Liste" in html
        assert "Årshjul" not in html
        assert "yearWheel" not in html
        assert "wheelView" not in html
        assert "Aldersgruppe" in html
        assert "Aktivitetstype" in html
        assert 'id="ageFilter"' in html
        assert 'id="typeFilter"' in html
        assert 'id="detailsDialog"' in html
        assert 'aria-labelledby="detailsTitle"' in html
        assert "detailsClose.addEventListener('click', closeDetails)" in html
        assert "detailsDialog.addEventListener('close', restoreFocus)" in html
        assert "lastActivator.focus" in html

    def test_page_positions_marker_overview_by_actual_date_including_leap_years(self, tmp_path):
        html = Path(generate_html(export_dir=str(tmp_path / "export"))).read_text(encoding="utf-8")

        assert "function dayOfYear(date)" in html
        assert "function daysInYear(year)" in html
        assert "? 366 : 365" in html
        assert "function dateToPercent(iso, year)" in html
        assert "daysInYear(targetYear) - 1" in html
        assert "class=\"timeline-marker" in html
        assert "width: 32px" in html
        assert "Fullt navn, sted og beskrivelse ligger i detaljer og i listen" in html

    def test_page_derives_lanes_handles_multi_age_all_age_and_collisions(self, tmp_path):
        html = Path(generate_html(export_dir=str(tmp_path / "export"))).read_text(encoding="utf-8")

        assert "function uniqueAgeGroups(activities)" in html
        assert "filter(group => group !== EVERYONE_GROUP)" in html
        assert "const EVERYONE_GROUP = 'ALL'" in html
        assert "groups.includes(EVERYONE_GROUP) ? lanes : groups" in html
        assert "buildLaneEntries(filteredActivities, lanes)" in html
        assert "function markerCollisionSpanPercent()" in html
        assert "getBoundingClientRect" in html
        assert "Math.abs(entry.x - previous.x) > collisionSpan" in html
        assert "occupiedStacks" in html
        assert "--stack:${entry.stack}" in html
        assert "Ingen aktiviteter har en konkret aldersgruppe" in html

    def test_page_uses_visible_category_legend_codes_and_non_color_cues(self, tmp_path):
        html = Path(generate_html(export_dir=str(tmp_path / "export"))).read_text(encoding="utf-8")

        assert "Forklaring av aktivitetstyper" in html
        assert "function renderLegend(types)" in html
        assert "fallbackVocabulary" in html
        assert "{ id: 'spillerutviklingssamling', code: 'SU'" in html
        assert "{ id: 'regionslagssamling', code: 'RS'" in html
        assert "{ id: 'regionsmesterskap', code: 'RM'" in html
        assert "{ id: 'regionsturnering', code: 'RT'" in html
        assert "const SHAPES = ['circle','square','diamond','ring','pill']" in html
        assert "shapeFor(type)" in html
        assert "marker-code" in html
        assert "categoryLabel(activity)" in html
        assert "categoryCode(activity)" in html

    def test_page_has_chronological_list_mobile_default_and_iframe_height_message(self, tmp_path):
        html = Path(generate_html(export_dir=str(tmp_path / "export"))).read_text(encoding="utf-8")

        assert "month-group" in html
        assert "activity-list" in html
        assert "matchMedia('(max-width: 760px)')" in html
        assert "mobile-default" in html
        assert "view-list" in html
        assert "rvv-activities-height" in html
        assert "HEIGHT_MESSAGE_NAMESPACE = 'rvv.activities'" in html
        assert "HEIGHT_MESSAGE_SCHEMA_VERSION = 1" in html
        assert "iframe_id: iframeId()" in html
        assert "postMessage" in html
        assert "}, parentTargetOrigin())" in html
        assert "postMessage({ type: 'rvv-activities-height', height: document.documentElement.scrollHeight }, '*')" not in html
        assert "overflow-x: hidden" in html
        assert "window.addEventListener('resize', () => { renderTimeline(); announceHeight('window-resize'); })" in html
        assert "window.addEventListener('orientationchange', () => announceHeight('orientation-change'))" in html

    def test_page_debounces_clamps_and_observes_height_changes(self, tmp_path):
        html = Path(generate_html(export_dir=str(tmp_path / "export"))).read_text(encoding="utf-8")

        assert "const MIN_EMBED_HEIGHT = 320" in html
        assert "const MAX_EMBED_HEIGHT = 6000" in html
        assert "function measuredDocumentHeight()" in html
        assert "Math.max(MIN_EMBED_HEIGHT, Math.min(MAX_EMBED_HEIGHT, measured))" in html
        assert "let heightRaf = 0" in html
        assert "cancelAnimationFrame(heightRaf)" in html
        assert "setTimeout(() => postHeight(reason || 'layout-fallback'), 180)" in html
        assert "if (height === lastPostedHeight) return" in html
        assert "function installHeightObserver()" in html
        assert "'ResizeObserver' in window" in html
        assert "new ResizeObserver(() => announceHeight('resize-observer'))" in html
        assert "document.fonts.ready.then(() => announceHeight('fonts-ready'))" in html

    def test_page_filters_and_views_share_normalized_data_without_reloading(self, tmp_path):
        html = Path(generate_html(export_dir=str(tmp_path / "export"))).read_text(encoding="utf-8")

        assert "function applyFilter()" in html
        assert "const selectedAge = ageFilter.value" in html
        assert "const selectedType = typeFilter.value" in html
        assert "categoryId(activity) === selectedType" in html
        assert "renderTimeline();" in html
        assert "renderList();" in html
        assert "overviewBtn.addEventListener('click', () => setView('overview'))" in html
        assert "listBtn.addEventListener('click', () => setView('list'))" in html
        assert "renderWheel" not in html

    def test_marker_and_list_buttons_have_accessible_non_hover_details(self, tmp_path):
        html = Path(generate_html(export_dir=str(tmp_path / "export"))).read_text(encoding="utf-8")

        assert "function accessibleName(activity, group)" in html
        assert "activity.location ?" in html
        assert "aria-label=\"${escapeHtml(accessibleName(activity, group))}\"" in html
        assert "title=\"${escapeHtml(accessibleName(activity, group))}\"" in html
        assert "showDetails(allActivities.find(a => a.id === item.dataset.id), item)" in html
        assert "showModal" in html
        assert "detailsClose.focus" in html
        assert "@media (prefers-reduced-motion: reduce)" in html

    def test_generate_html_output_is_deterministic_and_self_contained(self, tmp_path):
        first = Path(generate_html(export_dir=str(tmp_path / "first"))).read_text(encoding="utf-8")
        second = Path(generate_html(export_dir=str(tmp_path / "second"))).read_text(encoding="utf-8")

        assert first == second
        assert "../activities.json" in first
        assert "SheetJS" not in first
        assert "xlsx" not in first.lower()

    def test_wordpress_embed_docs_include_marker_view_and_iframe_resize_contract(self):
        docs = (Path(__file__).parents[1] / "docs" / "rvv-miniputt-pipeline.md").read_text(encoding="utf-8")

        assert "Sesongsløp" in docs
        assert "markør" in docs.lower()
        assert "rvv-activities-frame" in docs
        assert "rvv-activities-height" in docs
        assert "namespace: 'rvv.activities'" in docs
        assert "schema_version: 1" in docs
        assert "data.iframe_id" in docs
        assert "frame.contentWindow !== event.source" in docs
        assert "MAX_HEIGHT = 6000" in docs
        assert "event.origin !== EXPECTED_ORIGIN" in docs
        assert "Årshjul` view is removed" in docs
        assert "full-width" in docs
        assert "manual WordPress follow-ups" in docs
