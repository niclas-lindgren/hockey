"""Tests for registration CSV to workbook Lag-sheet sync."""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from tournament_scheduler.pipeline.registration_sync import (
    RegistrationSyncError,
    sync_registered_teams_to_workbook,
)


def _make_workbook(path: Path, lag_rows: list[dict[str, str]] | None = None) -> Path:
    """Create a minimal workbook with ``Lag`` and ``Innstillinger`` sheets."""
    wb = openpyxl.Workbook()
    # Innstillinger sheet (to verify it survives)
    ws_settings = wb.active
    ws_settings.title = "Innstillinger"
    ws_settings.append(["felt", "verdi"])
    ws_settings.append(["season_name", "Test 2026"])

    # Lag sheet
    ws_lag = wb.create_sheet("Lag")
    ws_lag.append(["club", "label", "age_group"])
    if lag_rows:
        for row in lag_rows:
            ws_lag.append([row.get("club", ""), row.get("label", ""), row.get("age_group", "")])

    wb.save(path)
    return path


def _make_csv(path: Path, rows: list[dict[str, str]] | None = None) -> Path:
    """Create a minimal registered-teams CSV."""
    lines = ["club,label,age_group"]
    if rows:
        for row in rows:
            lines.append(f"{row.get('club', '')},{row.get('label', '')},{row.get('age_group', '')}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return path


class TestSyncNoChanges:
    def test_empty_csv_to_empty_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = _make_csv(Path(tmp) / "teams.csv", [])
            wb_path = _make_workbook(Path(tmp) / "input.xlsx", [])
            result = sync_registered_teams_to_workbook(csv_path, wb_path)
            assert result["changed"] is False
            assert result["team_count"] == 0
            assert result["lag_rows_before"] == 0
            assert result["lag_rows_after"] == 0

    def test_same_rows_no_change(self):
        with tempfile.TemporaryDirectory() as tmp:
            rows = [{"club": "Jar", "label": "Jar 1", "age_group": "U10"}]
            csv_path = _make_csv(Path(tmp) / "teams.csv", rows)
            wb_path = _make_workbook(Path(tmp) / "input.xlsx", rows)
            result = sync_registered_teams_to_workbook(csv_path, wb_path)
            assert result["changed"] is False
            assert result["team_count"] == 1
            assert result["lag_rows_before"] == 1
            assert result["lag_rows_after"] == 1

    def test_same_rows_different_whitespace_no_change(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_rows = [{"club": "  Jar  ", "label": "Jar 1", "age_group": "U10"}]
            wb_rows = [{"club": "Jar", "label": "Jar 1", "age_group": "U10"}]
            csv_path = _make_csv(Path(tmp) / "teams.csv", csv_rows)
            wb_path = _make_workbook(Path(tmp) / "input.xlsx", wb_rows)
            result = sync_registered_teams_to_workbook(csv_path, wb_path)
            assert result["changed"] is False


class TestSyncWithChanges:
    def test_new_team_added(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = _make_csv(Path(tmp) / "teams.csv", [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
            ])
            wb_path = _make_workbook(Path(tmp) / "input.xlsx", [])
            result = sync_registered_teams_to_workbook(csv_path, wb_path)
            assert result["changed"] is True
            assert result["team_count"] == 1
            assert result["lag_rows_after"] == 1

            # Verify the workbook was updated.
            wb = openpyxl.load_workbook(wb_path, data_only=True)
            ws = wb["Lag"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            assert len(rows) == 1
            assert rows[0][0] == "Jar"

            # Verify other sheets survived.
            assert "Innstillinger" in wb.sheetnames
            assert wb["Innstillinger"].cell(2, 1).value == "season_name"

    def test_team_removed(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = _make_csv(Path(tmp) / "teams.csv", [])
            wb_path = _make_workbook(Path(tmp) / "input.xlsx", [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
            ])
            result = sync_registered_teams_to_workbook(csv_path, wb_path)
            assert result["changed"] is True
            assert result["team_count"] == 0
            assert result["lag_rows_after"] == 0

            wb = openpyxl.load_workbook(wb_path, data_only=True)
            ws = wb["Lag"]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            assert len(rows) == 0

    def test_team_label_changed(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = _make_csv(Path(tmp) / "teams.csv", [
                {"club": "Jar", "label": "Jar 2", "age_group": "U10"},
            ])
            wb_path = _make_workbook(Path(tmp) / "input.xlsx", [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
            ])
            result = sync_registered_teams_to_workbook(csv_path, wb_path)
            assert result["changed"] is True

    def test_multiple_teams_across_clubs(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = _make_csv(Path(tmp) / "teams.csv", [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
                {"club": "Jar", "label": "Jar 2", "age_group": "U10"},
                {"club": "Kongsberg", "label": "KIF 1", "age_group": "U12"},
                {"club": "Holmen", "label": "Holmen 1", "age_group": "JU10"},
            ])
            wb_path = _make_workbook(Path(tmp) / "input.xlsx", [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
            ])
            result = sync_registered_teams_to_workbook(csv_path, wb_path)
            assert result["changed"] is True
            assert result["team_count"] == 4
            assert result["lag_rows_after"] == 4


class TestSyncErrors:
    def test_missing_csv(self):
        with tempfile.TemporaryDirectory() as tmp:
            wb_path = _make_workbook(Path(tmp) / "input.xlsx")
            with pytest.raises(RegistrationSyncError, match="finnes ikke"):
                sync_registered_teams_to_workbook(Path(tmp) / "nonexistent.csv", wb_path)

    def test_missing_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = _make_csv(Path(tmp) / "teams.csv", [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
            ])
            with pytest.raises(RegistrationSyncError, match="finnes ikke"):
                sync_registered_teams_to_workbook(csv_path, Path(tmp) / "nonexistent.xlsx")

    def test_invalid_csv_missing_column(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = Path(tmp) / "teams.csv"
            csv_path.write_text("club,label\nJar,Jar 1\n", encoding="utf-8")
            wb_path = _make_workbook(Path(tmp) / "input.xlsx")
            with pytest.raises(RegistrationSyncError):  # wraps RegisteredTeamsValidationError
                sync_registered_teams_to_workbook(csv_path, wb_path)

    def test_csv_duplicate_rows_raises(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = _make_csv(Path(tmp) / "teams.csv", [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
            ])
            wb_path = _make_workbook(Path(tmp) / "input.xlsx")
            with pytest.raises(RegistrationSyncError):
                sync_registered_teams_to_workbook(csv_path, wb_path)


class TestDryRun:
    def test_dry_run_does_not_modify_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            csv_path = _make_csv(Path(tmp) / "teams.csv", [
                {"club": "Jar", "label": "Jar 1", "age_group": "U10"},
            ])
            wb_path = _make_workbook(Path(tmp) / "input.xlsx", [])
            original_bytes = wb_path.read_bytes()

            result = sync_registered_teams_to_workbook(csv_path, wb_path, dry_run=True)
            assert result["changed"] is True
            assert result["dry_run"] is True
            assert wb_path.read_bytes() == original_bytes
