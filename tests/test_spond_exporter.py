"""Tests for tournament_scheduler.spond.spond_exporter."""

from datetime import date

import openpyxl

from tournament_scheduler.models import Game, SeasonPlan, Team, Tournament
from tournament_scheduler.spond.spond_exporter import SpondExporter, _SPOND_HEADERS


def _sample_plan() -> SeasonPlan:
    teams_u10 = [
        Team(club="Kongsberg", label="Kongsberg U10A", age_group="U10"),
        Team(club="Skien", label="Skien U10A", age_group="U10"),
        Team(club="Holmen", label="Holmen U10A", age_group="U10"),
    ]
    t1 = Tournament(
        date=date(2026, 10, 10),
        arena="Kongsberghallen",
        age_group="U10",
        teams=teams_u10,
        games=[
            Game(home=teams_u10[0], away=teams_u10[1], round_number=1),
            Game(home=teams_u10[0], away=teams_u10[2], round_number=2),
            Game(home=teams_u10[1], away=teams_u10[2], round_number=3),
        ],
        host_club="Kongsberg",
        start_time="09:00",
    )

    teams_ju11 = [
        Team(club="Jutul", label="Jutul JU11A", age_group="JU11"),
        Team(club="Jar", label="Jar JU11A", age_group="JU11"),
    ]
    t2 = Tournament(
        date=date(2026, 11, 7),
        arena="Bærum ishall",
        age_group="JU11",
        teams=teams_ju11,
        games=[Game(home=teams_ju11[0], away=teams_ju11[1], round_number=1)],
        host_club="Jutul",
    )

    return SeasonPlan(tournaments=[t1, t2])


class TestSpondExporter:
    def test_default_export_is_tournament_level_with_filter_columns(self, tmp_path):
        plan = _sample_plan()
        output_path = tmp_path / "spond.xlsx"

        SpondExporter().export(
            plan,
            str(output_path),
            round_length_for_age_group={"U10": 15},
        )

        workbook = openpyxl.load_workbook(str(output_path))
        sheet = workbook["Sesongplan"]
        rows = list(sheet.iter_rows(values_only=True))

        assert rows[0] == tuple(_SPOND_HEADERS)
        assert sheet.auto_filter.ref == "A1:J3"
        assert rows[1][1] == "U10 Turnering — Kongsberghallen"
        assert rows[1][3] == "09:00"
        assert rows[1][4] == "09:45"
        assert rows[1][5] == "U10"
        assert rows[1][6] == "Kongsberg"
        assert rows[1][7] == "Kongsberg, Skien, Holmen"
        assert rows[1][8] == "Kongsberg U10A, Skien U10A, Holmen U10A"
        assert rows[1][9] == "turnering"
        assert rows[2][1].startswith("JU11 Turnering")
        assert rows[2][3] is None
        assert rows[2][4] is None

    def test_export_for_clubs_writes_prefiltered_workbooks(self, tmp_path):
        plan = _sample_plan()
        out_dir = tmp_path / "clubs"

        written = SpondExporter().export_for_clubs(
            plan,
            out_dir,
            clubs=["Kongsberg"],
            round_length_for_age_group={"U10": 15},
        )

        assert list(written) == ["Kongsberg"]
        club_path = out_dir / "season_plan_spond_Kongsberg.xlsx"
        assert written["Kongsberg"] == str(club_path)
        assert club_path.exists()

        workbook = openpyxl.load_workbook(str(club_path))
        sheet = workbook["Sesongplan"]
        rows = list(sheet.iter_rows(values_only=True))
        assert len(rows) == 2  # header + one Kongsberg-relevant tournament
        assert rows[1][6] == "Kongsberg"
        assert rows[1][7] == "Kongsberg, Skien, Holmen"
    def test_export_schedule_attachment_writes_one_sheet_per_tournament(self, tmp_path):
        plan = _sample_plan()
        output_path = tmp_path / "spond_games.xlsx"

        SpondExporter().export_schedule_attachment(
            plan,
            str(output_path),
            round_length_for_age_group={"U10": 15},
        )

        workbook = openpyxl.load_workbook(str(output_path))
        assert len(workbook.sheetnames) == 2

        first_sheet = workbook[workbook.sheetnames[0]]
        rows = list(first_sheet.iter_rows(values_only=True))
        header_row = next(i for i, row in enumerate(rows, start=1) if row[0] == "Runde")

        assert rows[0][0].startswith("10.10.2026")
        assert rows[1][0] == "Vertsklubb: Kongsberg"
        assert rows[2][0] == "Deltakende lag: Kongsberg U10A, Skien U10A, Holmen U10A"
        assert rows[header_row - 1][0:4] == ("Runde", "Hjemmelag", "Bortelag", "Parallellbane")
        assert rows[header_row][1] == "Kongsberg U10A"
        assert rows[header_row][3] == 1
