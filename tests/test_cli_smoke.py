"""CLI integration/smoke test: exercise the portable CLI end-to-end via a
real subprocess, with zero calendar sources so no network call is ever made.

Marked ``integration`` (subprocess-based, per pytest.ini's existing
convention) so it's excluded from the default quick ``pytest`` run but can
be targeted explicitly, e.g. by CI: ``pytest -m integration tests/test_cli_smoke.py``.
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]

pytestmark = pytest.mark.integration


def _write_minimal_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    settings.append(["start_date", "2025-09-01"])
    settings.append(["end_date", "2025-12-15"])

    age_groups = wb.create_sheet("Aldersgrupper")
    age_groups.append(["age_group", "parallel_games", "round_length_minutes"])
    age_groups.append(["U10", 2, None])

    teams = wb.create_sheet("Lag")
    teams.append(["club", "label", "age_group"])
    for club in ("Kongsberg", "Skien", "Ringerike", "Jar"):
        teams.append([club, f"{club} U10A", "U10"])

    # Deliberately no "Kilder" (sources) sheet — Stage 2 has nothing to
    # scrape, so this smoke test makes zero network calls.
    wb.save(path)


def _run_cli(args: list[str], *, timeout: int = 120) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, "-m", "tournament_scheduler.cli.rvv_cli", *args],
        cwd=ROOT,
        capture_output=True,
        text=True,
        timeout=timeout,
    )


class TestOperatorRunSmoke:
    def test_operator_run_completes_offline_without_crashing(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_minimal_workbook(input_path)
        work_dir = tmp_path / ".pipeline"
        export_dir = tmp_path / "export"

        result = _run_cli([
            "operator", "run",
            "--input", str(input_path),
            "--work-dir", str(work_dir),
            "--export-dir", str(export_dir),
            "--non-strict",
            "--allow-missing-sources",
            "--no-timestamped-export",
        ])

        assert result.returncode in (0, 1), result.stdout + result.stderr
        assert "Traceback" not in result.stderr

        manifest_path = work_dir / "run_manifest.json"
        assert manifest_path.exists()
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        assert manifest["final_outcome"] in ("ok", "warning", "blocked", "failed")
        assert manifest["objective"]
        capability_names = [c["capability"] for c in manifest["capabilities"]]
        assert capability_names == ["config", "scraping", "planning", "export"]

    def test_second_invocation_reports_nothing_pending(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_minimal_workbook(input_path)
        work_dir = tmp_path / ".pipeline"
        export_dir = tmp_path / "export"
        common_args = [
            "operator", "run",
            "--input", str(input_path),
            "--work-dir", str(work_dir),
            "--export-dir", str(export_dir),
            "--non-strict",
            "--allow-missing-sources",
            "--no-timestamped-export",
        ]

        first = _run_cli(common_args)
        assert first.returncode in (0, 1), first.stdout + first.stderr

        second = _run_cli(common_args)
        assert second.returncode == 0, second.stdout + second.stderr
        assert "ingenting å gjøre" in second.stdout


class TestPortableCliSmoke:
    def test_status_json_on_fresh_workspace(self, tmp_path):
        result = _run_cli(["status", "--json", "--work-dir", str(tmp_path / ".pipeline")])
        assert result.returncode == 0, result.stderr
        data = json.loads(result.stdout)
        assert "final_outcome" in data

    def test_sources_status_on_fresh_workspace(self, tmp_path):
        result = _run_cli(["sources", "status", "--work-dir", str(tmp_path / ".pipeline")])
        assert result.returncode == 0, result.stderr

    def test_operator_questions_on_fresh_workspace(self, tmp_path):
        result = _run_cli(["operator", "questions", "--json", "--work-dir", str(tmp_path / ".pipeline")])
        assert result.returncode == 0, result.stderr
        assert json.loads(result.stdout) == []

    def test_candidates_on_fresh_workspace(self, tmp_path):
        result = _run_cli(["candidates", "--work-dir", str(tmp_path / ".pipeline")])
        assert result.returncode == 0, result.stderr

    def test_unknown_command_prints_help_and_exits_0(self):
        result = _run_cli([])
        assert result.returncode == 0
        assert "rvv-miniputt" in result.stdout
