"""Tests for semantic_validation module and Stage 1 LLM integration."""

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

from tournament_scheduler.pipeline.semantic_validation import (
    build_semantic_prompt,
    parse_semantic_warnings,
)
from tournament_scheduler.pipeline.stage1_config import run
from tournament_scheduler.pipeline.state import PipelineState, StageName


# ---------------------------------------------------------------------------
# Helpers (mirrors test_stage1_config.py pattern)
# ---------------------------------------------------------------------------


def _make_valid_raw() -> dict:
    return {
        "start_date": "2025-09-01",
        "end_date": "2025-12-01",
        "teams": [
            {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
            {"club": "Skien",     "label": "Skien U10A",     "age_group": "U10"},
        ],
    }


def _write_input_workbook(path: Path, raw: dict | None = None) -> None:
    raw = raw or _make_valid_raw()
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    for key in ("start_date", "end_date", "target_tournament_count", "deltakelser_per_lag"):
        if key in raw:
            settings.append([key, raw[key]])

    if "age_groups" in raw:
        age_groups = wb.create_sheet("Aldersgrupper")
        age_groups.append(["age_group", "parallel_games", "round_length_minutes"])
        for ag in raw["age_groups"]:
            age_groups.append([
                ag,
                raw.get("parallel_games", {}).get(ag, 3),
                raw.get("round_length_minutes", {}).get(ag, 10),
            ])

    teams = wb.create_sheet("Lag")
    teams.append(["club", "label", "age_group"])
    for team in raw.get("teams", []):
        teams.append([team.get("club"), team.get("label"), team.get("age_group")])

    sources = wb.create_sheet("Kilder")
    sources.append(["name", "type", "url"])
    for source in raw.get("sources", [
        {"name": "Kongsberg", "type": "outlook", "url": "https://example.test/calendar"}
    ]):
        sources.append([source.get("name"), source.get("type"), source.get("url")])

    wb.save(path)


def _make_llm_response(text: str) -> MagicMock:
    """Return a mock LLM response object with a .text attribute."""
    resp = MagicMock()
    resp.text = text
    return resp


# ---------------------------------------------------------------------------
# Tests: build_semantic_prompt
# ---------------------------------------------------------------------------


class TestBuildSemanticPrompt:
    _base_config = {
        "start_date": "2025-09-01",
        "end_date": "2025-12-01",
        "teams": [
            {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
            {"club": "Skien",     "label": "Skien U10A",     "age_group": "U10"},
        ],
        "age_groups": ["U10", "U12"],
        "parallel_games": {"U10": 3, "U12": 2},
        "target_tournament_count": 5,
        "round_length_minutes": {"U10": 20, "U12": 25},
    }

    def test_returns_two_non_empty_strings(self):
        sys_p, usr_p = build_semantic_prompt(self._base_config)
        assert isinstance(sys_p, str) and len(sys_p) > 0
        assert isinstance(usr_p, str) and len(usr_p) > 0

    def test_user_prompt_contains_weekend_count(self):
        # 2025-09-01 to 2025-12-01: count Saturdays
        import datetime
        start = datetime.date(2025, 9, 1)
        end = datetime.date(2025, 12, 1)
        saturdays = sum(
            1 for d in (start + datetime.timedelta(n) for n in range((end - start).days + 1))
            if d.weekday() == 5
        )
        _, usr_p = build_semantic_prompt(self._base_config)
        assert str(saturdays) in usr_p

    def test_user_prompt_contains_age_group_names(self):
        _, usr_p = build_semantic_prompt(self._base_config)
        assert "U10" in usr_p
        assert "U12" in usr_p

    def test_user_prompt_contains_host_club_names(self):
        _, usr_p = build_semantic_prompt(self._base_config)
        assert "Kongsberg" in usr_p
        assert "Skien" in usr_p

    def test_user_prompt_contains_target_tournament_count(self):
        _, usr_p = build_semantic_prompt(self._base_config)
        assert "5" in usr_p


# ---------------------------------------------------------------------------
# Tests: parse_semantic_warnings
# ---------------------------------------------------------------------------


class TestParseSemanticWarnings:
    def test_parses_numbered_list(self):
        text = "1. Too many tournaments\n2. Not enough clubs"
        result = parse_semantic_warnings(text)
        assert len(result) == 2
        assert "Too many tournaments" in result[0]
        assert "Not enough clubs" in result[1]

    def test_parses_bullet_list(self):
        text = "- Too many tournaments\n- Not enough clubs"
        result = parse_semantic_warnings(text)
        assert len(result) == 2
        assert "Too many tournaments" in result[0]
        assert "Not enough clubs" in result[1]

    def test_returns_empty_for_no_issues_detected(self):
        assert parse_semantic_warnings("no issues detected") == []

    def test_returns_empty_for_empty_string(self):
        assert parse_semantic_warnings("") == []

    def test_no_leading_number_or_bullet_in_warnings(self):
        text = "1. Too many tournaments\n2. Not enough clubs"
        result = parse_semantic_warnings(text)
        for w in result:
            assert not w.startswith("1.")
            assert not w.startswith("2.")
            assert not w.startswith("-")
            assert not w.startswith("*")

    def test_returns_empty_for_no_issues_phrase_variants(self):
        for phrase in ("no issues", "no problems", "plan looks feasible"):
            assert parse_semantic_warnings(phrase) == [], f"Expected [] for: {phrase!r}"


# ---------------------------------------------------------------------------
# Tests: stage1_config.run() with llm_client
# ---------------------------------------------------------------------------


class TestStage1ConfigWithLlmClient:
    def test_run_without_llm_client_no_semantic_warnings_in_checkpoint(self, tmp_path):
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file)
        state = PipelineState(tmp_path / "pipeline")

        result = run(input_file, state, llm_client=None)

        assert "semantic_warnings" not in result
        checkpoint = state.read_stage(StageName.CONFIG)
        assert "semantic_warnings" not in (checkpoint or {})

    def test_run_with_llm_client_calls_complete_once(self, tmp_path):
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file)
        state = PipelineState(tmp_path / "pipeline")

        mock_client = MagicMock()
        mock_client.complete.return_value = _make_llm_response("no issues detected")

        run(input_file, state, llm_client=mock_client)

        mock_client.complete.assert_called_once()

    def test_run_with_llm_warnings_stored_in_checkpoint(self, tmp_path):
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file)
        state = PipelineState(tmp_path / "pipeline")

        mock_client = MagicMock()
        mock_client.complete.return_value = _make_llm_response(
            "1. Not enough clubs to host all age groups.\n2. Too many tournaments needed."
        )

        result = run(input_file, state, llm_client=mock_client)

        assert "semantic_warnings" in result
        assert len(result["semantic_warnings"]) == 2
        # Also persisted in checkpoint
        checkpoint = state.read_stage(StageName.CONFIG) or {}
        assert "semantic_warnings" in checkpoint

    def test_run_graceful_when_llm_raises(self, tmp_path):
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file)
        state = PipelineState(tmp_path / "pipeline")

        mock_client = MagicMock()
        mock_client.complete.side_effect = RuntimeError("LLM unavailable")

        result = run(input_file, state, llm_client=mock_client)

        # run() must succeed despite the LLM error
        assert "teams" in result
        assert "semantic_warnings" not in result
