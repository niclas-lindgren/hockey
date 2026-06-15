"""Tests for tournament_scheduler.pipeline.stage1_config."""

from pathlib import Path

import pytest
import openpyxl

from tournament_scheduler.pipeline.stage1_config import (
    Stage1Error,
    load_effective_config,
    run,
    validate_config,
)
from tournament_scheduler.pipeline.state import PipelineState, StageName, StageStatus


def _make_valid_raw():
    return {
        "start_date": "2025-09-01",
        "end_date": "2025-12-01",
        "teams": [
            {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
            {"club": "Skien",     "label": "Skien U10A",     "age_group": "U10"},
        ],
    }


def _write_input_workbook(path: Path, raw: dict | None = None) -> None:
    raw = raw or _make_valid_raw()
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    for key in ("start_date", "end_date", "target_tournament_count", "deltakelser_per_lag"):
        if key in raw:
            settings.append([key, raw[key]])

    if "age_groups" in raw:
        age_groups = wb.create_sheet("Aldersgrupper")
        age_groups.append(["age_group", "parallel_games", "round_length_minutes"])
        for age_group in raw["age_groups"]:
            age_groups.append([
                age_group,
                raw.get("parallel_games", {}).get(age_group, 3),
                raw.get("round_length_minutes", {}).get(age_group, 10),
            ])

    teams = wb.create_sheet("Lag")
    header_cols = ["club", "label", "age_group"]
    team_has_ttc = any("target_tournament_count" in team for team in raw.get("teams", []))
    if team_has_ttc:
        header_cols.append("target_tournament_count")
    teams.append(header_cols)
    for team in raw.get("teams", []):
        row = [team.get("club"), team.get("label"), team.get("age_group")]
        if team_has_ttc:
            row.append(team.get("target_tournament_count"))
        teams.append(row)

    sources = wb.create_sheet("Kilder")
    sources.append(["name", "type", "url"])
    for source in raw.get("sources", [
        {"name": "Kongsberg", "type": "outlook", "url": "https://example.test/calendar"}
    ]):
        sources.append([source.get("name"), source.get("type"), source.get("url")])

    wb.save(path)


class TestValidateConfig:
    def test_valid_config_has_no_errors(self):
        assert validate_config(_make_valid_raw()) == []

    def test_missing_start_date(self):
        raw = _make_valid_raw()
        del raw["start_date"]
        errors = validate_config(raw)
        assert any("start_date" in e for e in errors)

    def test_missing_end_date(self):
        raw = _make_valid_raw()
        del raw["end_date"]
        errors = validate_config(raw)
        assert any("end_date" in e for e in errors)

    def test_end_before_start_produces_error(self):
        raw = _make_valid_raw()
        raw["end_date"] = "2025-08-01"
        errors = validate_config(raw)
        assert any("end_date" in e for e in errors)

    def test_period_too_short(self):
        raw = _make_valid_raw()
        raw["start_date"] = "2025-09-01"
        raw["end_date"] = "2025-09-03"
        errors = validate_config(raw)
        assert any("dager" in e for e in errors)

    def test_invalid_date_format(self):
        raw = _make_valid_raw()
        raw["start_date"] = "01.09.2025"
        errors = validate_config(raw)
        assert any("ÅÅÅÅ-MM-DD" in e or "format" in e.lower() for e in errors)

    def test_missing_teams(self):
        raw = _make_valid_raw()
        del raw["teams"]
        errors = validate_config(raw)
        assert any("teams" in e for e in errors)

    def test_empty_teams_list(self):
        raw = _make_valid_raw()
        raw["teams"] = []
        errors = validate_config(raw)
        assert errors

    def test_unknown_age_group_in_team(self):
        raw = _make_valid_raw()
        raw["teams"][0]["age_group"] = "U99"
        errors = validate_config(raw)
        assert any("U99" in e for e in errors)

    def test_invalid_parallel_games_type(self):
        raw = _make_valid_raw()
        raw["parallel_games"] = "lots"
        errors = validate_config(raw)
        assert any("parallel_games" in e for e in errors)

    def test_parallel_games_exceed_federation_max(self):
        raw = _make_valid_raw()
        raw["parallel_games"] = {"U10": 999}
        errors = validate_config(raw)
        assert any("999" in e or "forbundets" in e for e in errors)

    def test_error_messages_are_norwegian(self):
        errors = validate_config({})
        # Norwegian error messages should contain Norwegian words
        combined = " ".join(errors)
        assert any(word in combined for word in ["Mangler", "felt", "Oppgi"])

    def test_deltakelser_per_lag_accepted(self):
        """The Norwegian alias `deltakelser_per_lag` passes validation."""
        raw = _make_valid_raw()
        raw.pop("target_tournament_count", None)
        raw["deltakelser_per_lag"] = 6
        assert validate_config(raw) == []

    def test_deltakelser_per_lag_takes_priority(self):
        """When both are set, `deltakelser_per_lag` wins (no validation error)."""
        raw = _make_valid_raw()
        raw["target_tournament_count"] = 3
        raw["deltakelser_per_lag"] = 8
        assert validate_config(raw) == []

    def test_deltakelser_per_lag_invalid_rejected(self):
        """An invalid `deltakelser_per_lag` produces a Norwegian validation error."""
        raw = _make_valid_raw()
        raw.pop("target_tournament_count", None)
        raw["deltakelser_per_lag"] = "mange"
        errors = validate_config(raw)
        assert any("deltakelser_per_lag" in e for e in errors)


class TestRunStage1:
    def test_run_writes_checkpoint_on_success(self, tmp_path):
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file)

        state = PipelineState(tmp_path / "pipeline")
        result = run(input_file, state)

        assert state.is_done(StageName.CONFIG)
        assert "teams" in result
        assert len(result["teams"]) == 2
        assert all(not key.startswith("derived_") for key in result)

    def test_run_accepts_excel_workbook_input(self, tmp_path):
        input_file = tmp_path / "input.xlsx"
        raw = _make_valid_raw()
        raw["age_groups"] = ["U10"]
        raw["parallel_games"] = {"U10": 3}
        raw["round_length_minutes"] = {"U10": 10}
        _write_input_workbook(input_file, raw)

        state = PipelineState(tmp_path / "pipeline")
        result = run(input_file, state)
        effective = load_effective_config(state, input_path=input_file)
        checkpoint = state.read_stage(StageName.CONFIG)

        assert state.is_done(StageName.CONFIG)
        assert len(result["teams"]) == 2
        assert checkpoint["teams"] == result["teams"]
        assert "start_date" not in checkpoint
        assert result["input_path"] == str(input_file.resolve())
        assert effective["start_date"] == "2025-09-01"
        assert effective["end_date"] == "2025-12-01"
        assert effective["age_groups"] == ["U10"]
        assert effective["parallel_games"] == {"U10": 3}
        assert effective["sources"] == [
            {"name": "Kongsberg", "type": "outlook", "url": "https://example.test/calendar"}
        ]

    def test_run_reports_missing_required_workbook_sheet(self, tmp_path):
        input_file = tmp_path / "input.xlsx"
        wb = openpyxl.Workbook()
        wb.active.title = "Innstillinger"
        wb.save(input_file)

        state = PipelineState(tmp_path / "pipeline")
        with pytest.raises(ValueError) as exc_info:
            run(input_file, state)

        assert "Lag" in str(exc_info.value)

    def test_run_accepts_deltakelser_per_lag_in_workbook(self, tmp_path):
        """The Norwegian alias `deltakelser_per_lag` works in the Innstillinger sheet."""
        raw = _make_valid_raw()
        raw.pop("target_tournament_count", None)
        raw["deltakelser_per_lag"] = 6
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file, raw)

        state = PipelineState(tmp_path / "pipeline")
        result = run(input_file, state)
        assert result["target_tournament_count"] == 6

    def test_run_preserves_per_team_target_tournament_count(self, tmp_path):
        """The per-team `target_tournament_count` column in the Lag sheet is preserved."""
        raw = _make_valid_raw()
        raw["teams"] = [
            {"club": "Kongsberg", "label": "Kongsberg 1", "age_group": "U10"},
            {"club": "Kongsberg", "label": "Kongsberg 2", "age_group": "U7", "target_tournament_count": 2},
            {"club": "Jar", "label": "Jar 1", "age_group": "U10", "target_tournament_count": 6},
        ]
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file, raw)

        state = PipelineState(tmp_path / "pipeline")
        result = run(input_file, state)
        teams = result["teams"]
        # Team without target
        kong1 = next(t for t in teams if t["label"] == "Kongsberg 1")
        assert "target_tournament_count" not in kong1 or kong1.get("target_tournament_count") is None
        # Team with target=2
        kong2 = next(t for t in teams if t["label"] == "Kongsberg 2")
        assert kong2["target_tournament_count"] == 2
        # Team with target=6
        jar1 = next(t for t in teams if t["label"] == "Jar 1")
        assert jar1["target_tournament_count"] == 6

    def test_run_rejects_json_input(self, tmp_path):
        input_file = tmp_path / "legacy.json"
        input_file.write_text("{}", encoding="utf-8")

        state = PipelineState(tmp_path / "pipeline")
        with pytest.raises(ValueError) as exc_info:
            run(input_file, state)

        assert "Excel" in str(exc_info.value)

    def test_run_does_not_inject_fallback_age_groups_when_input_omits_them(self, tmp_path):
        raw = _make_valid_raw()
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file, raw)
        state = PipelineState(tmp_path / "pipeline")
        result = run(input_file, state)

        assert all(not key.startswith("derived_") for key in result)
        effective = load_effective_config(state, input_path=input_file)
        assert effective.get("age_groups") == []
        assert effective.get("age_groups_from_input") is False

    def test_run_raises_on_invalid_config(self, tmp_path):
        raw = {"start_date": "bad", "end_date": "also-bad", "teams": []}
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file, raw)

        state = PipelineState(tmp_path / "pipeline")
        with pytest.raises(Stage1Error) as exc_info:
            run(input_file, state)
        assert len(exc_info.value.errors) > 0

    def test_run_raises_on_missing_field_no_checkpoint(self, tmp_path):
        raw = _make_valid_raw()
        del raw["start_date"]
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file, raw)

        state = PipelineState(tmp_path / "pipeline")
        with pytest.raises(Stage1Error):
            run(input_file, state)
        assert not state.checkpoint_path(StageName.CONFIG).exists()

    def test_run_raises_on_missing_file(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        with pytest.raises(FileNotFoundError):
            run(tmp_path / "nonexistent.xlsx", state)

    def test_run_strict_false_does_not_raise(self, tmp_path):
        raw = {}  # invalid but strict=False
        input_file = tmp_path / "input.xlsx"
        _write_input_workbook(input_file, raw)

        state = PipelineState(tmp_path / "pipeline")
        # Should not raise
        result = run(input_file, state, strict=False)
        # Result may be empty or partial
        assert isinstance(result, dict)
