"""Tests for tournament_scheduler.pipeline.stage4_export."""

from datetime import date, timedelta
from pathlib import Path
import json
import re

import openpyxl
import pytest

from tournament_scheduler.html.html_exporter import HtmlExporter
from tournament_scheduler.models import Game, Roster, SeasonPlan, Team, Tournament
from tournament_scheduler.pipeline.cache_manager import ScrapedDataCache
from tournament_scheduler.pipeline.stage4_export import Stage4Error, _dict_to_plan, run
from tournament_scheduler.pipeline.state import PipelineState, StageName, StageStatus


def _write_input_workbook(path: Path, raw: dict) -> None:
    wb = openpyxl.Workbook()
    settings = wb.active
    settings.title = "Innstillinger"
    settings.append(["felt", "verdi"])
    for key in ("start_date", "end_date", "target_tournament_count"):
        if key in raw:
            settings.append([key, raw[key]])

    if "age_groups" in raw:
        age_groups = wb.create_sheet("Aldersgrupper")
        age_groups.append(["age_group", "parallel_games", "round_length_minutes"])
        for age_group in raw["age_groups"]:
            age_groups.append([age_group, raw.get("parallel_games", {}).get(age_group), None])

    teams = wb.create_sheet("Lag")
    teams.append(["club", "label", "age_group"])
    for team in raw.get("teams", []):
        teams.append([team.get("club"), team.get("label"), team.get("age_group")])

    sources = wb.create_sheet("Kilder")
    sources.append(["name", "type", "url"])
    for source in raw.get("sources", []):
        sources.append([source.get("name"), source.get("type"), source.get("url")])

    wb.save(path)


def _make_plan_dict():
    """Build a minimal but valid plan checkpoint dict."""
    t1 = {
        "date": "2025-10-05",
        "arena": "Kongsberghallen",
        "age_group": "U10",
        "host_club": "Kongsberg",
        "teams": [
            {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
            {"club": "Skien",     "label": "Skien U10A",     "age_group": "U10"},
        ],
        "games": [
            {"home": "Kongsberg U10A", "away": "Skien U10A", "parallel_slot": 0, "round_number": 3},
        ],
        "start_time": "09:00",
    }
    return {
        "plan": {
            "start_date": "2025-09-01",
            "end_date": "2025-12-01",
            "diversity_score": 1.0,
            "pairwise_matchup_score": 1.0,
            "month_balance_score": 1.0,
            "arena_counts": {"Kongsberghallen": 1},
            "manual_adjustments": {
                "locked_dates": ["2025-10-05"],
                "banned_dates": ["2025-11-01"],
            },
            "fairness_gate": {
                "status": "pass",
                "score": 100,
                "metrics": [
                    {"label": "Kamper per lag", "value": 0, "threshold": 2, "status": "pass", "score": 100, "unit": "", "detail": "Lik kampfordeling."},
                    {
                        "key": "hosting_deviation",
                        "label": "Hjemmebanebelastning",
                        "value": 0.0,
                        "threshold": 1,
                        "status": "pass",
                        "score": 100,
                        "unit": "",
                        "detail": "Aldersgruppevis fordeling av hjemmeturneringer: U10 Kongsberg 1 vs ~1.0.",
                        "age_group_breakdown": [
                            {"age_group": "U10", "club": "Kongsberg", "actual": 1, "expected": 1.0},
                            {"age_group": "U10", "club": "Skien", "actual": 0, "expected": 0.0},
                        ],
                    },
                    {"label": "Månedsbalanse", "value": 1.0, "threshold": 0.75, "status": "pass", "score": 100, "unit": "", "detail": "Jevn sesongbelastning."},
                ],
            },
            "tournaments": [t1],
        },
        "llm_confidence": 0.9,
        "llm_reasoning": "great",
        "attempts": 1,
        "llm_skipped": True,
    }


def _make_multi_age_group_plan_dict():
    data = _make_plan_dict()
    data["plan"]["tournaments"].append(
        {
            "date": "2025-11-02",
            "arena": "Bærum ishall",
            "age_group": "JU11",
            "host_club": "Jutul",
            "teams": [
                {"club": "Jutul", "label": "Jutul JU11A", "age_group": "JU11"},
                {"club": "Holmen", "label": "Holmen JU11A", "age_group": "JU11"},
            ],
            "games": [
                {"home": "Jutul JU11A", "away": "Holmen JU11A", "parallel_slot": 0, "round_number": 1},
            ],
        }
    )
    data["plan"]["arena_counts"]["Bærum ishall"] = 1
    return data


def _make_spond_plan_dict():
    data = _make_plan_dict()
    data["plan"]["tournaments"][0]["teams"] = [
        {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
        {"club": "Skien", "label": "Skien U10A", "age_group": "U10"},
        {"club": "Holmen", "label": "Holmen U10A", "age_group": "U10"},
    ]
    data["plan"]["tournaments"][0]["games"] = [
        {"home": "Kongsberg U10A", "away": "Skien U10A", "parallel_slot": 0, "round_number": 1},
        {"home": "Kongsberg U10A", "away": "Holmen U10A", "parallel_slot": 0, "round_number": 2},
        {"home": "Skien U10A", "away": "Holmen U10A", "parallel_slot": 0, "round_number": 3},
    ]
    return data


class TestDictToPlan:
    def test_round_trips_plan(self):
        plan_dict = _make_plan_dict()["plan"]
        plan = _dict_to_plan(plan_dict)
        assert isinstance(plan, SeasonPlan)
        assert len(plan.tournaments) == 1
        t = plan.tournaments[0]
        assert t.arena == "Kongsberghallen"
        assert t.start_time == "09:00"
        assert len(t.games) == 1
        assert t.games[0].home.label == "Kongsberg U10A"
        assert t.games[0].round_number == 3
        assert plan.manual_adjustments["locked_dates"] == ["2025-10-05"]

    def test_round_trips_arena_day_collisions(self):
        plan_dict = _make_plan_dict()["plan"]
        plan_dict["arena_day_collisions"] = [
            {
                "date": "2025-10-05",
                "arena": "Jarahallen",
                "age_group": "U7",
                "host_club": "Jar",
                "conflicting_age_group": "U10",
                "conflicting_host_club": "Jar",
                "reason": "same_arena_same_day",
            }
        ]
        plan = _dict_to_plan(plan_dict)
        assert plan.arena_day_collisions[0]["arena"] == "Jarahallen"
        assert plan.arena_day_collisions[0]["conflicting_age_group"] == "U10"

    def test_handles_missing_dates(self):
        plan_dict = {"tournaments": [], "diversity_score": 0.0,
                     "pairwise_matchup_score": 0.0, "month_balance_score": 0.0,
                     "arena_counts": {}}
        plan = _dict_to_plan(plan_dict)
        assert plan.start_date is None


class TestRunStage4:
    def test_produces_excel_file(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(StageName.CONFIG, {"round_length_minutes": {"U10": 15}}, status=StageStatus.DONE)
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        assert "excel" in files
        assert Path(files["excel"]).exists()
        workbook = openpyxl.load_workbook(files["excel"])
        overview = workbook["Sesongoversikt"]
        rows = list(overview.iter_rows(values_only=True))
        assert rows[1][7] == "09:00"
        assert rows[1][8] == "09:45"

    def test_includes_fairness_gate_sheet(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        workbook = openpyxl.load_workbook(files["excel"])
        assert "Rettferdighetskontroll" in workbook.sheetnames
        assert "Rettferdighetsjusteringer" in workbook.sheetnames
        old_gate_label = "Rettferdighets" + "gate"
        old_adjustment_sheet = "Fairness" + "justeringer"
        assert old_gate_label not in workbook.sheetnames
        assert old_adjustment_sheet not in workbook.sheetnames
        sheet = workbook["Rettferdighetskontroll"]
        rows = list(sheet.iter_rows(values_only=True))
        assert rows[0][0] == "Overordnet status"
        assert rows[2][0] == "Metrikk"
        assert rows[3][0] == "Kamper per lag"
        adj = workbook["Rettferdighetsjusteringer"]
        adj_rows = list(adj.iter_rows(values_only=True))
        assert adj_rows[0][0] == "Rettferdighetsjusteringer per lag"
        assert "fairness" not in str(adj_rows[1][0]).lower()
        assert adj_rows[3][0] == "Lag"

    def test_generates_html_with_configured_age_group_filters(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(
            input_path,
            {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "age_groups": ["U10", "U11", "U12", "JU11"],
                "parallel_games": {"U10": 2, "U11": 2, "U12": 2, "JU11": 2},
                "teams": [],
                "sources": [],
            },
        )
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(
            StageName.CONFIG,
            {"input_path": str(input_path), "round_length_minutes": {}},
            status=StageStatus.DONE,
        )
        result = run(
            _make_multi_age_group_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        html_path = Path(files["html"])
        report_path = Path(files["html_report"])
        html = html_path.read_text(encoding="utf-8")
        report_html = report_path.read_text(encoding="utf-8")

        assert '<option value="U10">U10</option>' in html
        assert '<option value="JU11">JU11</option>' in html
        assert '<option value="U12">U12</option>' in html
        assert 'Alle (U10 + U11 + U12 + JU11)' in html
        assert 'id="themeToggle"' in html
        assert 'class="theme-toggle"' in html
        assert 'href="season_plan.xlsx"' in html
        assert 'href="season_plan.csv"' in html
        assert 'href="season_plan.ics"' in html
        assert 'href="season_plan.csv" class="export-link-btn"' in html or 'href="season_plan.csv"' in html
        assert html.index('class="export-links"') < html.index('class="header-main"')
        assert report_html.index('class="export-links"') < report_html.index('class="header-main"')
        assert html.index('class="export-links"') < html.index('class="stat-badge"')
        assert report_html.index('class="export-links"') < report_html.index('class="stat-badge"')
        old_gate_label = 'Rettferdighets' + 'gate'
        old_adjustment_label = 'Fairness' + '-justeringer'
        assert old_gate_label not in html
        assert old_adjustment_label not in html
        assert 'Rettferdighetskontroll' not in html
        assert 'Rettferdighetsjusteringer' not in html
        assert 'Kvalitetsgjennomgang' not in html
        assert 'id="timeline"' in html
        assert 'class="filters"' in html
        assert 'class="count-bar"' in html
        assert 'Ser planen jevn ut?' in report_html
        assert 'Rettferdighetsjusteringer' in report_html
        assert 'Per aldersgruppe og klubb: faktisk vs forventet hjemmeturneringer' in report_html
        assert 'Aldersgruppevis fordeling av hjemmeturneringer: U10 Kongsberg 1 vs ~1.0.' in report_html
        assert 'id="reportOverview"' in report_html
        assert 'Kan planen brukes?' in report_html
        assert 'Hva må sjekkes eller endres?' in report_html
        assert 'Ja, planen ser brukbar ut for klubbvis gjennomgang.' in report_html or 'Nesten, men punktene under' in report_html or 'Ikke ennå.' in report_html
        assert 'Hva skjer per aldersgruppe?' in report_html
        assert 'Hva må hver klubb vurdere?' in report_html
        assert 'Turneringer som skal gjennomgås' in report_html
        assert 'Detaljerte måltall og kontroller' in report_html
        assert 'Klubben bør sjekke' in report_html
        assert 'Egen vurdering' in report_html
        assert 'Dette er en separat tolkning av tallene' in report_html
        assert 'Min ærlige dom på hele planen' in report_html
        assert report_html.index('id="reportOverview"') < report_html.index('Ser planen jevn ut?')
        assert report_html.index('Hva må sjekkes eller endres?') < report_html.index('Detaljerte måltall og kontroller')
        assert report_html.index('Ser planen jevn ut?') < report_html.index('Detaljerte måltall og kontroller')
        assert report_html.index('id="opinionatedJudgment"') < report_html.index('id="detailedDiagnosticsIntro"')
        assert report_html.index('id="reportOverview"') < report_html.index('Egen vurdering')
        assert old_gate_label not in report_html
        assert old_adjustment_label not in report_html
        assert 'Rådgivende kontroll' in report_html
        assert 'Manglende klubber' in report_html
        assert 'id="clubReviewSummary"' in report_html
        assert 'id="teamStats"' in report_html
        assert 'id="travelStats"' in report_html
        assert 'id="heatmapSection"' in report_html
        # Heatmap must appear immediately after the hero block — before the card-grid and all other sections.
        assert report_html.index('id="heatmapSection"') < report_html.index('class="report-card-grid"')
        assert report_html.index('id="heatmapSection"') < report_html.index('id="priorityActions"')
        assert 'id="clubDashboard"' not in report_html
        assert 'style="display:none' not in report_html
        assert 'id="timeline"' not in report_html
        assert 'class="filters"' not in report_html
        assert 'class="count-bar"' not in report_html
        assert 'Klassetrinn' not in report_html
        assert 'Nullstill filter' not in report_html
        assert 'Viser 80 av 80 turneringer' not in report_html
        schedule_script = re.search(r"<script>\n(.*?)\n</script>", html, re.S).group(1)
        report_script = re.search(r"<script>\n(.*?)\n</script>", report_html, re.S).group(1)
        for schedule_identifier in ('filterAge', 'timeline', 'buildMatchHTML', 'function render()'):
            assert schedule_identifier in schedule_script
            assert schedule_identifier not in report_script
        assert 'rvv-theme' in report_script
        assert 'HEATMAP' in report_script
        assert 'clubSummaryBody' in report_script
        assert 'clubDashName' not in report_script
        assert 'debug-dashboard' not in html.lower()
        assert not re.search(r"[\U0001F300-\U0001FAFF]", html)
        assert not re.search(r"[\U0001F300-\U0001FAFF]", report_html)
        # Hero verdict pill must contain the action count — plan has 1 missing-host action.
        assert '1 punkt(er)' in report_html

    def test_review_summary_collapses_when_it_only_repeats_main_assessment(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(),
            state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")

        assert 'review-summary-panel--compact' in report_html
        assert 'Kortversjon av kontrollen' in report_html

    def test_report_missing_hosts_uses_canonical_club_aliases(self, tmp_path):
        """Short RVV club aliases should not trigger false missing-host warnings."""
        rvv_hosts = [
            "Ringerike",
            "Tønsberg",
            "Frisk Asker",
            "Sandefjord",  # Alias for canonical Sandefjord Penguins.
            "Jar",
            "Holmen",
            "Skien",
            "Jutul",
            "Kongsberg",
        ]
        start = date(2025, 10, 5)
        tournaments = []
        for index, host in enumerate(rvv_hosts):
            opponent = "Skien" if host != "Skien" else "Kongsberg"
            tournaments.append(
                {
                    "date": (start + timedelta(days=index * 7)).isoformat(),
                    "arena": f"{host} arena",
                    "age_group": "U10",
                    "host_club": host,
                    "teams": [
                        {"club": host, "label": f"{host} U10A", "age_group": "U10"},
                        {"club": opponent, "label": f"{opponent} U10A", "age_group": "U10"},
                    ],
                    "games": [
                        {
                            "home": f"{host} U10A",
                            "away": f"{opponent} U10A",
                            "parallel_slot": 0,
                            "round_number": 1,
                        },
                    ],
                }
            )
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            {
                "plan": {
                    "start_date": "2025-10-01",
                    "end_date": "2025-12-31",
                    "diversity_score": 1.0,
                    "pairwise_matchup_score": 1.0,
                    "month_balance_score": 1.0,
                    "arena_counts": {},
                    "tournaments": tournaments,
                }
            },
            state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")

        assert "Rådgivende kontroll" in report_html
        assert "Alle 9 RVV-klubber har minst én hjemmeturnering." in report_html
        assert "Følgende RVV-klubber har ingen hjemmeturnering" not in report_html

    def test_html_filters_fall_back_to_plan_age_groups_when_input_omits_them(self, tmp_path):
        input_path = tmp_path / "input.xlsx"
        _write_input_workbook(
            input_path,
            {
                "start_date": "2025-09-01",
                "end_date": "2025-12-01",
                "teams": [
                    {"club": "Kongsberg", "label": "Kongsberg U10A", "age_group": "U10"},
                ],
                "parallel_games": {"U10": 2},
                "sources": [],
            },
        )
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(
            StageName.CONFIG,
            {"input_path": str(input_path), "round_length_minutes": {}},
            status=StageStatus.DONE,
        )
        result = run(
            _make_multi_age_group_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        html = Path(result["output_files"]["html"]).read_text(encoding="utf-8")

        assert '<option value="U10">U10</option>' in html
        assert '<option value="JU11">JU11</option>' in html
        assert 'Alle (JU11 + U10)' in html or 'Alle (U10 + JU11)' in html

    def test_html_tournament_details_group_matches_by_round(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        html = Path(files["html"]).read_text(encoding="utf-8")

        payload = json.loads(HtmlExporter._plan_to_json(_dict_to_plan(_make_plan_dict()["plan"])))
        assert payload[0]["m"][0][3] == 3
        assert 'Kamper per runde' in html
        assert 'round-group-header' in html
        assert 'Runde ' in html

    def test_produces_ical_file(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(StageName.CONFIG, {"round_length_minutes": {"U10": 15}}, status=StageStatus.DONE)
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        assert "ical" in files
        ics_path = Path(files["ical"])
        assert ics_path.exists()
        assert ics_path.suffix == ".ics"
        content = ics_path.read_text()
        assert "BEGIN:VCALENDAR" in content
        assert "VEVENT" in content
        assert "DTSTART:20251005T090000Z" in content
        assert "DTEND:20251005T100000Z" in content

    def test_writes_timestamped_exports_without_flat_copies(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=True,
        )
        files = result.get("output_files", {})

        timestamped_paths = [Path(files[key]) for key in ("excel", "ical", "csv_games", "csv_overview", "html", "html_report", "spond", "spond_games")]
        timestamp_dirs = {path.parent for path in timestamped_paths}
        assert len(timestamp_dirs) == 1
        timestamp_dir = timestamp_dirs.pop()
        assert timestamp_dir.parent == tmp_path / "export"
        assert timestamp_dir.name

        for path in timestamped_paths:
            assert path.exists()
            assert path.parent == timestamp_dir

        assert not any(key.endswith("_flat") for key in files)
        assert list((tmp_path / "export").glob("*.xlsx")) == []
        assert list((tmp_path / "export").glob("*.csv")) == []
        assert list((tmp_path / "export").glob("*.ics")) == []
        assert list((tmp_path / "export").glob("*.html")) == []

    def test_stage4_spond_export_uses_tournament_rows(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        state.write_stage(StageName.CONFIG, {"round_length_minutes": {"U10": 15}}, status=StageStatus.DONE)
        result = run(
            _make_spond_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        workbook = openpyxl.load_workbook(files["spond"])
        sheet = workbook["Sesongplan"]
        rows = list(sheet.iter_rows(values_only=True))

        assert rows[0][0:5] == ("Dato", "Aktivitet", "Sted", "Start", "Slutt")
        assert rows[1][3] == "09:00"
        assert rows[1][4] == "09:45"
        assert rows[1][9] == "turnering"
        assert len(rows) == 2  # header + one tournament row, not one row per game

        attachment = openpyxl.load_workbook(files["spond_games"])
        assert len(attachment.sheetnames) == 1
        attachment_rows = list(attachment[attachment.sheetnames[0]].iter_rows(values_only=True))
        header_row = next(i for i, row in enumerate(attachment_rows) if row[:4] == ("Runde", "Hjemmelag", "Bortelag", "Parallellbane"))
        assert attachment_rows[header_row][0:4] == ("Runde", "Hjemmelag", "Bortelag", "Parallellbane")
        assert attachment_rows[header_row + 1][1] == "Kongsberg U10A"

    def test_produces_csv_files(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        result = run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        files = result.get("output_files", {})
        assert "csv_games" in files
        assert "csv_overview" in files
        games_path = Path(files["csv_games"])
        assert games_path.exists()
        lines = games_path.read_text().splitlines()
        assert lines[0] == "date,arena,age_group,home,away,parallel_slot"
        assert len(lines) > 1  # header + at least one game row

    def test_marks_checkpoint_done(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        run(
            _make_plan_dict(), state,
            export_dir=str(tmp_path / "export"),
            timestamped_export=False,
        )
        assert state.is_done(StageName.EXPORT)

    def test_raises_on_missing_plan(self, tmp_path):
        state = PipelineState(tmp_path / "pipeline")
        with pytest.raises(Stage4Error, match="Stage 3"):
            run({}, state, export_dir=str(tmp_path / "export"), strict=True)

    def test_conclusion_injects_weakest_metric_name(self, tmp_path):
        """Conclusion must include weakest metric label when metric_warnings exist."""
        data = _make_plan_dict()
        data["plan"]["fairness_gate"] = {
            "status": "warn",
            "score": 70,
            "metrics": [
                {"label": "Kampbalanse", "value": 3, "threshold": 2, "status": "warn", "score": 60, "unit": "", "detail": "Ujevn fordeling."},
                {"label": "Hjemmebanebelastning", "value": 1.5, "threshold": 1, "status": "fail", "score": 40, "unit": "", "detail": "For stor belastning."},
            ],
        }
        state = PipelineState(tmp_path / "pipeline")
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        # Weakest metric is "Hjemmebanebelastning" (status=fail, score=40)
        assert "Svakeste metrikk: Hjemmebanebelastning" in report_html
        # Hero pill must show the issue count: gate warn (1) + 2 metric warnings (2) + missing hosts (1) = 4.
        assert "4 punkt(er)" in report_html

    def test_calendars_html_generated_when_scrape_cache_populated(self, tmp_path):
        """stage4 should write calendars.html when the scrape cache contains events."""
        work_dir = tmp_path / "pipeline"
        state = PipelineState(work_dir)
        # Populate scrape cache so generate_html has data to render
        cache = ScrapedDataCache(str(work_dir))
        cache.write({
            "_meta": {
                "source_count": 1,
                "total_events": 2,
                "updated_at": "2025-01-01T00:00:00",
            },
            "sources": {
                "TestClub": {
                    "events": [
                        {"date": "05.10.2025", "title": "Test", "source": "TestClub", "url": ""},
                    ],
                    "scrape_timestamp": "2025-01-01T00:00:00",
                }
            },
        })
        export_dir = tmp_path / "export"
        result = run(_make_plan_dict(), state, export_dir=str(export_dir), timestamped_export=False)
        files = result.get("output_files", {})
        assert "calendars_html" in files
        assert Path(files["calendars_html"]).exists()

    def test_calendars_html_absent_and_nav_link_hidden_when_no_scrape_cache(self, tmp_path):
        """When no scrape cache exists, calendars.html should not be generated and the navbar link should be hidden."""
        state = PipelineState(tmp_path / "pipeline")
        export_dir = tmp_path / "export"
        result = run(_make_plan_dict(), state, export_dir=str(export_dir), timestamped_export=False)
        files = result.get("output_files", {})
        assert "calendars_html" not in files
        calendars_path = export_dir / "calendars.html"
        assert not calendars_path.exists()
        html_path = Path(files["html"])
        html = html_path.read_text(encoding="utf-8")
        assert "Skrapede kalendere" not in html

    def test_conclusion_injects_blocked_count(self, tmp_path):
        """Conclusion must include blocked source count when blocked sources exist."""
        data = _make_plan_dict()
        state = PipelineState(tmp_path / "pipeline")
        # Write scraping stage with blocked sources so stage4 picks them up
        state.write_stage(
            StageName.SCRAPING,
            {"sources": [], "blocked": ["Ringerike", "Tønsberg"]},
        )
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        assert "2 kilde(r) blokkert." in report_html

    def test_conclusion_injects_cancelled_count(self, tmp_path):
        """Conclusion must include cancellation count when cancelled tournaments exist."""
        data = _make_plan_dict()
        # Add a cancelled tournament
        data["plan"]["tournaments"].append({
            "date": "2025-11-08",
            "arena": "Kongsberghallen",
            "age_group": "U10",
            "host_club": "Kongsberg",
            "teams": [],
            "games": [],
            "cancelled": True,
        })
        state = PipelineState(tmp_path / "pipeline")
        result = run(data, state, export_dir=str(tmp_path / "export"), timestamped_export=False)
        report_html = Path(result["output_files"]["html_report"]).read_text(encoding="utf-8")
        assert "1 turnering(er) avlyst." in report_html
