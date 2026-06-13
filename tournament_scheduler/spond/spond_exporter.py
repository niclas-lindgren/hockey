"""Spond Excel exporter — produces a single-sheet workbook for Spond's Season Planner import.

The workbook keeps Spond's expected core columns (Dato/Aktivitet/Sted/
Start/Slutt) but adds filter-friendly tournament metadata so organizers can
sort and filter the sheet before importing.

Two output modes are supported:

* **tournament-level** (default) — one row per tournament: ``"U10 Turnering — Jarhallen"``
* **game-level** — one row per internal game: ``"U10: Jar 1 vs Jar 2"``

The resulting ``.xlsx`` file can be imported directly into Spond Club's
Season Planner (Sesongplanlegger → Importer fra Excel).
"""

from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

from ..models import SeasonPlan, Tournament

console = Console()

_SPOND_HEADERS = [
    "Dato",
    "Aktivitet",
    "Sted",
    "Start",
    "Slutt",
    "Aldersgruppe",
    "Vertsklubb",
    "Deltakende klubber",
    "Deltakende lag",
    "Import scope",
]

_SCHEDULE_HEADERS = ["Runde", "Hjemmelag", "Bortelag", "Parallellbane"]


class SpondExporter:
    """Export a :class:`~tournament_scheduler.models.SeasonPlan` to Spond's Excel-import format.

    Parameters
    ----------
    game_level:
        If ``True``, each row is a single game (``"U10: Jar 1 vs Jar 2"``).
        If ``False`` (default), each row is a tournament summary
        (``"U10 Turnering — Jarhallen"``).
    """

    def __init__(self, *, game_level: bool = False) -> None:
        self.game_level = game_level

    def export(
        self,
        plan: SeasonPlan,
        output_path: str,
        *,
        club: str | None = None,
        round_length_for_age_group: Optional[dict[str, int]] = None,
    ) -> str:
        """Build and save a Spond-compatible Excel workbook to *output_path*."""
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Sesongplan"

        self._write_sheet(
            sheet,
            plan,
            club=club,
            round_length_for_age_group=round_length_for_age_group,
        )
        self._style_header_row(sheet, 1)
        self._configure_sheet(sheet)
        self._autosize_columns(sheet)

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
        console.print(f"[green]Spond-eksport lagret til[/green] [bold]{out}[/bold]")
        return str(out)

    def export_for_clubs(
        self,
        plan: SeasonPlan,
        output_dir: str | os.PathLike[str],
        *,
        basename: str = "season_plan_spond",
        clubs: Iterable[str] | None = None,
        round_length_for_age_group: Optional[dict[str, int]] = None,
    ) -> dict[str, str]:
        """Write one prefiltered workbook per club and return club -> path."""
        out_dir = Path(output_dir)
        out_dir.mkdir(parents=True, exist_ok=True)
        club_list = list(clubs) if clubs is not None else sorted(
            {team.club for tournament in plan.tournaments for team in tournament.teams}
        )

        written: dict[str, str] = {}
        for club in club_list:
            slug = self._slugify(club)
            path = out_dir / f"{basename}_{slug}.xlsx"
            written[club] = self.export(
                plan,
                str(path),
                club=club,
                round_length_for_age_group=round_length_for_age_group,
            )
        return written

    def export_schedule_attachment(
        self,
        plan: SeasonPlan,
        output_path: str,
        *,
        club: str | None = None,
        round_length_for_age_group: Optional[dict[str, int]] = None,
    ) -> str:
        """Build a printable workbook with one game-schedule sheet per tournament."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        round_length_for_age_group = round_length_for_age_group or {}
        used_titles: set[str] = set()
        tournaments = self._tournaments_for_club(plan, club)

        if not tournaments:
            sheet = wb.create_sheet(title="Kamper")
            sheet.append(["Ingen turneringer å vise"])
            self._style_title_row(sheet, 1)
            self._configure_attachment_sheet(sheet)
            self._autosize_columns(sheet)
        else:
            for index, tournament in enumerate(tournaments, start=1):
                sheet = wb.create_sheet(
                    title=self._unique_attachment_sheet_title(tournament, index, used_titles)
                )
                used_titles.add(sheet.title)
                self._write_schedule_attachment_sheet(
                    sheet,
                    tournament,
                    round_length_for_age_group,
                )

        out = Path(output_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(out))
        console.print(f"[green]Spond-kampoppsett lagret til[/green] [bold]{out}[/bold]")
        return str(out)

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _write_sheet(
        self,
        sheet: Worksheet,
        plan: SeasonPlan,
        *,
        club: str | None = None,
        round_length_for_age_group: Optional[dict[str, int]] = None,
    ) -> None:
        round_length_for_age_group = round_length_for_age_group or {}
        sheet.append(_SPOND_HEADERS)

        for tournament in sorted(plan.tournaments, key=lambda t: t.date):
            if club and not any(team.club == club for team in tournament.teams):
                continue

            if self.game_level:
                for row in self._game_rows_for_tournament(tournament, round_length_for_age_group):
                    sheet.append(row)
            else:
                sheet.append(self._summary_row_for_tournament(tournament, round_length_for_age_group))

        if sheet.max_row:
            end_ref = sheet.cell(row=sheet.max_row, column=sheet.max_column).coordinate
            sheet.auto_filter.ref = f"A1:{end_ref}"
        sheet.freeze_panes = "A2"

    def _write_schedule_attachment_sheet(
        self,
        sheet: Worksheet,
        tournament: Tournament,
        round_length_for_age_group: dict[str, int],
    ) -> None:
        date_str = tournament.date.strftime("%d.%m.%Y")
        title = f"{date_str} ({self._weekday_name(tournament.date)}) — {tournament.age_group} — {tournament.arena}"
        if tournament.cancelled:
            title = f"(AVLYST) {title}"

        sheet.append([title])
        self._style_title_row(sheet, 1)

        if tournament.cancelled:
            reason = tournament.cancellation_reason or "ingen grunn oppgitt"
            sheet.append([f"AVLYST: {reason}"])
            self._style_title_row(sheet, 2)

        sheet.append([f"Vertsklubb: {tournament.host_club or ''}"])
        sheet.append([f"Deltakende lag: {', '.join(team.label for team in tournament.teams)}"])

        time_bits = []
        if tournament.start_time:
            time_bits.append(f"Start: {tournament.start_time}")
            round_length = round_length_for_age_group.get(tournament.age_group)
            if round_length:
                end_time = tournament.end_time(round_length)
                if end_time:
                    time_bits.append(f"Slutt: {end_time}")
        if time_bits:
            sheet.append([" • ".join(time_bits)])

        sheet.append([])
        sheet.append(_SCHEDULE_HEADERS)
        self._style_header_row(sheet, sheet.max_row)

        for game in tournament.games:
            sheet.append([
                game.round_number,
                game.home.label,
                game.away.label,
                game.parallel_slot + 1,
            ])

        if not tournament.games:
            sheet.append(["-", "Ingen kamper", "", ""])

        bye_rounds = tournament.get_bye_rounds()
        if bye_rounds:
            for round_num in sorted(bye_rounds):
                for team_label in bye_rounds[round_num]:
                    sheet.append([round_num, "(Pause)", team_label, ""])

        self._configure_attachment_sheet(sheet)
        self._autosize_columns(sheet)

    def _summary_row_for_tournament(
        self,
        tournament: Tournament,
        round_length_for_age_group: dict[str, int],
    ) -> list[str]:
        date_str = tournament.date.strftime("%d.%m.%Y")
        arena = tournament.arena
        age_group = tournament.age_group
        clubs = self._join_unique(team.club for team in tournament.teams)
        teams = ", ".join(team.label for team in tournament.teams)
        start_time = tournament.start_time or ""
        end_time = ""
        if tournament.start_time:
            round_length = round_length_for_age_group.get(tournament.age_group)
            if round_length:
                end_time = tournament.end_time(round_length) or ""

        activity = f"{age_group} Turnering — {arena}"
        if tournament.cancelled:
            activity = f"AVLYST: {activity}"

        return [
            date_str,
            activity,
            arena,
            start_time,
            end_time,
            age_group,
            tournament.host_club or "",
            clubs,
            teams,
            "turnering",
        ]

    def _game_rows_for_tournament(
        self,
        tournament: Tournament,
        round_length_for_age_group: dict[str, int],
    ) -> list[list[str]]:
        rows: list[list[str]] = []
        date_str = tournament.date.strftime("%d.%m.%Y")
        arena = tournament.arena
        age_group = tournament.age_group
        clubs = self._join_unique(team.club for team in tournament.teams)
        start_time = tournament.start_time or ""
        end_time = ""
        if tournament.start_time:
            round_length = round_length_for_age_group.get(tournament.age_group)
            if round_length:
                end_time = tournament.end_time(round_length) or ""

        for game in tournament.games:
            activity = f"{age_group}: {game.home.label} vs {game.away.label}"
            if tournament.cancelled:
                activity = f"AVLYST: {activity}"
            rows.append([
                date_str,
                activity,
                arena,
                start_time,
                end_time,
                age_group,
                tournament.host_club or "",
                clubs,
                f"{game.home.label}, {game.away.label}",
                "kamp",
            ])

        if not rows:
            rows.append(self._summary_row_for_tournament(tournament, round_length_for_age_group))
        return rows

    @staticmethod
    def _tournaments_for_club(plan: SeasonPlan, club: str | None) -> list[Tournament]:
        tournaments = sorted(plan.tournaments, key=lambda t: t.date)
        if not club:
            return tournaments
        return [t for t in tournaments if any(team.club == club for team in t.teams)]

    @staticmethod
    def _join_unique(values: Iterable[str]) -> str:
        seen: list[str] = []
        for value in values:
            if value and value not in seen:
                seen.append(value)
        return ", ".join(seen)

    @staticmethod
    def _slugify(value: str) -> str:
        slug = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
        return slug or "club"

    @staticmethod
    def _weekday_name(value) -> str:
        weekdays = ["mandag", "tirsdag", "onsdag", "torsdag", "fredag", "lørdag", "søndag"]
        return weekdays[value.weekday()]

    @staticmethod
    def _style_title_row(sheet: Worksheet, row_number: int) -> None:
        for cell in sheet[row_number]:
            cell.font = cell.font.copy(bold=True)

    @staticmethod
    def _style_header_row(sheet: Worksheet, row_number: int = 1) -> None:
        for cell in sheet[row_number]:
            cell.font = cell.font.copy(bold=True)

    @staticmethod
    def _configure_sheet(sheet: Worksheet) -> None:
        sheet.sheet_view.showGridLines = False

    @staticmethod
    def _configure_attachment_sheet(sheet: Worksheet) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A6"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    @staticmethod
    def _autosize_columns(sheet: Worksheet, max_width: int = 60) -> None:
        widths: dict[str, int] = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col_letter = cell.column_letter
                width = len(str(cell.value))
                widths[col_letter] = max(widths.get(col_letter, 0), width)
        for col_letter, width in widths.items():
            sheet.column_dimensions[col_letter].width = min(width + 2, max_width)

    @staticmethod
    def _unique_attachment_sheet_title(tournament: Tournament, index: int, used_titles) -> str:
        date_part = tournament.date.strftime("%d.%m")
        base = f"{date_part} {tournament.age_group} {tournament.arena}".strip()
        return SpondExporter._unique_sheet_title_from_base(base, index, used_titles)

    @staticmethod
    def _unique_sheet_title_from_base(base: str, index: int, used_titles) -> str:
        base = SpondExporter._sanitize_sheet_title(base)
        title = base[:31]
        if title not in used_titles:
            return title
        while True:
            suffix = f" ({index})"
            trimmed_len = 31 - len(suffix)
            title = f"{base[:trimmed_len]}{suffix}"
            if title not in used_titles:
                return title
            index += 1

    @staticmethod
    def _sanitize_sheet_title(title: str) -> str:
        invalid_chars = set('[]:*?/\\')
        return "".join(ch for ch in title if ch not in invalid_chars).strip()
