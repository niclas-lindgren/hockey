"""Excel exporter for season plans.

`SeasonPlanExporter` is a write-side sibling to the read-only
`ExcelTournamentReader` in `tournament_scheduler.excel.tournament_reader`. It
builds a workbook with:

  (a) a season-overview sheet — one row per tournament: date, weekday,
      age group/gender, arena/venue, and the participating teams; and
  (b) a per-tournament sheet (one sheet per tournament, falling back to a
      grouped block on a single sheet if the plan is too large for Excel's
      31-character / unique-name sheet limits) listing that tournament's
      full round-robin game schedule (home/away teams + parallel slot).

The resulting workbook is saved to a user-specified `.xlsx` path.
"""

from datetime import date
from typing import Dict, List, Optional

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

from tournament_scheduler.club_distances import furthest_traveling_team
from tournament_scheduler.fairness_model import SeasonFairnessModel
from tournament_scheduler.models import SeasonPlan, Tournament

console = Console()


# Norwegian weekday names (Monday=0 .. Sunday=6), matching the project's
# Norwegian-language interactive CLI conventions.
_NORWEGIAN_WEEKDAYS = [
    "mandag", "tirsdag", "onsdag", "torsdag", "fredag", "lørdag", "søndag",
]

_OVERVIEW_HEADERS = ["Dato", "Ukedag", "Aldersgruppe", "Arena", "Vertsklubb", "Lag", "Lengst anslått reise", "Starttid", "Sluttid"]
_GAMES_HEADERS = ["Runde", "Hjemmelag", "Bortelag", "Parallellbane"]
_CLUB_SUMMARY_HEADERS = ["Lag", "Aldersgruppe", "Dato", "Ukedag", "Motstander(e)", "Vertsarena"]
_RULES_HEADERS = ["Regel", "Forklaring", "Kategori"]
_FAIRNESS_HEADERS = ["Metrikk", "Verdi", "Terskel", "Score", "Status", "Detalj"]
_FAIRNESS_ADJUSTMENT_HEADERS = ["Lag", "Klubb", "Aldersgruppe", "Faktisk", "Mål", "Justering", "Status", "Kommentar"]

# Club worksheet titles are prefixed to distinguish them from tournament
# sheets and to keep them grouped together when sorted alphabetically.
_CLUB_SHEET_PREFIX = "Klubb "

# Excel worksheet titles cannot exceed 31 characters and must be unique.
_MAX_SHEET_TITLE_LENGTH = 31


class SeasonPlanExporter:
    """Writes a `SeasonPlan` to an Excel workbook (season overview + per-tournament games)."""

    def __init__(self):
        self.workbook: Optional[openpyxl.Workbook] = None

    def export(
        self,
        plan: SeasonPlan,
        output_path: str,
        *,
        rules_report: Optional[List[Dict[str, str]]] = None,
        round_length_for_age_group: Optional[Dict[str, int]] = None,
    ) -> str:
        """Build and save the workbook for `plan` to `output_path`.

        Optionally includes a ``Regler og avgjørelser`` sheet when
        ``rules_report`` (from ``SeasonPlanner.rules_report()``) is
        provided.

        ``round_length_for_age_group`` (optional mapping of age group ->
        round length in minutes) is used to compute each tournament's end
        time for the "Sluttid" overview column via
        ``Tournament.end_time()``.

        Returns the path the workbook was saved to.
        """
        self.workbook = openpyxl.Workbook()

        overview_sheet = self.workbook.active
        overview_sheet.title = "Sesongoversikt"
        self._write_overview_sheet(overview_sheet, plan, round_length_for_age_group or {})

        used_titles = {overview_sheet.title}

        if rules_report:
            rules_sheet = self.workbook.create_sheet(title="Regler og avgjørelser")
            used_titles.add(rules_sheet.title)
            self._write_rules_sheet(rules_sheet, rules_report)

        if plan.fairness_gate:
            fairness_sheet = self.workbook.create_sheet(title="Rettferdighetskontroll")
            used_titles.add(fairness_sheet.title)
            self._write_fairness_sheet(fairness_sheet, plan.fairness_gate)

        fairness_adjustments_sheet = self.workbook.create_sheet(title="Rettferdighetsjusteringer")
        used_titles.add(fairness_adjustments_sheet.title)
        self._write_fairness_adjustments_sheet(fairness_adjustments_sheet, plan)

        for index, tournament in enumerate(plan.tournaments, start=1):
            sheet = self.workbook.create_sheet(title=self._unique_sheet_title(tournament, index, used_titles))
            used_titles.add(sheet.title)
            self._write_tournament_sheet(sheet, tournament)

        for club_index, club in enumerate(self._clubs_in_plan(plan), start=1):
            sheet = self.workbook.create_sheet(title=self._unique_club_sheet_title(club, club_index, used_titles))
            used_titles.add(sheet.title)
            self._write_club_summary_sheet(sheet, plan, club)

        self.workbook.save(output_path)
        console.print(f"[green]Sesongplanen ble eksportert til[/green] [bold]{output_path}[/bold]")
        return output_path

    # ------------------------------------------------------------------
    # Season-overview sheet
    # ------------------------------------------------------------------

    def _write_overview_sheet(
        self,
        sheet: Worksheet,
        plan: SeasonPlan,
        round_length_for_age_group: Optional[Dict[str, int]] = None,
    ) -> None:
        sheet.append(_OVERVIEW_HEADERS)
        self._style_header_row(sheet)

        round_length_for_age_group = round_length_for_age_group or {}

        _cancelled_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        for tournament in plan.tournaments:
            travel_info = self._travel_info(tournament)
            status_prefix = "(AVLYST) " if tournament.cancelled else ""
            round_length = round_length_for_age_group.get(tournament.age_group)
            end_time = tournament.end_time(round_length) if round_length else None
            row = [
                status_prefix + self._format_date(tournament.date),
                self._weekday_name(tournament.date),
                tournament.age_group,
                tournament.arena,
                tournament.host_club or "",
                ", ".join(team.label for team in tournament.teams),
                travel_info,
                tournament.start_time or "",
                end_time or "",
            ]
            sheet.append(row)
            if tournament.cancelled:
                for cell in sheet[sheet.max_row]:
                    cell.fill = _cancelled_fill

        self._autosize_columns(sheet)

    # ------------------------------------------------------------------
    # Rules-and-decisions sheet
    # ------------------------------------------------------------------

    def _write_rules_sheet(self, sheet: Worksheet, rules_report: List[Dict[str, str]]) -> None:
        """Write the rules-and-decisions transparency report.

        Expected keys per rule dict: ``regel``, ``forklaring``, ``kategori``.
        """
        sheet.append(_RULES_HEADERS)
        self._style_header_row(sheet)

        for rule in rules_report:
            sheet.append([
                rule.get("regel", ""),
                rule.get("forklaring", ""),
                rule.get("kategori", ""),
            ])

        self._autosize_columns(sheet)

    # ------------------------------------------------------------------
    # Fairness gate sheet
    # ------------------------------------------------------------------

    def _write_fairness_sheet(self, sheet: Worksheet, fairness_gate: dict) -> None:
        sheet.append(["Overordnet status", fairness_gate.get("status", ""), "", fairness_gate.get("score", ""), "", ""])
        sheet.append([])
        sheet.append(_FAIRNESS_HEADERS)
        header_row_index = sheet.max_row
        for cell in sheet[header_row_index]:
            cell.font = cell.font.copy(bold=True)

        _status_fills = {
            "pass": PatternFill(start_color="D9F99D", end_color="D9F99D", fill_type="solid"),
            "warn": PatternFill(start_color="FDE68A", end_color="FDE68A", fill_type="solid"),
            "fail": PatternFill(start_color="FCA5A5", end_color="FCA5A5", fill_type="solid"),
        }

        metrics = fairness_gate.get("metrics", []) if isinstance(fairness_gate, dict) else []
        for metric in metrics:
            unit = metric.get("unit", "")
            value = metric.get("value", "")
            threshold = metric.get("threshold", "")
            if unit and value != "":
                value = f"{value} {unit}"
            if unit and threshold != "":
                threshold = f"{threshold} {unit}"
            sheet.append([
                metric.get("label", ""),
                value,
                threshold,
                metric.get("score", ""),
                str(metric.get("status", "")).upper(),
                metric.get("detail", ""),
            ])
            status = str(metric.get("status", "")).lower()
            fill = _status_fills.get(status)
            if fill:
                for cell in sheet[sheet.max_row]:
                    cell.fill = fill

        self._autosize_columns(sheet)

    def _write_fairness_adjustments_sheet(self, sheet: Worksheet, plan: SeasonPlan) -> None:
        model = SeasonFairnessModel()
        rows = model.adjustment_rows_for_plan(plan)

        sheet.append(["Rettferdighetsjusteringer per lag"])
        sheet.append(["Positiv rettferdighetsjustering betyr at laget ligger under rettferdighetsmålet og kan få flere kamper."])
        sheet.append([])
        sheet.append(_FAIRNESS_ADJUSTMENT_HEADERS)
        header_row_index = sheet.max_row
        for cell in sheet[header_row_index]:
            cell.font = cell.font.copy(bold=True)

        fills = {
            "under": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            "over": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            "on_target": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        }
        for row in rows:
            adjustment = float(row.get("adjustment", 0.0))
            comment = (
                f"Mangler {adjustment:.1f} kamper til mål" if adjustment > 0.5 else
                (f"{abs(adjustment):.1f} kamper over mål" if adjustment < -0.5 else "På mål")
            )
            sheet.append([
                row.get("label", ""),
                row.get("club", ""),
                row.get("age_group", ""),
                row.get("actual", 0),
                round(float(row.get("target", 0.0)), 3),
                round(adjustment, 3),
                str(row.get("status", "")).replace("on_target", "på mål").upper(),
                comment,
            ])
            fill = fills.get(str(row.get("status", "")))
            if fill:
                for cell in sheet[sheet.max_row]:
                    cell.fill = fill

        self._autosize_columns(sheet)

    # ------------------------------------------------------------------
    # Per-tournament game-schedule sheets
    # ------------------------------------------------------------------

    def _write_tournament_sheet(self, sheet: Worksheet, tournament: Tournament) -> None:
        _cancelled_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        title_row = (
            f"{self._format_date(tournament.date)} ({self._weekday_name(tournament.date)}) — "
            f"{tournament.age_group} — {tournament.arena}"
        )
        if tournament.cancelled:
            title_row = f"(AVLYST) {title_row}"
        sheet.append([title_row])
        if tournament.cancelled:
            reason = tournament.cancellation_reason or "ingen grunn oppgitt"
            sheet.append([f"AVLYST: {reason}"])
        sheet.append([])
        sheet.append([f"Deltakende lag: {', '.join(team.label for team in tournament.teams)}"])
        travel_info = self._travel_info(tournament)
        if travel_info:
            sheet.append([f"Lengst anslått reise: {travel_info}"])
        sheet.append([])

        sheet.append(_GAMES_HEADERS)
        header_row_index = sheet.max_row
        for cell in sheet[header_row_index]:
            cell.font = cell.font.copy(bold=True)

        for game in tournament.games:
            sheet.append([
                game.round_number,
                game.home.label,
                game.away.label,
                game.parallel_slot + 1,  # 1-based for human-friendly display
            ])

        # Bye rows for odd-numbered tournaments: one team sits out each round.
        bye_rounds = tournament.get_bye_rounds()
        if bye_rounds:
            for round_num in sorted(bye_rounds):
                for team_label in bye_rounds[round_num]:
                    sheet.append([round_num, "(Pause)", team_label, ""])

        self._autosize_columns(sheet)

    # ------------------------------------------------------------------
    # Per-club summary sheets
    # ------------------------------------------------------------------

    def _write_club_summary_sheet(self, sheet: Worksheet, plan: SeasonPlan, club: str) -> None:
        """Write one row per (team, tournament-date, opponent, arena) for `club`.

        Iterates `plan.tournaments`, finds the games involving teams that
        belong to `club`, and renders a row for each such game from that
        team's perspective (its label, the tournament's age group/date, the
        opponent's label, and the host arena).
        """
        sheet.append([f"Klubb: {club}"])
        sheet.append([])
        sheet.append(_CLUB_SUMMARY_HEADERS)
        header_row_index = sheet.max_row
        for cell in sheet[header_row_index]:
            cell.font = cell.font.copy(bold=True)

        for tournament in plan.tournaments:
            club_teams = [team for team in tournament.teams if team.club == club]
            if not club_teams:
                continue

            for team in club_teams:
                opponents = [
                    (game.away if game.home is team else game.home).label
                    for game in tournament.games
                    if game.home is team or game.away is team
                ]
                sheet.append([
                    team.label,
                    tournament.age_group,
                    self._format_date(tournament.date),
                    self._weekday_name(tournament.date),
                    ", ".join(opponents),
                    tournament.arena,
                ])

        self._autosize_columns(sheet)

    @staticmethod
    def _clubs_in_plan(plan: SeasonPlan) -> List[str]:
        """Return distinct club names across all tournaments' teams, in first-seen order."""
        seen: List[str] = []
        for tournament in plan.tournaments:
            for team in tournament.teams:
                if team.club not in seen:
                    seen.append(team.club)
        return seen

    @staticmethod
    def _unique_club_sheet_title(club: str, index: int, used_titles) -> str:
        """Build a unique, Excel-valid (<=31 chars) sheet title for a club summary.

        Format: "<_CLUB_SHEET_PREFIX><club>" truncated as needed, with a
        numeric suffix appended on collision — mirrors `_unique_sheet_title`.
        """
        base = f"{_CLUB_SHEET_PREFIX}{club}"
        return SeasonPlanExporter._unique_sheet_title_from_base(base, index, used_titles)

    # ------------------------------------------------------------------
    # Travel-time display helper
    # ------------------------------------------------------------------

    @staticmethod
    def _travel_info(tournament: Tournament) -> str:
        """Return a short Norwegian-language travel-info string for *tournament*.

        Example: "Jar 1 (~80 km)" or empty string when no travel data exists.
        """
        result = furthest_traveling_team(tournament)
        if result is None:
            return ""
        team, km = result
        return f"{team.label} (~{km} km)"

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _format_date(value: date) -> str:
        return value.strftime("%d.%m.%Y")

    @staticmethod
    def _weekday_name(value: date) -> str:
        return _NORWEGIAN_WEEKDAYS[value.weekday()]

    @staticmethod
    def _style_header_row(sheet: Worksheet) -> None:
        for cell in sheet[1]:
            cell.font = cell.font.copy(bold=True)

    @staticmethod
    def _autosize_columns(sheet: Worksheet, max_width: int = 60) -> None:
        """Roughly size each column to fit its widest cell (capped at `max_width`)."""
        widths = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                column_letter = cell.column_letter
                length = len(str(cell.value))
                widths[column_letter] = max(widths.get(column_letter, 0), length)

        for column_letter, width in widths.items():
            sheet.column_dimensions[column_letter].width = min(width + 2, max_width)

    @staticmethod
    def _unique_sheet_title(tournament: Tournament, index: int, used_titles) -> str:
        """Build a unique, Excel-valid (<=31 chars) sheet title for a tournament.

        Format: "<DD.MM> <age_group> <arena>" truncated as needed, with a
        numeric suffix appended if a collision would otherwise occur.
        """
        date_part = tournament.date.strftime("%d.%m")
        base = f"{date_part} {tournament.age_group} {tournament.arena}".strip()
        return SeasonPlanExporter._unique_sheet_title_from_base(base, index, used_titles)

    @staticmethod
    def _unique_sheet_title_from_base(base: str, index: int, used_titles) -> str:
        """Sanitize `base` into a unique, Excel-valid (<=31 chars) sheet title.

        Shared by tournament and club sheet naming: truncates `base` to
        Excel's sheet-title limits and, on collision with `used_titles`,
        appends a numeric suffix (` (<index>)`), trimming the base as needed
        to keep the result within `_MAX_SHEET_TITLE_LENGTH`.
        """
        base = SeasonPlanExporter._sanitize_sheet_title(base)

        title = base[:_MAX_SHEET_TITLE_LENGTH]
        if title not in used_titles:
            return title

        # Collision — append a numeric suffix, trimming the base as needed,
        # bumping the index until the result is unique.
        while True:
            suffix = f" ({index})"
            trimmed_len = _MAX_SHEET_TITLE_LENGTH - len(suffix)
            title = f"{base[:trimmed_len]}{suffix}"
            if title not in used_titles:
                return title
            index += 1

    @staticmethod
    def _sanitize_sheet_title(title: str) -> str:
        """Strip characters that Excel does not allow in sheet titles."""
        invalid_chars = set('[]:*?/\\')
        return "".join(ch for ch in title if ch not in invalid_chars).strip()
