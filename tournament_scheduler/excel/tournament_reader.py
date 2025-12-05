"""Excel tournament reader for extracting tournament and team information."""

import sys
from datetime import datetime, date
from typing import Set, List, Optional
import openpyxl
from tournament_scheduler.models import TournamentInfo
from tournament_scheduler.utils.date_parser import DateParser


class ExcelTournamentReader:
    """Reads tournament and team data from Excel files."""

    def __init__(self, file_path: str, date_parser: DateParser):
        """Initialize Excel reader.

        Args:
            file_path: Path to Excel file
            date_parser: DateParser instance for consistent date parsing
        """
        self.file_path = file_path
        self.date_parser = date_parser
        self.workbook = None
        self.worksheet = None

    def _load_workbook(self):
        """Load workbook if not already loaded."""
        if self.workbook is None:
            try:
                self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
                self.worksheet = self.workbook.active
            except FileNotFoundError:
                print(f"Error: Excel file not found: {self.file_path}", file=sys.stderr)
                sys.exit(1)
            except Exception as e:
                print(f"Error: Failed to open Excel file: {e}", file=sys.stderr)
                sys.exit(1)

    def get_all_tournament_dates(self) -> Set[date]:
        """Get all tournament dates from Excel file.

        Returns:
            Set of dates containing tournaments
        """
        self._load_workbook()
        dates = set()

        for row in self.worksheet.iter_rows(min_row=1, values_only=True):
            for cell in row:
                parsed = self.date_parser.parse_datetime_cell(cell)
                if parsed:
                    dates.add(parsed.date())

        return dates

    def get_tournament_info(self, tournament_date: date) -> TournamentInfo:
        """Get tournament information for a specific date.

        Args:
            tournament_date: Date to find tournament for

        Returns:
            TournamentInfo with teams and details

        Raises:
            SystemExit: If no tournament found on the specified date
        """
        self._load_workbook()
        teams = self.extract_teams_for_date(tournament_date)
        location = self._extract_location_for_date(tournament_date)

        if not teams:
            available_dates = self.get_all_tournament_dates()
            print(f"\nError: No tournament found on {tournament_date.strftime('%Y-%m-%d')} in Excel file", file=sys.stderr)
            if available_dates:
                print("\nAvailable tournament dates in file:", file=sys.stderr)
                for d in sorted(available_dates)[:10]:
                    print(f"  - {d.strftime('%Y-%m-%d')}", file=sys.stderr)
                if len(available_dates) > 10:
                    print(f"  ... and {len(available_dates) - 10} more dates", file=sys.stderr)
            sys.exit(1)

        # Debug output
        print(f"\n{'='*60}")
        print(f"TOURNAMENT DETAILS FOR {tournament_date.strftime('%Y-%m-%d')}")
        print(f"{'='*60}")
        if location:
            print(f"Location: {location}")
        print(f"Teams found: {len(teams)}")
        for i, team in enumerate(teams, 1):
            print(f"  {i}. {team}")
        print(f"{'='*60}\n")

        return TournamentInfo(
            date=tournament_date,
            name=f"Tournament on {tournament_date.strftime('%Y-%m-%d')}",
            teams=teams,
            location=location
        )

    def extract_teams_for_date(self, tournament_date: date) -> List[str]:
        """Extract all teams participating in tournament on specified date.

        Args:
            tournament_date: Date to extract teams for

        Returns:
            List of team names
        """
        self._load_workbook()
        teams = []
        found_date_row = None

        # Find the row containing the tournament date
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=1, values_only=True), start=1):
            for cell in row:
                parsed = self.date_parser.parse_datetime_cell(cell)
                if parsed and parsed.date() == tournament_date:
                    found_date_row = row_idx
                    break
            if found_date_row:
                break

        if not found_date_row:
            return teams

        # Look for "Deltagende lag:" or team names in subsequent rows
        # Team names typically appear after the date row
        max_rows_to_check = 30
        for row_idx in range(found_date_row, min(found_date_row + max_rows_to_check, self.worksheet.max_row + 1)):
            row = list(self.worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]

            # Stop if we hit another date (next tournament)
            for cell in row:
                parsed = self.date_parser.parse_datetime_cell(cell)
                if parsed and parsed.date() != tournament_date:
                    return teams

            # Check for start of next tournament section
            for cell in row:
                if cell and isinstance(cell, str):
                    cell_lower = cell.lower()
                    # Stop if we hit headers for next tournament
                    if 'turnering nr' in cell_lower or cell_lower == 'arrangør':
                        return teams

            # Extract team names
            for cell in row:
                if cell and isinstance(cell, str):
                    cell_str = str(cell).strip()

                    # Skip header rows and common non-team entries
                    skip_patterns = [
                        'deltagende lag', 'runde nr', 'kampstart', 'kamp nr',
                        'bane nr', 'hjemmelag', 'bortelag', 'dato', 'kl',
                        'dommerkommisjon', 'lederstevne', 'møtetid', 'kamp',
                        'turnering', 'serie', 'arrangør'
                    ]

                    if any(pattern in cell_str.lower() for pattern in skip_patterns):
                        continue

                    # Skip entries that look like "Team - Venue" format
                    # (e.g., "Frisk Asker - Askerhallen")
                    if ' - ' in cell_str:
                        parts = cell_str.split(' - ')
                        if len(parts) == 2 and any(keyword in parts[1].lower() for keyword in ['hall', 'arena', 'forum']):
                            continue

                    # Skip empty strings and numbers
                    if not cell_str or cell_str.isdigit():
                        continue

                    # Skip single characters or very short strings
                    if len(cell_str) < 3:
                        continue

                    # Skip time patterns (HH:MM)
                    if ':' in cell_str and len(cell_str) <= 5:
                        continue

                    # Check if this looks like a team name
                    # Team names often contain letters and may have numbers
                    if any(c.isalpha() for c in cell_str):
                        # Check if it's not already in the list
                        if cell_str not in teams:
                            teams.append(cell_str)

        return teams

    def _extract_location_for_date(self, tournament_date: date) -> Optional[str]:
        """Extract location information for a tournament on specified date.

        Args:
            tournament_date: Date to extract location for

        Returns:
            Location string if found, None otherwise
        """
        found_date_row = None

        # Find the row containing the tournament date
        for row_idx, row in enumerate(self.worksheet.iter_rows(min_row=1, values_only=True), start=1):
            for cell in row:
                parsed = self.date_parser.parse_datetime_cell(cell)
                if parsed and parsed.date() == tournament_date:
                    found_date_row = row_idx
                    break
            if found_date_row:
                break

        if not found_date_row:
            return None

        # Look in nearby rows for location keywords
        for row_idx in range(max(1, found_date_row - 5), min(found_date_row + 10, self.worksheet.max_row + 1)):
            row = list(self.worksheet.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
            for cell in row:
                if cell and isinstance(cell, str):
                    cell_lower = cell.lower()
                    # Common location keywords
                    if any(keyword in cell_lower for keyword in ['hall', 'arena', 'forum', 'isforum', 'ishall']):
                        return cell.strip()

        return None
