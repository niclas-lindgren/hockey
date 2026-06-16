"""Per-club review packets for the finalized season plan."""

from __future__ import annotations

import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console

from ..club_distances import compute_team_travel_distances
from ..models import SeasonPlan, Tournament
from ..spond.spond_exporter import SpondExporter

console = Console()


@dataclass(frozen=True)
class ReviewPacketPaths:
    """Filesystem paths for one club's review packet."""

    folder: str
    summary_workbook: str
    spond_import_workbook: str
    schedule_attachment: str
    manifest: str
    response_template: str


@dataclass(frozen=True)
class ReviewPacketResponse:
    """Parsed club response from a review packet template."""

    club: str
    decision: str
    requested_changes: dict[str, list[str]]
    comment: str = ""

    @classmethod
    def from_path(cls, path: str | Path) -> "ReviewPacketResponse":
        response_path = Path(path)
        if response_path.is_dir():
            response_path = response_path / "response_template.json"
        data = json.loads(response_path.read_text(encoding="utf-8"))
        requested_raw = data.get("requested_changes", {}) or {}

        def _as_list(values: Any) -> list[str]:
            if values is None:
                return []
            if isinstance(values, str):
                values = [values]
            return [str(value) for value in values if str(value).strip()]

        requested_changes = {
            key: _as_list(values)
            for key, values in requested_raw.items()
        }
        decision = str(data.get("decision", "accept") or "accept").strip().lower()
        if decision not in {"accept", "change_request"}:
            decision = "accept"
        if decision == "accept" and any(requested_changes.values()):
            decision = "change_request"
        return cls(
            club=str(data.get("club", "") or "").strip(),
            decision=decision,
            requested_changes=requested_changes,
            comment=str(data.get("comment", "") or ""),
        )

    @property
    def is_change_request(self) -> bool:
        return self.decision == "change_request"

    def as_manual_adjustments(self) -> dict[str, list[str]]:
        aliases = {
            "lock_dates": "locked_dates",
            "ban_dates": "banned_dates",
            "pin_tournaments": "pinned_tournament_ids",
            "pinned_tournament_ids": "pinned_tournament_ids",
            "force_host_clubs": "forced_host_clubs",
            "forced_host_clubs": "forced_host_clubs",
            "exclude_host_clubs": "excluded_host_clubs",
            "excluded_host_clubs": "excluded_host_clubs",
        }
        merged: dict[str, list[str]] = {
            "locked_dates": [],
            "banned_dates": [],
            "pinned_tournament_ids": [],
            "forced_host_clubs": [],
            "excluded_host_clubs": [],
        }
        for key, values in self.requested_changes.items():
            target = aliases.get(key, key)
            if target in merged:
                merged[target].extend(str(value) for value in values if str(value).strip())
        return merged


class ReviewPacketExporter:
    """Write per-club review/approval packets for a season plan."""

    def export(
        self,
        plan: SeasonPlan,
        output_dir: str | Path,
        *,
        clubs: Iterable[str] | None = None,
        round_length_for_age_group: Optional[dict[str, int]] = None,
    ) -> dict[str, str]:
        """Write one packet directory per club and return club -> folder path."""
        root = Path(output_dir)
        root.mkdir(parents=True, exist_ok=True)
        club_list = list(clubs) if clubs is not None else sorted(
            {team.club for tournament in plan.tournaments for team in tournament.teams}
        )

        written: dict[str, str] = {}
        for club in club_list:
            packet_dir = root / self._slugify(club)
            packet_dir.mkdir(parents=True, exist_ok=True)

            summary_workbook = packet_dir / "club_review.xlsx"
            spond_path = packet_dir / "club_review_spond.xlsx"
            schedule_path = packet_dir / "club_review_spond_games.xlsx"
            manifest_path = packet_dir / "manifest.json"
            response_path = packet_dir / "response_template.json"

            club_tournaments = self._club_tournaments(plan, club)
            self._write_summary_workbook(
                summary_workbook,
                plan,
                club,
                club_tournaments,
                round_length_for_age_group or {},
            )

            spond_exporter = SpondExporter()
            spond_exporter.export(
                plan,
                str(spond_path),
                club=club,
                round_length_for_age_group=round_length_for_age_group,
            )
            spond_exporter.export_schedule_attachment(
                plan,
                str(schedule_path),
                club=club,
                round_length_for_age_group=round_length_for_age_group,
            )

            manifest = self._build_manifest(
                plan,
                club,
                packet_dir,
                summary_workbook,
                spond_path,
                schedule_path,
                club_tournaments,
            )
            manifest_path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")
            response_path.write_text(
                json.dumps(self._response_template(club), indent=2, ensure_ascii=False),
                encoding="utf-8",
            )

            written[club] = str(packet_dir)
            console.print(f"[green]Club review packet lagret til[/green] [bold]{packet_dir}[/bold]")

        return written

    def _write_summary_workbook(
        self,
        path: Path,
        plan: SeasonPlan,
        club: str,
        club_tournaments: list[Tournament],
        round_length_for_age_group: dict[str, int],
    ) -> None:
        wb = openpyxl.Workbook()
        overview = wb.active
        overview.title = "Oversikt"
        tournaments_sheet = wb.create_sheet("Turneringer")
        hosting_sheet = wb.create_sheet("Hjemmeturneringer")
        travel_sheet = wb.create_sheet("Reise")

        team_travel = compute_team_travel_distances(plan)
        club_team_labels = sorted(
            {
                team.label
                for tournament in plan.tournaments
                for team in tournament.teams
                if team.club == club
            }
        )
        club_team_travel = {label: team_travel.get(label, 0) for label in club_team_labels}
        hosted = [t for t in club_tournaments if t.host_club == club]
        away_count = sum(1 for t in club_tournaments if t.host_club != club)

        self._write_overview_sheet(
            overview,
            club,
            plan,
            club_tournaments,
            hosted,
            away_count,
            club_team_travel,
            round_length_for_age_group,
        )
        self._write_tournaments_sheet(tournaments_sheet, club_tournaments, round_length_for_age_group)
        self._write_hosting_sheet(hosting_sheet, hosted, round_length_for_age_group)
        self._write_travel_sheet(travel_sheet, club_team_travel)

        for index, sheet in enumerate(wb.worksheets, start=1):
            self._style_header_row(sheet, 1)
            if sheet.max_row:
                sheet.freeze_panes = "A2"
            self._configure_sheet(sheet)
            self._autosize_columns(sheet)

        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(path))

    def _write_overview_sheet(
        self,
        sheet: Worksheet,
        club: str,
        plan: SeasonPlan,
        club_tournaments: list[Tournament],
        hosted: list[Tournament],
        away_count: int,
        club_team_travel: dict[str, int],
        round_length_for_age_group: dict[str, int],
    ) -> None:
        total_travel = sum(club_team_travel.values())
        sheet.append([f"Klubb: {club}"])
        sheet.append([f"Totalt antall relevante turneringer: {len(club_tournaments)}"])
        sheet.append([f"Turneringer klubben arrangerer: {len(hosted)}"])
        sheet.append([f"Turneringer på bortebane: {away_count}"])
        sheet.append([f"Anslått totalt reisebehov (km): {total_travel}"])
        sheet.append([f"Anbefaling: {'Godkjenn' if club_tournaments else 'Ingen turneringer å vurdere'}"])

        # Skipped age groups — groups with <3 teams that were not planned.
        if plan.skipped_age_groups:
            sheet.append([])
            sheet.append(["Hoppet over — aldersgrupper som ikke planlegges:"])
            for entry in plan.skipped_age_groups:
                sheet.append([
                    f"{entry['age_group']}: {entry['reason']}"
                ])
            sheet.append([])

        sheet.append(["Dato", "Aldersgruppe", "Arena", "Vertsklubb", "Start", "Slutt", "Lag", "Kamper"])
        for tournament in club_tournaments:
            start_time = tournament.start_time or ""
            end_time = ""
            if tournament.start_time:
                round_length = round_length_for_age_group.get(tournament.age_group)
                if round_length:
                    end_time = tournament.end_time(round_length) or ""
            sheet.append([
                tournament.date.isoformat(),
                tournament.age_group,
                tournament.arena,
                tournament.host_club or "",
                start_time,
                end_time,
                ", ".join(team.label for team in tournament.teams),
                len(tournament.games),
            ])

    def _write_tournaments_sheet(
        self,
        sheet: Worksheet,
        tournaments: list[Tournament],
        round_length_for_age_group: dict[str, int],
    ) -> None:
        sheet.append([
            "Dato",
            "Aldersgruppe",
            "Aktivitet",
            "Sted",
            "Start",
            "Slutt",
            "Vertsklubb",
            "Deltakende lag",
        ])
        for tournament in tournaments:
            start_time = tournament.start_time or ""
            end_time = ""
            if tournament.start_time:
                round_length = round_length_for_age_group.get(tournament.age_group)
                if round_length:
                    end_time = tournament.end_time(round_length) or ""
            activity = f"{tournament.age_group} Turnering — {tournament.arena}"
            if tournament.cancelled:
                activity = f"AVLYST: {activity}"
            sheet.append([
                tournament.date.strftime("%d.%m.%Y"),
                tournament.age_group,
                activity,
                tournament.arena,
                start_time,
                end_time,
                tournament.host_club or "",
                ", ".join(team.label for team in tournament.teams),
            ])

    def _write_hosting_sheet(
        self,
        sheet: Worksheet,
        hosted: list[Tournament],
        round_length_for_age_group: dict[str, int],
    ) -> None:
        sheet.append(["Dato", "Aldersgruppe", "Arena", "Start", "Slutt", "Lag", "Kamper"])
        if not hosted:
            sheet.append(["-", "Ingen hjemmeturneringer", "", "", "", "", ""])
            return
        for tournament in hosted:
            start_time = tournament.start_time or ""
            end_time = ""
            if tournament.start_time:
                round_length = round_length_for_age_group.get(tournament.age_group)
                if round_length:
                    end_time = tournament.end_time(round_length) or ""
            sheet.append([
                tournament.date.strftime("%d.%m.%Y"),
                tournament.age_group,
                tournament.arena,
                start_time,
                end_time,
                ", ".join(team.label for team in tournament.teams),
                len(tournament.games),
            ])

    def _write_travel_sheet(self, sheet: Worksheet, team_travel: dict[str, int]) -> None:
        sheet.append(["Lag", "Anslått kjøreavstand (km)"])
        if not team_travel:
            sheet.append(["-", 0])
            return
        for team_label, km in sorted(team_travel.items(), key=lambda item: (-item[1], item[0])):
            sheet.append([team_label, km])

    def _build_manifest(
        self,
        plan: SeasonPlan,
        club: str,
        packet_dir: Path,
        summary_workbook: Path,
        spond_path: Path,
        schedule_path: Path,
        club_tournaments: list[Tournament],
    ) -> dict[str, object]:
        hosted = [t for t in club_tournaments if t.host_club == club]
        return {
            "club": club,
            "packet_dir": str(packet_dir),
            "summary_workbook": summary_workbook.name,
            "spond_import_workbook": spond_path.name,
            "schedule_attachment": schedule_path.name,
            "response_template": "response_template.json",
            "tournament_count": len(club_tournaments),
            "hosted_count": len(hosted),
            "club_names": sorted({team.club for tournament in club_tournaments for team in tournament.teams}),
            "age_groups": sorted({tournament.age_group for tournament in club_tournaments}),
            "decisions": {
                "accept": "Godkjenn pakken som den er.",
                "change_request": "Fyll ut ønskede endringer i response_template.json.",
            },
        }

    @staticmethod
    def _response_template(club: str) -> dict[str, object]:
        return {
            "club": club,
            "decision": "accept",
            "comment": "",
            "requested_changes": {
                "lock_dates": [],
                "ban_dates": [],
                "pin_tournaments": [],
                "force_host_clubs": [],
                "exclude_host_clubs": [],
            },
        }

    def _club_tournaments(self, plan: SeasonPlan, club: str) -> list[Tournament]:
        return [
            tournament
            for tournament in sorted(plan.tournaments, key=lambda t: (t.date, t.age_group, t.id))
            if any(team.club == club for team in tournament.teams)
        ]

    @staticmethod
    def _slugify(value: str) -> str:
        slug = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_")
        return slug or "club"

    @staticmethod
    def _configure_sheet(sheet: Worksheet) -> None:
        sheet.sheet_view.showGridLines = False

    @staticmethod
    def _style_header_row(sheet: Worksheet, row_number: int = 1) -> None:
        for cell in sheet[row_number]:
            cell.font = cell.font.copy(bold=True)

    @staticmethod
    def _autosize_columns(sheet: Worksheet, max_width: int = 60) -> None:
        widths: dict[str, int] = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                col_letter = cell.column_letter
                width = len(str(cell.value))
                widths[col_letter] = max(widths.get(col_letter, 0), width)
        for col_letter, width in widths.items():
            sheet.column_dimensions[col_letter].width = min(width + 2, max_width)
