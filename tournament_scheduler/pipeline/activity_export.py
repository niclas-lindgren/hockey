"""Public activity/year-wheel workbook export helpers.

The RVV activity calendar is a public projection of one intentionally selected
worksheet/table. This module never exposes the workbook itself and never scans
internal sheets such as ``Innstillinger``, ``Lag`` or ``Kilder`` for content.
"""

from __future__ import annotations

import json
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .input_workbook import WorkbookInputError, _normalize_value

ACTIVITY_SHEET_NAMES: tuple[str, ...] = (
    "Aktiviteter",
    "Aktivitetsplan",
    "Årshjul",
    "Aarshjul",
    "Activities",
    "Year wheel",
)

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "date": ("date", "dato", "dag", "når", "nar", "when"),
    "month": ("month", "måned", "maned", "maaned", "mnd"),
    "title": ("title", "tittel", "aktivitet", "activity", "navn", "tema"),
    "type": ("type", "kategori", "aktivitetstype", "activity type"),
    "age_groups": ("age_groups", "age group", "aldersgruppe", "aldersgrupper", "gruppe", "målgruppe", "malgruppe"),
    "location": ("location", "sted", "arena", "lokasjon", "hall"),
    "description": ("description", "beskrivelse", "notat", "kommentar", "notes"),
    "url": ("url", "lenke", "link", "påmelding", "pamelding"),
}

_MONTHS: dict[str, int] = {
    "januar": 1,
    "jan": 1,
    "februar": 2,
    "feb": 2,
    "mars": 3,
    "mar": 3,
    "april": 4,
    "apr": 4,
    "mai": 5,
    "juni": 6,
    "jun": 6,
    "juli": 7,
    "jul": 7,
    "august": 8,
    "aug": 8,
    "september": 9,
    "sep": 9,
    "oktober": 10,
    "okt": 10,
    "october": 10,
    "oct": 10,
    "november": 11,
    "nov": 11,
    "desember": 12,
    "des": 12,
    "december": 12,
    "dec": 12,
}

_HELP_TEXT_RE = re.compile(r"\b(eksempel|example|hjelp|help|veiledning|forklaring)\b", re.IGNORECASE)
_AGE_GROUP_RE = re.compile(r"\bJ?U\d{1,2}\b", re.IGNORECASE)
_SPLIT_RE = re.compile(r"[,;/|]+|\s+og\s+", re.IGNORECASE)


def has_activity_table(path: str | Path) -> bool:
    """Return ``True`` when *path* contains a supported activity worksheet."""
    return _find_activity_sheet(path) is not None


def build_activities_payload(
    path: str | Path,
    *,
    default_year: int | None = None,
    generated_at: str | None = None,
) -> dict[str, Any] | None:
    """Read a public activity table and return the normalized JSON payload.

    Returns ``None`` when the workbook has no supported activity worksheet.
    Raises :class:`WorkbookInputError` when a supported sheet has malformed or
    incomplete activity rows.
    """
    found = _find_activity_sheet(path)
    if found is None:
        return None
    workbook_path, sheet_name = found
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl exception types vary
        raise WorkbookInputError(f"Kunne ikke lese Excel-filen '{workbook_path}': {exc}") from exc

    ws = wb[sheet_name]
    header_row, columns = _find_header(ws)
    if header_row is None:
        raise WorkbookInputError(
            f"{sheet_name}: fant ikke en aktivitetstabell med kolonner for dato/måned og aktivitet."
        )

    records = _read_activity_rows(ws, sheet_name=sheet_name, header_row=header_row, columns=columns, default_year=default_year)
    records.sort(key=lambda r: (r["date"], r["title"]))
    years = sorted({int(str(record["date"])[:4]) for record in records})
    payload_year = default_year if default_year is not None else (years[0] if years else None)
    return {
        "generated_at": generated_at or datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "year": payload_year,
        "activities": records,
    }


def write_activities_json(
    path: str | Path,
    export_dir: str | Path,
    *,
    default_year: int | None = None,
    generated_at: str | None = None,
) -> str | None:
    """Write ``activities.json`` in *export_dir* and return the path.

    Returns ``None`` when the workbook has no supported activity table.
    """
    payload = build_activities_payload(path, default_year=default_year, generated_at=generated_at)
    if payload is None:
        return None
    export_path = Path(export_dir)
    export_path.mkdir(parents=True, exist_ok=True)
    json_path = export_path / "activities.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return str(json_path)


def _find_activity_sheet(path: str | Path) -> tuple[Path, str] | None:
    workbook_path = Path(path)
    if not workbook_path.exists():
        return None
    try:
        wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl exception types vary
        raise WorkbookInputError(f"Kunne ikke lese Excel-filen '{workbook_path}': {exc}") from exc
    normalized_names = {_normalize_header(name): name for name in wb.sheetnames}
    for candidate in ACTIVITY_SHEET_NAMES:
        sheet = normalized_names.get(_normalize_header(candidate))
        if sheet:
            return workbook_path, sheet
    return None


def _find_header(ws: Worksheet) -> tuple[int | None, dict[str, int]]:
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        header_cells: list[tuple[str, int]] = []
        for col_index, cell in enumerate(row):
            header = _canonical_header(cell)
            if header:
                header_cells.append((header, col_index))
        columns = _choose_activity_table_columns(header_cells)
        if columns:
            return row_index, columns
    return None, {}


def _choose_activity_table_columns(header_cells: list[tuple[str, int]]) -> dict[str, int]:
    """Choose the most compact activity table from a header row.

    Some source workbooks keep helper data and the public activity table on the
    same sheet, producing duplicate headers such as ``Måned``/``Aktivitet`` in
    separate table blocks. Scoring around each title/activity header prevents a
    helper table's title column from being paired with the intended table's date
    column many cells away.
    """
    by_header: dict[str, list[int]] = {}
    for header, col_index in header_cells:
        by_header.setdefault(header, []).append(col_index)

    title_cols = by_header.get("title", [])
    if not title_cols:
        return {}
    date_or_month_cols = sorted(by_header.get("date", []) + by_header.get("month", []))
    if not date_or_month_cols:
        return {}

    def closest(candidates: list[int], anchor: int) -> int | None:
        if not candidates:
            return None
        return min(candidates, key=lambda col: (abs(col - anchor), col))

    best: tuple[int, int, dict[str, int]] | None = None
    for title_col in title_cols:
        date_col = closest(by_header.get("date", []), title_col)
        month_col = closest(by_header.get("month", []), title_col)
        nearest_required = closest(date_or_month_cols, title_col)
        if nearest_required is None:
            continue
        columns = {"title": title_col}
        if date_col is not None:
            columns["date"] = date_col
        if month_col is not None:
            columns["month"] = month_col
        for optional in ("type", "age_groups", "location", "description", "url"):
            col = closest(by_header.get(optional, []), title_col)
            if col is not None:
                columns[optional] = col
        # Compactness matters most; prefer the rightmost compact table when
        # scores tie because helper/summary tables commonly sit to the left.
        score = sum(abs(col - title_col) for key, col in columns.items() if key != "title")
        candidate = (score, -title_col, columns)
        if best is None or candidate < best:
            best = candidate
    return best[2] if best else {}


def _read_activity_rows(
    ws: Worksheet,
    *,
    sheet_name: str,
    header_row: int,
    columns: dict[str, int],
    default_year: int | None,
) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for row_index, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        values = {key: _cell(row, col_index) for key, col_index in columns.items()}
        if not any(value not in (None, "") for value in values.values()):
            continue
        if _is_help_or_example_row(values):
            continue

        title = _clean_text(values.get("title"))
        raw_date = values.get("date")
        raw_month = values.get("month")
        if not title:
            raise WorkbookInputError(f"{sheet_name}!rad {row_index}: mangler påkrevd tittel/aktivitet.")
        activity_date = _parse_activity_date(raw_date, raw_month, default_year=default_year, sheet_name=sheet_name, row_index=row_index)

        raw_type = _clean_text(values.get("type"))
        raw_age_groups = _clean_text(values.get("age_groups"))
        age_groups = _parse_age_groups(raw_age_groups or title)
        activity_type = raw_type or _infer_type(title, age_groups)

        records.append(
            {
                "date": activity_date.isoformat(),
                "type": activity_type,
                "age_groups": age_groups,
                "title": title,
                "location": _clean_text(values.get("location")),
                "description": _clean_text(values.get("description")),
                "url": _clean_text(values.get("url")),
            }
        )
    return records


def _cell(row: Iterable[Any], col_index: int) -> Any:
    values = list(row)
    if col_index >= len(values):
        return None
    return _normalize_value(values[col_index])


def _canonical_header(value: Any) -> str | None:
    normalized = _normalize_header(value)
    if not normalized:
        return None
    for canonical, aliases in _HEADER_ALIASES.items():
        if normalized in {_normalize_header(alias) for alias in aliases}:
            return canonical
    return None


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("å", "a").replace("æ", "ae").replace("ø", "o")
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def _clean_text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    text = str(value).strip()
    return text or None


def _is_help_or_example_row(values: dict[str, Any]) -> bool:
    # Limit this heuristic to display text columns. URLs such as
    # ``https://example.com/...`` are perfectly valid public activity links.
    text = " ".join(
        str(values.get(key))
        for key in ("title", "description")
        if values.get(key) not in (None, "")
    )
    return bool(_HELP_TEXT_RE.search(text))


def _parse_activity_date(
    raw_date: Any,
    raw_month: Any,
    *,
    default_year: int | None,
    sheet_name: str,
    row_index: int,
) -> date:
    if isinstance(raw_date, datetime):
        return raw_date.date()
    if isinstance(raw_date, date):
        return raw_date

    month_number = _parse_month(raw_month)
    if isinstance(raw_date, int) and month_number:
        return _date_from_parts(default_year, month_number, raw_date, raw_date, sheet_name, row_index)

    text = str(raw_date or "").strip()
    if not text:
        raise WorkbookInputError(f"{sheet_name}!rad {row_index}: mangler påkrevd dato.")

    # Full date forms first.
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass

    # Day number in the date column plus month in a separate column.
    if month_number and re.fullmatch(r"\d{1,2}", text):
        day = int(text)
        return _date_from_parts(default_year, month_number, day, text, sheet_name, row_index)

    # Norwegian text such as "17. januar" or "17 januar".
    match = re.fullmatch(r"(\d{1,2})\.?\s+([A-Za-zÆØÅæøå]+)", text)
    if match:
        month_from_text = _parse_month(match.group(2))
        if month_from_text:
            return _date_from_parts(default_year, month_from_text, int(match.group(1)), text, sheet_name, row_index)

    raise WorkbookInputError(f"{sheet_name}!rad {row_index}: ugyldig dato '{text}'. Bruk YYYY-MM-DD eller en gyldig dato.")


def _date_from_parts(
    default_year: int | None,
    month_number: int,
    day: int,
    raw_value: Any,
    sheet_name: str,
    row_index: int,
) -> date:
    if default_year is None:
        raise WorkbookInputError(
            f"{sheet_name}!rad {row_index}: dato '{raw_value}' mangler år. Oppgi YYYY-MM-DD eller konfigurer sesongår."
        )
    try:
        return date(default_year, month_number, day)
    except ValueError as exc:
        raise WorkbookInputError(f"{sheet_name}!rad {row_index}: ugyldig dato '{raw_value}'.") from exc


def _parse_month(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, int) and 1 <= value <= 12:
        return value
    text = str(value).strip().lower().replace("å", "a").replace("ø", "o").replace("æ", "ae")
    if text.isdigit() and 1 <= int(text) <= 12:
        return int(text)
    return _MONTHS.get(text)


def _parse_age_groups(value: str | None) -> list[str]:
    if not value:
        return []
    matches = [match.group(0).upper() for match in _AGE_GROUP_RE.finditer(value)]
    if matches:
        return list(dict.fromkeys(matches))
    parts = [part.strip().upper() for part in _SPLIT_RE.split(value) if part.strip()]
    return list(dict.fromkeys(part for part in parts if _AGE_GROUP_RE.fullmatch(part)))


def _infer_type(title: str, age_groups: list[str]) -> str | None:
    text = title
    for age_group in age_groups:
        text = re.sub(rf"\b{re.escape(age_group)}\b", "", text, flags=re.IGNORECASE)
    token = text.strip(" -–—:").split()
    return token[0].lower() if token else None
