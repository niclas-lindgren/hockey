"""Excel workbook input adapter for the RVV Miniputt pipeline.

The pipeline standard is now a human-editable Excel workbook. This module
converts the workbook sheets into the same raw config dict that Stage 1
validation and downstream stages already understand.
"""

from __future__ import annotations

import warnings
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from ..models import DatePreference


class WorkbookInputError(ValueError):
    """Raised when a workbook cannot be mapped to the pipeline input schema."""


REQUIRED_SHEETS = ("Innstillinger", "Lag")


def load_workbook_config(path: str | Path) -> dict[str, Any]:
    """Load the standard ``.xlsx`` pipeline input workbook as a config dict.

    Expected sheets:
    - ``Innstillinger``: columns ``felt`` and ``verdi`` for scalar settings.
      Tournament-participation target can be set as ``deltakelser_per_lag``
      (preferred) or the legacy ``target_tournament_count``.
      Optional ``vekt_cap`` (float, default 10.0) sets the absolute cap for
      preference-weight values; exceeded values trigger a :class:`UserWarning`.
    - ``Aldersgrupper``: columns ``age_group``, ``parallel_games``, optional
      ``round_length_minutes``, optional ``preferanse_vekt`` (float, default 0.0),
      optional ``deltakelser_per_lag`` / ``target_tournament_count`` and the
      split fields ``deltakelser_per_lag_før_jul`` /
      ``target_tournament_count_before_christmas`` plus
      ``deltakelser_per_lag_etter_jul`` /
      ``target_tournament_count_after_christmas``.
    - ``Lag``: columns ``club``, ``label``, ``age_group``.
    - ``Kilder``: columns ``name``, ``type``, ``url``.
    - ``Datopreferanser``: columns ``fra``, ``til``, ``vekt`` for global
      date-range scoring adjustments (positive = penalise, negative = reward).
    """
    workbook_path = Path(path)
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
    except Exception as exc:  # pragma: no cover - openpyxl gives varied exceptions
        raise WorkbookInputError(f"Kunne ikke lese Excel-filen '{workbook_path}': {exc}") from exc

    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in wb.sheetnames]
    if missing:
        raise WorkbookInputError(
            f"Excel-filen mangler påkrevde ark: {', '.join(missing)}."
        )

    raw: dict[str, Any] = {}
    raw.update(_read_settings(wb["Innstillinger"]))

    # Configurable absolute cap on vekt values (default 10.0).
    # Prevents runaway weights from dominating the scoring function.
    vekt_cap: float = float(raw.pop("vekt_cap", 10.0))

    if "Aldersgrupper" in wb.sheetnames:
        age_groups, parallel_games, round_lengths, pref_weights, target_counts = _read_age_groups(
            wb["Aldersgrupper"], vekt_cap=vekt_cap
        )
        if age_groups:
            raw["age_groups"] = age_groups
        if parallel_games:
            raw["parallel_games"] = parallel_games
        if round_lengths:
            raw["round_length_minutes"] = round_lengths
        if pref_weights:
            raw["preferanse_vekt"] = pref_weights
        if target_counts:
            raw["target_tournament_counts_by_age_group"] = target_counts

    raw["teams"] = _read_table(
        wb["Lag"],
        required_columns=("club", "label", "age_group"),
        optional_columns=("region", "skill_level", "target_tournament_count"),
    )

    if "Kilder" in wb.sheetnames:
        sources = _read_table(
            wb["Kilder"],
            required_columns=("name", "type", "url"),
        )
        sources = [s for s in sources if s.get("url")]
        if sources:
            raw["sources"] = sources

    if "Datopreferanser" in wb.sheetnames:
        date_prefs = _parse_datopreferanser(wb["Datopreferanser"], vekt_cap=vekt_cap)
        if date_prefs:
            raw["datopreferanser"] = date_prefs

    return raw


def _parse_datopreferanser(ws: Worksheet, *, vekt_cap: float = 10.0) -> list[DatePreference]:
    """Parse the ``Datopreferanser`` sheet into a list of :class:`DatePreference`.

    Expected columns: ``fra`` (date), ``til`` (date), ``vekt`` (float).
    Rows missing ``fra`` or ``til`` are skipped; ``vekt`` defaults to 0.0.
    Emits a :class:`UserWarning` when ``abs(vekt) > vekt_cap``.
    """
    prefs: list[DatePreference] = []
    for row in _rows_as_dicts(ws):
        fra_raw = row.get("fra")
        til_raw = row.get("til")
        if fra_raw in (None, "") or til_raw in (None, ""):
            continue
        # Normalise to date objects — openpyxl may return datetime or date-string
        fra_val = _to_date(fra_raw)
        til_val = _to_date(til_raw)
        if fra_val is None or til_val is None:
            continue
        vekt_raw = row.get("vekt")
        vekt = float(vekt_raw) if vekt_raw not in (None, "") else 0.0
        if abs(vekt) > vekt_cap:
            warnings.warn(
                f"Datopreferanser: vekt={vekt:.2f} for perioden {fra_val}–{til_val} "
                f"overskrider grensen ±{vekt_cap:.2f}. Verdien blir ikke automatisk begrenset her.",
                UserWarning,
                stacklevel=3,
            )
        prefs.append(DatePreference(fra=fra_val, til=til_val, vekt=vekt))
    return prefs


def _to_date(value: Any) -> date | None:
    """Convert a cell value to a :class:`datetime.date`, or return ``None``."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def _read_settings(ws: Worksheet) -> dict[str, Any]:
    rows = _rows_as_dicts(ws)
    result: dict[str, Any] = {}
    for row in rows:
        key = str(row.get("felt") or "").strip()
        if not key:
            continue
        value = row.get("verdi")
        if value is None or value == "":
            continue
        result[key] = _normalize_value(value)
    return result


def _read_age_groups(
    ws: Worksheet,
    *,
    vekt_cap: float = 10.0,
) -> tuple[list[str], dict[str, int], dict[str, int], dict[str, float], dict[str, dict[str, int]]]:
    age_groups: list[str] = []
    parallel_games: dict[str, int] = {}
    round_lengths: dict[str, int] = {}
    preferanse_vekt: dict[str, float] = {}
    target_counts: dict[str, dict[str, int]] = {}
    for row in _rows_as_dicts(ws):
        age_group = str(row.get("age_group") or "").strip()
        if not age_group:
            continue
        age_groups.append(age_group)
        if row.get("parallel_games") not in (None, ""):
            parallel_games[age_group] = int(row["parallel_games"])
        if row.get("round_length_minutes") not in (None, ""):
            round_lengths[age_group] = int(row["round_length_minutes"])
        if row.get("preferanse_vekt") not in (None, ""):
            vekt = float(row["preferanse_vekt"])
            if abs(vekt) > vekt_cap:
                warnings.warn(
                    f"Aldersgrupper: preferanse_vekt for '{age_group}' er {vekt:.2f}, "
                    f"som overskrider grensen ±{vekt_cap:.2f}. Verdien blir ikke automatisk begrenset her.",
                    UserWarning,
                    stacklevel=3,
                )
            preferanse_vekt[age_group] = vekt

        target_total = None
        for key in ("deltakelser_per_lag", "target_tournament_count"):
            if row.get(key) not in (None, ""):
                target_total = int(row[key])
                break
        target_before = None
        for key in ("deltakelser_per_lag_før_jul", "target_tournament_count_before_christmas"):
            if row.get(key) not in (None, ""):
                target_before = int(row[key])
                break
        target_after = None
        for key in ("deltakelser_per_lag_etter_jul", "target_tournament_count_after_christmas"):
            if row.get(key) not in (None, ""):
                target_after = int(row[key])
                break
        if target_total is not None or target_before is not None or target_after is not None:
            entry: dict[str, int] = {}
            if target_total is not None:
                entry["total"] = target_total
            if target_before is not None:
                entry["before_christmas"] = target_before
            if target_after is not None:
                entry["after_christmas"] = target_after
            target_counts[age_group] = entry

    return age_groups, parallel_games, round_lengths, preferanse_vekt, target_counts


def _read_table(
    ws: Worksheet,
    *,
    required_columns: Iterable[str],
    optional_columns: Iterable[str] = (),
) -> list[dict[str, Any]]:
    allowed = set(required_columns) | set(optional_columns)
    records: list[dict[str, Any]] = []
    for row in _rows_as_dicts(ws):
        if not any(value not in (None, "") for value in row.values()):
            continue
        record: dict[str, Any] = {}
        for column in allowed:
            value = row.get(column)
            if value not in (None, ""):
                record[column] = _normalize_value(value)
        records.append(record)
    return records


def _rows_as_dicts(ws: Worksheet) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if not any(headers):
        return []
    result: list[dict[str, Any]] = []
    for values in rows[1:]:
        row = {
            header: value
            for header, value in zip(headers, values, strict=False)
            if header
        }
        result.append(row)
    return result


def _normalize_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value
