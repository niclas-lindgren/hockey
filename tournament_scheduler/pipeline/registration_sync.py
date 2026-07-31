"""Sync registered-team CSV data into the ``Lag`` sheet of the season workbook.

This is intentionally a narrow write path: it only touches the ``Lag`` sheet,
preserving every other sheet and all existing workbook formatting.  The caller
is responsible for committing the updated workbook to the repository.
"""

from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

import openpyxl

from .input_workbook import WorkbookInputError
from .registered_teams import (
    PUBLIC_COLUMNS,
    RegisteredTeamsValidationError,
    build_registered_teams_payload,
)

_LAG_SHEET = "Lag"


class RegistrationSyncError(RuntimeError):
    """Raised when registration data cannot be synced to the workbook."""


def sync_registered_teams_to_workbook(
    csv_path: str | Path,
    workbook_path: str | Path,
    *,
    config_path: str | Path | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    """Validate *csv_path* and update only the ``Lag`` sheet of *workbook_path*.

    Returns a dictionary with ``changed`` (bool), ``team_count``, ``lag_rows_before``,
    ``lag_rows_after``, and ``dry_run``.  Raises :class:`RegistrationSyncError` on
    validation or I/O failures — the caller must not publish when this happens.
    """
    csv_path = Path(csv_path)
    workbook_path = Path(workbook_path)

    if not csv_path.exists():
        raise RegistrationSyncError(f"Registrerings-CSV finnes ikke: {csv_path}")

    # Validate the CSV through the existing pipeline — wrap its errors.
    try:
        public_payload, validation_report = build_registered_teams_payload(
            csv_path, config_path=config_path
        )
    except RegisteredTeamsValidationError as exc:
        raise RegistrationSyncError(str(exc)) from exc

    # Read existing Lag rows for comparison.
    lag_before = _read_lag_rows(workbook_path)

    # Build the target rows from the CSV.
    new_rows: list[dict[str, Any]] = []
    for group in public_payload.get("age_groups", []) or []:
        age_group = group.get("age_group", "")
        for club_entry in group.get("clubs", []) or []:
            club = club_entry.get("club", "")
            for team in club_entry.get("teams", []) or []:
                new_rows.append({"club": club, "label": team, "age_group": age_group})

    # Compare: are they semantically the same?
    lag_key = lambda r: (
        str(r.get("club", "")).strip().casefold(),
        str(r.get("label", "")).strip().casefold(),
        str(r.get("age_group", "")).strip().casefold(),
    )
    before_keys = {lag_key(r) for r in lag_before}
    after_keys = {lag_key(r) for r in new_rows}
    changed = before_keys != after_keys

    result: dict[str, Any] = {
        "changed": changed,
        "team_count": len(new_rows),
        "lag_rows_before": len(lag_before),
        "lag_rows_after": len(new_rows),
        "dry_run": dry_run,
        "validation_report": validation_report,
    }

    if dry_run or not changed:
        return result

    # Write the updated workbook.
    _write_lag_rows(workbook_path, new_rows)
    return result


def _read_lag_rows(workbook_path: Path) -> list[dict[str, Any]]:
    """Read existing rows from the ``Lag`` sheet (header excluded)."""
    if not workbook_path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:
        raise RegistrationSyncError(f"Kunne ikke lese arbeidsboken '{workbook_path}': {exc}") from exc
    if _LAG_SHEET not in wb.sheetnames:
        return []
    ws = wb[_LAG_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    columns = [h.casefold() for h in headers]
    records: list[dict[str, Any]] = []
    for values in rows[1:]:
        record: dict[str, Any] = {}
        for col_idx, header in enumerate(headers):
            if col_idx < len(values) and values[col_idx] not in (None, ""):
                record[header] = str(values[col_idx]).strip()
        if record:
            records.append(record)
    return records


def _write_lag_rows(workbook_path: Path, rows: list[dict[str, Any]]) -> None:
    """Write *rows* into the ``Lag`` sheet, preserving all other sheets."""
    if not workbook_path.exists():
        raise RegistrationSyncError(f"Arbeidsboken finnes ikke: {workbook_path}")

    # Work on a copy in case something goes wrong mid-write.
    tmp_path = workbook_path.with_suffix(".sync-tmp.xlsx")
    shutil.copy2(workbook_path, tmp_path)
    try:
        wb = openpyxl.load_workbook(tmp_path)
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        raise RegistrationSyncError(f"Kunne ikke lese arbeidsboken '{workbook_path}': {exc}") from exc

    if _LAG_SHEET not in wb.sheetnames:
        tmp_path.unlink(missing_ok=True)
        raise RegistrationSyncError(
            f"Arbeidsboken '{workbook_path}' mangler '{_LAG_SHEET}'-arket. "
            "Opprett arket med kolonnene club, label, age_group før synkronisering."
        )

    ws = wb[_LAG_SHEET]
    # Clear existing rows (keep header row 1, clear everything else).
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    headers = ["club", "label", "age_group"]
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))

    try:
        wb.save(tmp_path)
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        raise RegistrationSyncError(f"Kunne ikke lagre arbeidsboken: {exc}") from exc

    # Atomically replace the original.
    tmp_path.replace(workbook_path)
