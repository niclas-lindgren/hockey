"""Deterministic SharePoint registration import helpers.

The reviewed SharePoint List is the operational source for club/team
registrations, while ``input.xlsx`` remains the controlled pipeline snapshot.
This module validates a CSV/XLSX export from SharePoint and can replace only
``input.xlsx``'s public ``Lag`` sheet with the approved/current registrations.
"""

from __future__ import annotations

import csv
import hashlib
import json
import shutil
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping

import openpyxl

from .pipeline.input_workbook import WorkbookInputError, load_workbook_config


class RegistrationImportError(ValueError):
    """Raised when a reviewed registration export cannot be imported safely."""


_ACTIVE_STATUSES = {
    "approved",
    "current",
    "active",
    "accepted",
    "godkjent",
    "aktiv",
    "gjeldende",
}

_EXCLUDED_STATUSES = {
    "rejected",
    "withdrawn",
    "duplicate",
    "incomplete",
    "avvist",
    "trukket",
    "duplikat",
    "ufullstendig",
}

_COLUMN_ALIASES: Mapping[str, tuple[str, ...]] = {
    "sharepoint_id": ("sharepoint_id", "sharepoint id", "item_id", "item id", "id", "sp_id", "list_item_id"),
    "club": ("club", "klubb", "forening"),
    "label": ("team label", "team_label", "team", "team_name", "lag", "lagnavn", "label", "name"),
    "age_group": ("age group", "age_group", "aldergruppe", "klasse"),
    "status": ("status", "approval_state", "approval status", "godkjenningsstatus", "registreringsstatus"),
    "created": ("created", "opprettet", "created_at"),
    "modified": ("modified", "endret", "modified_at"),
    "contact": ("contact", "kontakt", "email", "e-post", "phone", "telefon"),
    "comment": ("comment", "comments", "kommentar", "notat"),
}

_REQUIRED_CANONICAL_COLUMNS = ("sharepoint_id", "club", "label", "age_group", "status")


@dataclass(frozen=True)
class RegistrationRow:
    """One row from the reviewed registration export."""

    sharepoint_id: str
    club: str
    label: str
    age_group: str
    status: str
    row_number: int
    created: str | None = None
    modified: str | None = None

    @property
    def identity(self) -> tuple[str, str, str]:
        return (self.club, self.label, self.age_group)

    def team_dict(self) -> dict[str, str]:
        return {"club": self.club, "label": self.label, "age_group": self.age_group}


@dataclass
class RegistrationRejectedRow:
    """A non-active row that was parsed but intentionally excluded."""

    row_number: int
    sharepoint_id: str
    status: str
    club: str = ""
    label: str = ""
    age_group: str = ""


@dataclass
class RegistrationDiff:
    added: list[dict[str, str]] = field(default_factory=list)
    removed: list[dict[str, str]] = field(default_factory=list)
    changed: list[dict[str, Any]] = field(default_factory=list)
    unchanged: list[dict[str, str]] = field(default_factory=list)
    rejected: list[RegistrationRejectedRow] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "added": self.added,
            "removed": self.removed,
            "changed": self.changed,
            "unchanged": self.unchanged,
            "rejected": [row.__dict__ for row in self.rejected],
        }


@dataclass
class RegistrationImportResult:
    """Validation/export result safe for CLI rendering and audit JSON."""

    source_path: str
    input_path: str
    output_path: str | None
    audit_path: str | None
    source_fingerprint: str
    active_rows: list[RegistrationRow]
    diff: RegistrationDiff
    dry_run: bool = False
    written: bool = False

    @property
    def summary_counts(self) -> dict[str, int]:
        return {
            "active": len(self.active_rows),
            "added": len(self.diff.added),
            "removed": len(self.diff.removed),
            "changed": len(self.diff.changed),
            "unchanged": len(self.diff.unchanged),
            "rejected": len(self.diff.rejected),
        }

    def to_audit_dict(self) -> dict[str, Any]:
        return {
            "schema_version": 1,
            "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "source_path": self.source_path,
            "input_path": self.input_path,
            "output_path": self.output_path,
            "source_fingerprint": self.source_fingerprint,
            "included_sharepoint_ids": [row.sharepoint_id for row in self.active_rows],
            "active_team_count": len(self.active_rows),
            "diff": self.diff.to_dict(),
        }


def validate_registrations(source_path: str | Path, *, input_path: str | Path) -> RegistrationImportResult:
    """Validate a reviewed SharePoint export without writing an output workbook."""

    return _build_result(source_path, input_path=input_path, output_path=None, dry_run=True)


def export_registrations(
    source_path: str | Path,
    *,
    input_path: str | Path,
    output_path: str | Path,
    dry_run: bool = False,
) -> RegistrationImportResult:
    """Create an updated input workbook, replacing only the ``Lag`` sheet.

    When ``dry_run`` is true the same validation/diff work is performed, but no
    workbook or audit file is created.
    """

    result = _build_result(source_path, input_path=input_path, output_path=output_path, dry_run=dry_run)
    if dry_run:
        return result

    input_workbook = Path(input_path)
    output_workbook = Path(output_path)
    output_workbook.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(input_workbook, output_workbook)
    _replace_lag_sheet(output_workbook, [row.team_dict() for row in result.active_rows])

    audit_path = _audit_path_for(output_workbook)
    audit_path.write_text(json.dumps(result.to_audit_dict(), indent=2, ensure_ascii=False), encoding="utf-8")
    result.audit_path = str(audit_path)
    result.written = True
    return result


def format_registration_summary(result: RegistrationImportResult) -> str:
    """Return a concise Norwegian plain-text summary for CLI output."""

    counts = result.summary_counts
    lines = [
        "Registreringer validert:",
        f"  Aktive lag: {counts['active']}",
        f"  Nye: {counts['added']}",
        f"  Fjernet/ikke lenger aktive: {counts['removed']}",
        f"  Endret: {counts['changed']}",
        f"  Uendret: {counts['unchanged']}",
        f"  Avvist/utelatt: {counts['rejected']}",
        f"  Kilde-fingerprint: {result.source_fingerprint}",
    ]
    if result.output_path:
        if result.dry_run:
            lines.append(f"  Dry-run: ville skrevet {result.output_path}")
        elif result.written:
            lines.append(f"  Skrev arbeidsbok: {result.output_path}")
            lines.append(f"  Audit: {result.audit_path}")
    return "\n".join(lines)


def _build_result(
    source_path: str | Path,
    *,
    input_path: str | Path,
    output_path: str | Path | None,
    dry_run: bool,
) -> RegistrationImportResult:
    source = Path(source_path)
    input_workbook = Path(input_path)
    if not source.exists():
        raise RegistrationImportError(f"Registreringseksporten finnes ikke: {source}")
    if not input_workbook.exists():
        raise RegistrationImportError(f"Input-arbeidsboken finnes ikke: {input_workbook}")

    try:
        config = load_workbook_config(input_workbook)
    except WorkbookInputError as exc:
        raise RegistrationImportError(str(exc)) from exc

    existing_teams = _normalize_existing_teams(config.get("teams") or [])
    allowed_clubs = {team["club"] for team in existing_teams}
    allowed_age_groups = _allowed_age_groups(config, existing_teams)

    parsed_rows, rejected = _read_registration_rows(source)
    active_rows = _validate_active_rows(parsed_rows, allowed_clubs=allowed_clubs, allowed_age_groups=allowed_age_groups)
    diff = _diff_teams(existing_teams, [row.team_dict() for row in active_rows], rejected)

    return RegistrationImportResult(
        source_path=str(source),
        input_path=str(input_workbook),
        output_path=str(output_path) if output_path is not None else None,
        audit_path=str(_audit_path_for(Path(output_path))) if output_path is not None and not dry_run else None,
        source_fingerprint=_fingerprint(source),
        active_rows=active_rows,
        diff=diff,
        dry_run=dry_run,
    )


def _read_registration_rows(path: Path) -> tuple[list[RegistrationRow], list[RegistrationRejectedRow]]:
    rows = _read_raw_rows(path)
    if not rows:
        raise RegistrationImportError(f"Registreringseksporten '{path}' inneholder ingen datarader.")

    header_map = _resolve_headers(rows[0].keys())
    missing = [column for column in _REQUIRED_CANONICAL_COLUMNS if column not in header_map]
    if missing:
        raise RegistrationImportError(
            "Registreringseksporten mangler påkrevde kolonner: "
            + ", ".join(missing)
            + ". Godkjente aliaser er SharePoint ID, club/klubb, team label/lag, age group/aldergruppe og status."
        )

    active_rows: list[RegistrationRow] = []
    rejected_rows: list[RegistrationRejectedRow] = []
    errors: list[str] = []

    for index, raw in enumerate(rows, start=2):
        canonical = {column: _clean(raw.get(header)) for column, header in header_map.items()}
        status_original = canonical.get("status", "")
        status = _normalize_status(status_original)
        sharepoint_id = canonical.get("sharepoint_id", "")
        club = canonical.get("club", "")
        label = canonical.get("label", "")
        age_group = canonical.get("age_group", "")

        if not status:
            errors.append(f"Rad {index}: mangler status.")
            continue
        if status not in _ACTIVE_STATUSES and status not in _EXCLUDED_STATUSES:
            errors.append(f"Rad {index}: ukjent status '{status_original}'.")
            continue
        if not sharepoint_id:
            errors.append(f"Rad {index}: mangler stabil SharePoint-ID.")
            continue

        if status in _EXCLUDED_STATUSES:
            rejected_rows.append(
                RegistrationRejectedRow(
                    row_number=index,
                    sharepoint_id=sharepoint_id,
                    status=status_original or status,
                    club=club,
                    label=label,
                    age_group=age_group,
                )
            )
            continue

        missing_fields = [name for name, value in (("club", club), ("label", label), ("age_group", age_group)) if not value]
        if missing_fields:
            errors.append(f"Rad {index}: aktiv registrering mangler {', '.join(missing_fields)}.")
            continue

        active_rows.append(
            RegistrationRow(
                sharepoint_id=sharepoint_id,
                club=club,
                label=label,
                age_group=age_group,
                status=status_original or status,
                row_number=index,
                created=canonical.get("created") or None,
                modified=canonical.get("modified") or None,
            )
        )

    if errors:
        raise RegistrationImportError("\n".join(errors))
    return active_rows, rejected_rows


def _read_raw_rows(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle)]
    if suffix in {".xlsx", ".xlsm"}:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        return [
            {header: value for header, value in zip(headers, values, strict=False) if header}
            for values in rows[1:]
            if any(value not in (None, "") for value in values)
        ]
    raise RegistrationImportError(f"Ustøttet registreringseksport-format '{path.suffix}'. Bruk CSV eller XLSX.")


def _resolve_headers(headers: Iterable[str]) -> dict[str, str]:
    normalized = {_normalize_header(header): header for header in headers if header}
    result: dict[str, str] = {}
    for canonical, aliases in _COLUMN_ALIASES.items():
        for alias in aliases:
            match = normalized.get(_normalize_header(alias))
            if match:
                result[canonical] = match
                break
    return result


def _validate_active_rows(
    rows: list[RegistrationRow],
    *,
    allowed_clubs: set[str],
    allowed_age_groups: set[str],
) -> list[RegistrationRow]:
    errors: list[str] = []
    by_id: dict[str, RegistrationRow] = {}
    by_identity: dict[tuple[str, str, str], RegistrationRow] = {}

    for row in rows:
        if allowed_clubs and row.club not in allowed_clubs:
            errors.append(f"Rad {row.row_number}: ukjent klubb '{row.club}'.")
        if allowed_age_groups and row.age_group not in allowed_age_groups:
            errors.append(f"Rad {row.row_number}: ukjent aldersgruppe '{row.age_group}'.")

        existing_id = by_id.get(row.sharepoint_id)
        if existing_id and existing_id.identity != row.identity:
            errors.append(
                f"Rad {row.row_number}: duplikat SharePoint-ID {row.sharepoint_id} brukes for flere ulike lag "
                f"({existing_id.identity} og {row.identity})."
            )
        elif existing_id:
            errors.append(f"Rad {row.row_number}: duplikat SharePoint-ID {row.sharepoint_id}.")
        else:
            by_id[row.sharepoint_id] = row

        existing_identity = by_identity.get(row.identity)
        if existing_identity and existing_identity.sharepoint_id != row.sharepoint_id:
            errors.append(
                f"Rad {row.row_number}: laget {row.club}/{row.label}/{row.age_group} finnes med flere SharePoint-ID-er "
                f"({existing_identity.sharepoint_id}, {row.sharepoint_id})."
            )
        elif existing_identity:
            errors.append(f"Rad {row.row_number}: duplikat lagregistrering {row.club}/{row.label}/{row.age_group}.")
        else:
            by_identity[row.identity] = row

    if errors:
        raise RegistrationImportError("\n".join(errors))
    return sorted(rows, key=lambda row: (row.age_group, row.club, row.label, row.sharepoint_id))


def _normalize_existing_teams(teams: Iterable[Mapping[str, Any]]) -> list[dict[str, str]]:
    normalized: list[dict[str, str]] = []
    for team in teams:
        club = _clean(team.get("club"))
        label = _clean(team.get("label"))
        age_group = _clean(team.get("age_group"))
        if club and label and age_group:
            normalized.append({"club": club, "label": label, "age_group": age_group})
    return sorted(normalized, key=lambda row: (row["age_group"], row["club"], row["label"]))


def _allowed_age_groups(config: Mapping[str, Any], teams: list[dict[str, str]]) -> set[str]:
    age_groups = {_clean(value) for value in config.get("age_groups") or [] if _clean(value)}
    if not age_groups:
        age_groups = {team["age_group"] for team in teams}
    return age_groups


def _diff_teams(
    existing: list[dict[str, str]],
    active: list[dict[str, str]],
    rejected: list[RegistrationRejectedRow],
) -> RegistrationDiff:
    existing_by_identity = {_team_identity(team): team for team in existing}
    active_by_identity = {_team_identity(team): team for team in active}

    added = [active_by_identity[key] for key in sorted(active_by_identity.keys() - existing_by_identity.keys())]
    removed = [existing_by_identity[key] for key in sorted(existing_by_identity.keys() - active_by_identity.keys())]
    unchanged = [active_by_identity[key] for key in sorted(active_by_identity.keys() & existing_by_identity.keys())]

    changed = _infer_changed_rows(removed, added)
    return RegistrationDiff(added=added, removed=removed, changed=changed, unchanged=unchanged, rejected=rejected)


def _infer_changed_rows(removed: list[dict[str, str]], added: list[dict[str, str]]) -> list[dict[str, Any]]:
    """Infer likely renames/age corrections for dry-run visibility.

    The controlled ``Lag`` sheet intentionally does not store SharePoint IDs, so
    stable ID based change tracking starts in the audit artifact. For the human
    diff we still flag likely changes when a removed and added row share two of
    three identity fields.
    """

    changes: list[dict[str, Any]] = []
    used_added: set[int] = set()
    for old in removed:
        old_values = (old["club"], old["label"], old["age_group"])
        for idx, new in enumerate(added):
            if idx in used_added:
                continue
            new_values = (new["club"], new["label"], new["age_group"])
            if sum(1 for a, b in zip(old_values, new_values, strict=False) if a == b) >= 2:
                changes.append({"from": old, "to": new})
                used_added.add(idx)
                break
    return changes


def _replace_lag_sheet(workbook_path: Path, teams: list[dict[str, str]]) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    sheet_index = wb.sheetnames.index("Lag") if "Lag" in wb.sheetnames else len(wb.sheetnames)
    if "Lag" in wb.sheetnames:
        del wb["Lag"]
    ws = wb.create_sheet("Lag", sheet_index)
    ws.append(["club", "label", "age_group"])
    for team in sorted(teams, key=lambda row: (row["age_group"], row["club"], row["label"])):
        ws.append([team["club"], team["label"], team["age_group"]])
    wb.save(workbook_path)


def _team_identity(team: Mapping[str, str]) -> tuple[str, str, str]:
    return (team["club"], team["label"], team["age_group"])


def _fingerprint(path: Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return f"sha256:{digest.hexdigest()}"


def _audit_path_for(output_workbook: Path) -> Path:
    return output_workbook.with_suffix(".registrations.audit.json")


def _normalize_header(value: str) -> str:
    return " ".join(str(value).strip().lower().replace("_", " ").split())


def _normalize_status(value: str) -> str:
    return _clean(value).lower().replace(" ", "_")


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()
