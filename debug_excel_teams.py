#!/usr/bin/env python3
"""Debug script to inspect Excel file structure for team events."""

import openpyxl
from datetime import datetime
from tournament_scheduler.utils.date_parser import DateParser

excel_file = "/Users/niclas/Downloads/U10_ETTER_JUL_REVIDERT__1_.xlsx"
date_parser = DateParser()

wb = openpyxl.load_workbook(excel_file, data_only=True)
ws = wb.active

# Track current tournament context
current_tournament_num = None
current_location = None
current_date = None

print("Scanning Excel file for tournament structure:\n")
print("=" * 80)

for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
    # Check each cell in the row for headers
    for cell_idx, cell in enumerate(row):
        if cell and isinstance(cell, str):
            cell_str = str(cell).strip()
            cell_lower = cell_str.lower()

            # Look for tournament number header
            if 'turnering nr' in cell_lower:
                current_tournament_num = cell_str
                current_location = None
                current_date = None
                print(f"\nRow {row_idx}: TOURNAMENT HEADER = {current_tournament_num}")
                # Show all cells in this row for debugging
                print(f"  Header row: {[str(c)[:30] if c else '' for c in row]}")

                # Show the next row (values row) too
                next_row_idx = row_idx
                for next_r in ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1, values_only=True):
                    print(f"  Values row: {[str(c)[:30] if c else '' for c in next_r]}")

            # Look for location (Sted) - check various formats
            if 'sted' in cell_lower:
                print(f"Row {row_idx}: Found 'sted' in cell: '{cell_str}'")
                # Next non-empty cell should be location
                for next_idx, next_cell in enumerate(row[cell_idx + 1:], start=cell_idx + 1):
                    if next_cell and str(next_cell).strip():
                        current_location = str(next_cell).strip()
                        print(f"Row {row_idx}: LOCATION = {current_location}")
                        break

    # Check for date in this row
    for cell in row:
        parsed = date_parser.parse_datetime_cell(cell)
        if parsed:
            old_date = current_date
            current_date = parsed.date()
            if current_date != old_date:
                print(f"Row {row_idx}: DATE = {current_date} (Location: {current_location or 'NOT SET'})")
            break

    # Check for team names if we have a date
    if current_date:
        for cell in row:
            if cell and isinstance(cell, str):
                cell_str = str(cell).strip()
                # Look for specific teams
                teams_to_check = ["Jar 6", "Skien", "Sandefjord", "Frisk Asker"]
                for team in teams_to_check:
                    if team.lower() in cell_str.lower():
                        print(f"Row {row_idx}: TEAM '{team}' found on {current_date} - Event: {current_location or current_tournament_num or 'UNKNOWN'}")

wb.close()

print("\n" + "=" * 80)
